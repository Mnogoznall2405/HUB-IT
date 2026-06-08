"""LibreOffice-based preview conversion for mail Office attachments."""
from __future__ import annotations

import hashlib
import logging
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_CONVERT_LOCK = threading.Lock()
_PREVIEW_CACHE: dict[str, tuple[float, "PreviewArtifact"]] = {}
_PREVIEW_CACHE_TTL_SEC = 600

_DEFAULT_WINDOWS_SOFFICE = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
_DEFAULT_LINUX_SOFFICE = Path("/usr/bin/soffice")


@dataclass(frozen=True)
class PreviewArtifact:
    pdf_bytes: bytes
    pdf_filename: str
    source_kind: str
    page_count: int
    sheets: list[dict[str, Any]]


class MailAttachmentPreviewError(RuntimeError):
    """Raised when Office preview conversion cannot be produced."""


def _positive_int_env(name: str, default: int) -> int:
    try:
        value = int(str(os.getenv(name, default) or default))
    except (TypeError, ValueError):
        return default
    return value if value > 0 else default


def _bool_env(name: str, default: bool) -> bool:
    raw = str(os.getenv(name, "1" if default else "0") or "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on"}


def _file_extension(filename: str) -> str:
    normalized = str(filename or "").strip().lower()
    dot_index = normalized.rfind(".")
    return normalized[dot_index + 1:] if dot_index >= 0 else ""


def resolve_soffice_path() -> Path:
    configured = str(os.getenv("LIBREOFFICE_SOFFICE_PATH", "") or os.getenv("SOFFICE_PATH", "") or "").strip()
    if configured:
        path = Path(configured)
        if path.is_file():
            return path
        raise MailAttachmentPreviewError(f"LibreOffice executable not found: {configured}")
    if _DEFAULT_WINDOWS_SOFFICE.is_file():
        return _DEFAULT_WINDOWS_SOFFICE
    if _DEFAULT_LINUX_SOFFICE.is_file():
        return _DEFAULT_LINUX_SOFFICE
    raise MailAttachmentPreviewError(
        "LibreOffice (soffice) is not configured. Set LIBREOFFICE_SOFFICE_PATH in .env."
    )


def classify_office_source(*, filename: str, content_type: str) -> str:
    extension = _file_extension(filename)
    normalized_type = str(content_type or "").strip().lower()
    if (
        "spreadsheetml" in normalized_type
        or "ms-excel" in normalized_type
        or "opendocument.spreadsheet" in normalized_type
        or extension in {"xls", "xlsx", "xlsm", "xlt", "xltx", "xltm", "ods"}
    ):
        return "excel"
    if (
        "wordprocessingml" in normalized_type
        or "msword" in normalized_type
        or "opendocument.text" in normalized_type
        or "rtf" in normalized_type
        or extension in {"doc", "docx", "docm", "dot", "dotx", "rtf", "odt"}
    ):
        return "word"
    return ""


def is_office_preview_enabled() -> bool:
    return _bool_env("MAIL_OFFICE_PREVIEW_ENABLED", True)


def office_preview_max_bytes() -> int:
    return _positive_int_env("MAIL_OFFICE_PREVIEW_MAX_BYTES", 25 * 1024 * 1024)


def office_preview_timeout_sec() -> int:
    return _positive_int_env("MAIL_OFFICE_PREVIEW_TIMEOUT_SEC", 120)


def _pdf_page_count(pdf_path: Path) -> int:
    from pypdf import PdfReader

    with pdf_path.open("rb") as handle:
        reader = PdfReader(handle)
        return max(0, len(reader.pages))


def _merge_pdf_paths(paths: list[Path], output_path: Path) -> None:
    from pypdf import PdfWriter

    writer = PdfWriter()
    for path in paths:
        if not path.is_file():
            continue
        with path.open("rb") as handle:
            writer.append(handle)
    with output_path.open("wb") as handle:
        writer.write(handle)


def _run_soffice_convert(*, soffice: Path, source_path: Path, output_dir: Path, timeout_sec: int) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    command = [
        str(soffice),
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--norestore",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_dir),
        str(source_path),
    ]
    with _CONVERT_LOCK:
        completed = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout_sec,
            check=False,
            text=True,
        )
    if completed.returncode != 0:
        stderr = (completed.stderr or completed.stdout or "").strip()
        raise MailAttachmentPreviewError(
            f"LibreOffice conversion failed (rc={completed.returncode}): {stderr[:500]}"
        )
    expected = output_dir / f"{source_path.stem}.pdf"
    if expected.is_file():
        return expected
    generated = sorted(output_dir.glob("*.pdf"), key=lambda item: item.stat().st_mtime, reverse=True)
    if generated:
        return generated[0]
    raise MailAttachmentPreviewError("LibreOffice conversion produced no PDF output.")


def _write_single_sheet_workbook(source_path: Path, *, sheet_index: int, target_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    source_wb = load_workbook(source_path, read_only=False, data_only=True)
    try:
        worksheets = list(source_wb.worksheets)
        if sheet_index < 0 or sheet_index >= len(worksheets):
            raise MailAttachmentPreviewError(f"Excel sheet index out of range: {sheet_index}")
        sheet = worksheets[sheet_index]
        target_wb = Workbook()
        target_ws = target_wb.active
        target_ws.title = sheet.title
        for row in sheet.iter_rows(values_only=True):
            target_ws.append(list(row))
        target_wb.save(target_path)
    finally:
        source_wb.close()


def _excel_sheet_metadata(*, source_path: Path, work_dir: Path, soffice: Path, timeout_sec: int) -> tuple[list[dict[str, Any]], bytes]:
    from openpyxl import load_workbook

    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheets = list(source_wb.worksheets)
    finally:
        source_wb.close()

    sheets_meta: list[dict[str, Any]] = []
    pdf_paths: list[Path] = []
    current_page = 1

    for index, worksheet in enumerate(worksheets):
        hidden = str(getattr(worksheet, "sheet_state", "visible") or "visible").lower() in {"hidden", "veryhidden"}
        if hidden:
            sheets_meta.append(
                {
                    "index": index,
                    "name": str(worksheet.title or f"Лист {index + 1}"),
                    "page": None,
                    "hidden": True,
                }
            )
            continue

        sheet_source = work_dir / f"sheet_{index}{source_path.suffix or '.xlsx'}"
        _write_single_sheet_workbook(source_path, sheet_index=index, target_path=sheet_source)
        sheet_pdf = _run_soffice_convert(
            soffice=soffice,
            source_path=sheet_source,
            output_dir=work_dir / f"sheet_{index}_pdf",
            timeout_sec=timeout_sec,
        )
        page_count = max(1, _pdf_page_count(sheet_pdf))
        page_end = current_page + page_count - 1
        sheets_meta.append(
            {
                "index": index,
                "name": str(worksheet.title or f"Лист {index + 1}"),
                "page": current_page,
                "page_end": page_end,
                "page_count": page_count,
                "hidden": False,
            }
        )
        pdf_paths.append(sheet_pdf)
        current_page += page_count

    if not pdf_paths:
        raise MailAttachmentPreviewError("Excel workbook has no visible sheets for preview.")

    merged_pdf = work_dir / "workbook.pdf"
    if len(pdf_paths) == 1:
        shutil.copyfile(pdf_paths[0], merged_pdf)
    else:
        _merge_pdf_paths(pdf_paths, merged_pdf)
    return sheets_meta, merged_pdf.read_bytes()


def _cache_key(*, filename: str, content: bytes) -> str:
    digest = hashlib.sha256(content).hexdigest()
    return f"{_file_extension(filename)}:{digest}"


def _cache_get(key: str) -> PreviewArtifact | None:
    entry = _PREVIEW_CACHE.get(key)
    if not entry:
        return None
    created_at, artifact = entry
    if time.monotonic() - created_at > _PREVIEW_CACHE_TTL_SEC:
        _PREVIEW_CACHE.pop(key, None)
        return None
    return artifact


def _cache_put(key: str, artifact: PreviewArtifact) -> None:
    _PREVIEW_CACHE[key] = (time.monotonic(), artifact)


def build_office_preview_artifact(*, filename: str, content_type: str, content: bytes) -> PreviewArtifact:
    if not is_office_preview_enabled():
        raise MailAttachmentPreviewError("Office attachment preview is disabled.")
    if not content:
        raise MailAttachmentPreviewError("Attachment is empty.")
    if len(content) > office_preview_max_bytes():
        raise MailAttachmentPreviewError("Attachment is too large for Office preview.")

    source_kind = classify_office_source(filename=filename, content_type=content_type)
    if not source_kind:
        raise MailAttachmentPreviewError("Attachment type is not supported for Office preview.")

    cache_key = _cache_key(filename=filename, content=content)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    soffice = resolve_soffice_path()
    timeout_sec = office_preview_timeout_sec()
    extension = _file_extension(filename) or ("xlsx" if source_kind == "excel" else "docx")
    stem = Path(str(filename or "attachment.bin")).stem or "attachment"
    pdf_filename = f"{stem}.pdf"

    with tempfile.TemporaryDirectory(prefix="itinvent-mail-preview-", ignore_cleanup_errors=True) as temp_dir:
        work_dir = Path(temp_dir)
        source_path = work_dir / f"source.{extension}"
        source_path.write_bytes(content)

        if source_kind == "excel" and extension in {"xlsx", "xlsm", "xltx", "xltm"}:
            try:
                sheets, pdf_bytes = _excel_sheet_metadata(
                    source_path=source_path,
                    work_dir=work_dir,
                    soffice=soffice,
                    timeout_sec=timeout_sec,
                )
            except Exception:
                logger.exception("Excel sheet preview failed; falling back to workbook PDF conversion")
                pdf_path = _run_soffice_convert(
                    soffice=soffice,
                    source_path=source_path,
                    output_dir=work_dir / "pdf",
                    timeout_sec=timeout_sec,
                )
                pdf_bytes = pdf_path.read_bytes()
                sheets = []
        else:
            pdf_path = _run_soffice_convert(
                soffice=soffice,
                source_path=source_path,
                output_dir=work_dir / "pdf",
                timeout_sec=timeout_sec,
            )
            pdf_bytes = pdf_path.read_bytes()
            sheets = []

        page_count = max(1, _pdf_page_count_from_bytes(pdf_bytes))
        artifact = PreviewArtifact(
            pdf_bytes=pdf_bytes,
            pdf_filename=pdf_filename,
            source_kind=source_kind,
            page_count=page_count,
            sheets=sheets,
        )
        _cache_put(cache_key, artifact)
        return artifact


def _pdf_page_count_from_bytes(pdf_bytes: bytes) -> int:
    from io import BytesIO
    from pypdf import PdfReader

    with PdfReader(BytesIO(pdf_bytes)) as reader:
        return len(reader.pages)


def build_preview_metadata(
    *,
    filename: str,
    content_type: str,
    artifact: PreviewArtifact,
    preview_pdf_path: str,
) -> dict[str, Any]:
    return {
        "preview_kind": "office_pdf",
        "source_kind": artifact.source_kind,
        "source_filename": str(filename or "attachment.bin"),
        "pdf_filename": artifact.pdf_filename,
        "page_count": int(artifact.page_count or 0),
        "sheets": list(artifact.sheets or []),
        "preview_url": str(preview_pdf_path or ""),
    }


def describe_preview_runtime() -> dict[str, Any]:
    enabled = is_office_preview_enabled()
    soffice = ""
    soffice_ready = False
    error = ""
    try:
        soffice = str(resolve_soffice_path())
        soffice_ready = True
    except MailAttachmentPreviewError as exc:
        error = str(exc)
    return {
        "enabled": enabled,
        "soffice_path": soffice,
        "soffice_ready": soffice_ready,
        "max_bytes": office_preview_max_bytes(),
        "timeout_sec": office_preview_timeout_sec(),
        "error": error,
        "python": sys.version.split()[0],
    }
