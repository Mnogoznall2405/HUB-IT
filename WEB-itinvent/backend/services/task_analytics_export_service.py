"""Excel export helpers for hub task analytics."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.services.hub_service import hub_service


_DATE_BASIS_LABELS = {
    "protocol_date": "По дате протокола",
    "completed_at": "По завершению",
    "due_at": "По сроку",
}

_METRIC_LABELS = (
    ("Всего задач", "total"),
    ("Открыто", "open"),
    ("Новые", "new"),
    ("В работе", "in_progress"),
    ("На проверке", "review"),
    ("Выполнено", "done"),
    ("В срок", "done_on_time"),
    ("Выполнено без срока", "done_without_due"),
    ("Просрочено", "overdue"),
    ("Со сроком", "with_due_total"),
    ("Выполнено, %", "completion_percent"),
    ("В срок, %", "completion_on_time_percent"),
)

_ANALYTICS_TABLE_COLUMNS = (
    ("Всего", "total"),
    ("Открыто", "open"),
    ("Новые", "new"),
    ("В работе", "in_progress"),
    ("На проверке", "review"),
    ("Выполнено", "done"),
    ("В срок", "done_on_time"),
    ("Без срока", "done_without_due"),
    ("Просрочено", "overdue"),
    ("Выполнено, %", "completion_percent"),
    ("В срок, %", "completion_on_time_percent"),
)


def _normalize_text(value: Any, default: str = "") -> str:
    text = str(value or "").strip()
    return text or default


def _fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                value = str(cell.value or "")
            except Exception:
                value = ""
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 48)


def _append_section_title(worksheet, title: str) -> None:
    worksheet.append([title])
    worksheet.cell(worksheet.max_row, 1).font = Font(bold=True, size=12)


def _append_table(worksheet, headers: list[str], rows: Iterable[list[Any]], *, empty_message: str = "Нет данных") -> None:
    worksheet.append(headers)
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)

    has_rows = False
    for row in rows:
        has_rows = True
        worksheet.append(list(row))

    if not has_rows:
        worksheet.append([empty_message])


def _format_filter_labels(
    *,
    hub_service_impl,
    project_ids: list[str] | None,
    object_ids: list[str] | None,
    participant_user_ids: list[int] | None,
) -> dict[str, str]:
    normalized_project_ids = [str(item).strip() for item in (project_ids or []) if str(item).strip()]
    normalized_object_ids = [str(item).strip() for item in (object_ids or []) if str(item).strip()]
    normalized_participant_ids = [int(item) for item in (participant_user_ids or []) if int(item) > 0]

    projects_by_id = {
        str(item.get("id")): _normalize_text(item.get("name")) or _normalize_text(item.get("code")) or str(item.get("id"))
        for item in hub_service_impl.list_task_projects(include_inactive=True)
    }
    objects_by_id = {
        str(item.get("id")): _normalize_text(item.get("name")) or _normalize_text(item.get("code")) or str(item.get("id"))
        for item in hub_service_impl.list_task_objects(include_inactive=True)
    }
    users_by_id = {
        int(item.get("id")): _normalize_text(item.get("full_name")) or _normalize_text(item.get("username")) or str(item.get("id"))
        for item in hub_service_impl.list_assignees()
        if int(item.get("id") or 0) > 0
    }

    project_label = ", ".join(projects_by_id.get(item, item) for item in normalized_project_ids) or "Все проекты"
    object_label = ", ".join(objects_by_id.get(item, item) for item in normalized_object_ids) or "Все объекты"
    participant_label = ", ".join(users_by_id.get(item, str(item)) for item in normalized_participant_ids) or "Все участники"

    return {
        "projects": project_label,
        "objects": object_label,
        "participants": participant_label,
    }


def _table_rows(items: list[dict[str, Any]], *, label_key: str) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for item in items or []:
        row = [_normalize_text(item.get(label_key)) or "—"]
        for _, key in _ANALYTICS_TABLE_COLUMNS:
            row.append(item.get(key, 0))
        rows.append(row)
    return rows


def build_task_analytics_excel(
    *,
    hub_service_impl=hub_service,
    start_date: str | None = None,
    end_date: str | None = None,
    date_basis: str = "protocol_date",
    project_ids: list[str] | None = None,
    object_ids: list[str] | None = None,
    participant_user_ids: list[int] | None = None,
) -> tuple[bytes, str]:
    analytics = hub_service_impl.get_task_analytics(
        start_date=start_date,
        end_date=end_date,
        date_basis=date_basis,
        project_ids=project_ids,
        object_ids=object_ids,
        participant_user_ids=participant_user_ids,
    )

    summary = analytics.get("summary") or {}
    filters = analytics.get("filters") or {}
    filter_labels = _format_filter_labels(
        hub_service_impl=hub_service_impl,
        project_ids=filters.get("project_ids") or project_ids,
        object_ids=filters.get("object_ids") or object_ids,
        participant_user_ids=filters.get("participant_user_ids") or participant_user_ids,
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet["A1"] = "Аналитика задач"
    summary_sheet["A1"].font = Font(bold=True, size=14)

    summary_sheet.append([])
    _append_section_title(summary_sheet, "Фильтры отчёта")
    _append_table(
        summary_sheet,
        ["Фильтр", "Значение"],
        [
            ["Период с", _normalize_text(filters.get("start_date")) or "Без ограничения"],
            ["Период по", _normalize_text(filters.get("end_date")) or "Без ограничения"],
            ["База дат", _DATE_BASIS_LABELS.get(_normalize_text(filters.get("date_basis"), "protocol_date"), "По дате протокола")],
            ["Проекты", filter_labels["projects"]],
            ["Объекты", filter_labels["objects"]],
            ["Участники", filter_labels["participants"]],
        ],
    )

    summary_sheet.append([])
    _append_section_title(summary_sheet, "Ключевые показатели")
    _append_table(
        summary_sheet,
        ["Показатель", "Значение"],
        [[label, summary.get(key, 0)] for label, key in _METRIC_LABELS],
    )

    summary_sheet.append([])
    _append_section_title(summary_sheet, "Статусы")
    _append_table(
        summary_sheet,
        ["Статус", "Количество"],
        [
            [_normalize_text(item.get("label")) or _normalize_text(item.get("status")) or "—", item.get("value", 0)]
            for item in (analytics.get("status_breakdown") or [])
        ],
    )

    participant_sheet = workbook.create_sheet(title="По участникам")
    _append_table(
        participant_sheet,
        ["Участник", *[label for label, _ in _ANALYTICS_TABLE_COLUMNS]],
        _table_rows(analytics.get("by_participant") or [], label_key="participant_name"),
    )

    project_sheet = workbook.create_sheet(title="По проектам")
    _append_table(
        project_sheet,
        ["Проект", *[label for label, _ in _ANALYTICS_TABLE_COLUMNS]],
        _table_rows(analytics.get("by_project") or [], label_key="project_name"),
    )

    object_sheet = workbook.create_sheet(title="По объектам")
    _append_table(
        object_sheet,
        ["Объект", *[label for label, _ in _ANALYTICS_TABLE_COLUMNS]],
        _table_rows(analytics.get("by_object") or [], label_key="object_name"),
    )

    trend_sheet = workbook.create_sheet(title="Тренд")
    _append_table(
        trend_sheet,
        ["Период", "Создано по протоколу", "Выполнено", "Выполнено в срок"],
        [
            [
                _normalize_text(item.get("bucket_label")) or _normalize_text(item.get("bucket_key")) or "—",
                item.get("created", 0),
                item.get("completed", 0),
                item.get("completed_on_time", 0),
            ]
            for item in ((analytics.get("trend") or {}).get("items") or [])
        ],
    )

    for worksheet in workbook.worksheets:
        _fit_columns(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"task_analytics_{timestamp}.xlsx"
    return output.read(), filename
