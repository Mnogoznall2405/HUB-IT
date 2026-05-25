"""Service for importing ticket data from Excel files.

Provides full workbook reading (read_only=False, data_only=False) to preserve
formulas, colors, comments, hyperlinks, and hidden sheet information.

Requirements: 1.1, 1.2, 1.3, 1.4, 1.5, 1.6, 1.7, 1.8, 1.9, 1.10, 1.11
"""
from __future__ import annotations

import json
import logging
import math
import re
import uuid as uuid_mod
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.appdb.db import app_session
from backend.appdb.tickets_models import (
    TicketEmployee,
    TicketFinancialOp,
    TicketImportJob,
    TicketImportRawTrace,
    TicketObject,
    TicketRequest,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

UPLOAD_DIR = "uploads/tickets/import"
MAX_IMPORT_SCAN_COLUMNS = 256

# Default reference color map: hex (uppercase, no #) → status
DEFAULT_REFERENCE_COLORS: dict[str, str] = {
    "FFFFFF": "new",
    "C6EFCE": "purchased",
    "FFFF00": "in_progress",
    "FF0000": "cancelled",
    "FFC000": "exchange_needed",
    "BDD7EE": "data_check",
    "D9D9D9": "closed",
}

VALID_REQUEST_STATUSES = {
    "new",
    "data_check",
    "missing_data",
    "ready_to_buy",
    "in_progress",
    "purchased",
    "exchange_needed",
    "refund",
    "cancelled",
    "no_show",
    "closed",
    "archive",
}

# Standard column schema for headerless sheets
STANDARD_COLUMN_SCHEMA: list[str] = [
    "full_name",
    "submitted_at",
    "departure_date",
    "route",
    "cost",
    "status_note",
    "assignee",
]

# Excel epoch for serial date conversion (1900-01-01 minus 2 days for Excel bug)
_EXCEL_EPOCH = datetime(1899, 12, 30, tzinfo=timezone.utc)

# Header aliases for mapping Excel headers to field names
_HEADER_ALIASES: dict[str, str] = {
    "фио": "full_name",
    "ф.и.о.": "full_name",
    "ф.и.о": "full_name",
    "имя": "full_name",
    "сотрудник": "full_name",
    "дата подачи": "submitted_at",
    "дата заявки": "submitted_at",
    "дата вылета": "departure_date",
    "вылет": "departure_date",
    "дата прибытия": "arrival_date",
    "прибытие": "arrival_date",
    "маршрут": "route",
    "направление": "route",
    "стоимость": "cost",
    "цена": "cost",
    "сумма": "cost",
    "статус": "status_note",
    "ответственный": "assignee",
    "логист": "assignee",
    "телефон": "phone",
    "тел": "phone",
    "паспорт": "passport",
    "паспортные данные": "passport",
}


# ---------------------------------------------------------------------------
# Data Transfer Objects
# ---------------------------------------------------------------------------


@dataclass
class SheetClassification:
    """Classification result for a single worksheet."""
    title: str
    classification: str  # рабочий, пустой, служебный, скрытый, ПОТЕРИ
    row_count: int = 0


@dataclass
class HiddenSheetInfo:
    """Visibility information for a single worksheet."""
    title: str
    state: str  # visible, hidden, veryHidden


@dataclass
class SheetPreviewInfo:
    """Preview information for a single worksheet."""
    title: str
    classification: str
    row_count: int
    headers: list[str]
    matched_object_id: int | None = None
    matched_object_name: str | None = None
    error_count: int = 0


@dataclass
class ImportPreview:
    """Full preview result for an import job."""
    job_id: str
    file_name: str
    sheets: list[SheetPreviewInfo] = field(default_factory=list)
    color_map: dict[str, str] = field(default_factory=dict)
    unmatched_sheets: list[str] = field(default_factory=list)
    unique_colors: list[str] = field(default_factory=list)


@dataclass
class ParsedRow:
    """Successfully parsed row from Excel."""
    full_name: str
    submitted_at: datetime | None = None
    departure_date: datetime | None = None
    arrival_date: datetime | None = None
    route: str | None = None
    cost: Decimal = field(default_factory=lambda: Decimal("0.00"))
    status: str | None = None
    assignee: str | None = None
    phone: str | None = None
    passport: str | None = None
    is_urgent: bool = False
    needs_review: bool = False
    review_fields: list[str] = field(default_factory=list)
    row_number: int = 0
    raw_cells: dict[str, Any] = field(default_factory=dict)
    cell_colors: dict[str, str] = field(default_factory=dict)
    cell_formulas: dict[str, str] = field(default_factory=dict)
    cell_comments: dict[str, str] = field(default_factory=dict)
    cell_hyperlinks: dict[str, str] = field(default_factory=dict)
    cell_addresses: dict[str, str] = field(default_factory=dict)


@dataclass
class ParseError:
    """Error encountered while parsing a row."""
    row_number: int
    error: str
    severity: str = "error"  # error or warning
    raw_cells: dict[str, Any] = field(default_factory=dict)
    cell_colors: dict[str, str] = field(default_factory=dict)
    cell_formulas: dict[str, str] = field(default_factory=dict)
    cell_comments: dict[str, str] = field(default_factory=dict)
    cell_hyperlinks: dict[str, str] = field(default_factory=dict)
    cell_addresses: dict[str, str] = field(default_factory=dict)


@dataclass
class DuplicateMatch:
    """A detected duplicate row."""
    parsed_row: ParsedRow
    existing_id: int
    strategy: str = "skip"  # skip, update, create


@dataclass
class FinancialOpRow:
    """Parsed row from ПОТЕРИ sheet."""
    full_name: str
    amount: Decimal = field(default_factory=lambda: Decimal("0.00"))
    reason: str | None = None
    op_date: datetime | None = None
    op_type: str = "loss"  # refund, exchange, loss
    row_number: int = 0


@dataclass
class ImportSettings:
    """Settings for executing an import."""
    color_map: dict[str, str] = field(default_factory=dict)
    duplicate_strategy: str = "skip"  # skip, update, create
    sheet_object_map: dict[str, int] = field(default_factory=dict)


@dataclass
class ImportResult:
    """Result of an import execution."""
    imported: int = 0
    skipped: int = 0
    errors: int = 0
    warnings: int = 0
    error_details: list[dict[str, Any]] = field(default_factory=list)
    warning_details: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ReconciliationReport:
    """Reconciliation report after import."""
    total_rows_in_file: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    rows_with_errors: int = 0
    rows_with_warnings: int = 0
    total_losses_amount: Decimal = field(default_factory=lambda: Decimal("0.00"))
    total_cost_imported: Decimal = field(default_factory=lambda: Decimal("0.00"))
    sheets_processed: int = 0
    sheets_skipped: int = 0
    hidden_sheets_found: int = 0


# ---------------------------------------------------------------------------
# TicketsImportService
# ---------------------------------------------------------------------------


class TicketsImportService:
    """Service for importing ticket data from Excel with full reading
    (colors, comments, hyperlinks, hidden sheets).
    """

    def __init__(self, upload_base_dir: str | None = None) -> None:
        _patch_openpyxl_timezone_datetimes()
        self._upload_base_dir = upload_base_dir or UPLOAD_DIR

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def upload_file(
        self,
        file_name: str,
        file_content: bytes,
        user_id: int,
    ) -> dict[str, Any]:
        """Save uploaded Excel file and create an import job record."""
        job_id = uuid_mod.uuid4().hex

        upload_dir = Path(self._upload_base_dir)
        upload_dir.mkdir(parents=True, exist_ok=True)

        stored_filename = f"{job_id}.xlsx"
        file_path = upload_dir / stored_filename
        file_path.write_bytes(file_content)

        with app_session() as session:
            job = TicketImportJob(
                id=job_id,
                file_name=file_name,
                file_path=str(file_path),
                status="uploaded",
                user_id=user_id,
            )
            session.add(job)

        logger.info(
            "Import file uploaded: job_id=%s, file_name=%s, user_id=%d",
            job_id, file_name, user_id,
        )
        return {"id": job_id, "file_name": file_name, "status": "uploaded"}

    def open_workbook(self, path: str) -> Workbook:
        """Open workbook in full reading mode."""
        return openpyxl.load_workbook(path, read_only=False, data_only=False)

    def get_preview(self, job_id: str) -> ImportPreview:
        """Generate a preview for an import job."""
        with app_session() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            if job is None:
                raise ValueError(f"Import job not found: {job_id}")
            if job.status not in ("uploaded", "previewed"):
                raise ValueError(
                    f"Import job {job_id} is in status '{job.status}', "
                    f"expected 'uploaded' or 'previewed'"
                )
            file_path = job.file_path
            file_name = job.file_name

            objects = session.query(TicketObject).filter_by(is_active=True).all()
            object_lookup: dict[str, int] = {}
            object_names: dict[int, str] = {}
            for obj in objects:
                object_lookup[obj.name.lower()] = obj.id
                object_names[obj.id] = obj.name
                if obj.code:
                    object_lookup[obj.code.lower()] = obj.id
                if obj.short_name:
                    object_lookup[obj.short_name.lower()] = obj.id

        wb = self.open_workbook(file_path)
        classifications = self._classify_sheets(wb)

        sheets_preview: list[SheetPreviewInfo] = []
        unmatched_sheets: list[str] = []
        all_unique_colors: set[str] = set()

        for cls in classifications:
            ws = wb[cls.title]
            headers = self._extract_headers(ws)
            matched_id: int | None = None
            matched_name: str | None = None
            title_lower = cls.title.lower().strip()

            if title_lower in object_lookup:
                matched_id = object_lookup[title_lower]
                matched_name = object_names.get(matched_id)

            if cls.classification == "рабочий" and matched_id is None:
                unmatched_sheets.append(cls.title)

            if cls.classification == "рабочий":
                sheet_colors = self._extract_unique_colors(ws)
                all_unique_colors.update(sheet_colors)

            sheets_preview.append(SheetPreviewInfo(
                title=cls.title,
                classification=cls.classification,
                row_count=cls.row_count,
                headers=headers,
                matched_object_id=matched_id,
                matched_object_name=matched_name,
            ))

        unique_colors_list = sorted(all_unique_colors)
        suggested_color_map = self._suggest_color_mapping(
            unique_colors_list, DEFAULT_REFERENCE_COLORS
        )

        preview = ImportPreview(
            job_id=job_id,
            file_name=file_name,
            sheets=sheets_preview,
            color_map=suggested_color_map,
            unmatched_sheets=unmatched_sheets,
            unique_colors=unique_colors_list,
        )

        with app_session() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            if job is not None:
                job.status = "previewed"
                job.preview_json = json.dumps(
                    {
                        "sheets": [
                            {
                                "title": s.title,
                                "classification": s.classification,
                                "row_count": s.row_count,
                                "headers": s.headers,
                                "matched_object_id": s.matched_object_id,
                                "matched_object_name": s.matched_object_name,
                            }
                            for s in sheets_preview
                        ],
                        "color_map": suggested_color_map,
                        "unmatched_sheets": unmatched_sheets,
                        "unique_colors": unique_colors_list,
                    },
                    ensure_ascii=False,
                )

        wb.close()
        return preview

    def execute_import(
        self, job_id: str, settings: ImportSettings, user_id: int
    ) -> ImportResult:
        """Execute the import pipeline for a previewed job.

        Args:
            job_id: The import job ID (must be in 'previewed' status).
            settings: Import settings (color_map, duplicate_strategy, sheet_object_map).
            user_id: ID of the user executing the import.

        Returns:
            ImportResult with counts of imported, skipped, errors, warnings.
        """
        # Load job
        with app_session() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            if job is None:
                raise ValueError(f"Import job not found: {job_id}")
            if job.status not in ("previewed",):
                raise ValueError(
                    f"Import job {job_id} is in status '{job.status}', "
                    f"expected 'previewed'"
                )
            file_path = job.file_path
            file_name = job.file_name
            job.status = "running"

        result = ImportResult()
        total_rows_in_file = 0
        total_cost_imported = Decimal("0.00")
        total_losses_amount = Decimal("0.00")
        sheets_processed = 0
        sheets_skipped = 0
        hidden_sheets_found = 0

        try:
            wb = self.open_workbook(file_path)
            classifications = self._classify_sheets(wb)
            hidden_sheets_found = sum(
                1 for c in classifications if c.classification == "скрытый"
            )

            # Load object lookup for matching
            with app_session() as session:
                objects = session.query(TicketObject).filter_by(is_active=True).all()
                object_lookup: dict[str, int] = {}
                for obj in objects:
                    object_lookup[obj.name.lower()] = obj.id
                    if obj.code:
                        object_lookup[obj.code.lower()] = obj.id
                    if obj.short_name:
                        object_lookup[obj.short_name.lower()] = obj.id

            # Extract worksheet-level metadata
            ws_comments_cache: dict[str, dict[str, str]] = {}
            ws_hyperlinks_cache: dict[str, dict[str, list[str]]] = {}

            for cls in classifications:
                ws = wb[cls.title]
                ws_comments_cache[cls.title] = self._extract_cell_comments(ws)
                ws_hyperlinks_cache[cls.title] = self._extract_cell_hyperlinks(ws)

            # Process each sheet
            for cls in classifications:
                if cls.classification in ("пустой", "служебный", "скрытый"):
                    sheets_skipped += 1
                    continue

                ws = wb[cls.title]
                sheet_visibility = getattr(ws, "sheet_state", "visible") or "visible"

                # Handle ПОТЕРИ sheet separately
                if cls.classification == "ПОТЕРИ":
                    sheets_processed += 1
                    fin_ops = self._parse_losses_sheet(ws)
                    self._store_losses(
                        fin_ops, job_id, file_name, cls.title,
                        sheet_visibility, ws_comments_cache.get(cls.title, {}),
                        ws_hyperlinks_cache.get(cls.title, {}),
                    )
                    for op in fin_ops:
                        total_losses_amount += op.amount
                    total_rows_in_file += cls.row_count
                    continue

                # Working sheet
                sheets_processed += 1
                total_rows_in_file += cls.row_count

                # Determine object_id for this sheet
                object_id: int | None = settings.sheet_object_map.get(cls.title)
                if object_id is None:
                    title_lower = cls.title.lower().strip()
                    object_id = object_lookup.get(title_lower)
                if object_id is None:
                    # Cannot import without object mapping
                    result.errors += cls.row_count - 1  # subtract header
                    result.error_details.append({
                        "sheet": cls.title,
                        "error": "No object mapping for sheet",
                    })
                    continue

                # Determine if sheet has headers
                headers = self._extract_headers(ws)
                has_headers = self._has_valid_headers(headers)

                # Parse rows
                parsed_rows: list[ParsedRow] = []
                parse_errors: list[ParseError] = []
                comments = ws_comments_cache.get(cls.title, {})
                hyperlinks = ws_hyperlinks_cache.get(cls.title, {})

                start_row = 2 if has_headers else 1
                col_order = self._build_header_map(headers) if has_headers else None

                for row_idx, row in self._iter_data_rows(ws, min_row=start_row):
                    if has_headers and col_order is not None:
                        row_result = self._parse_row(
                            row, col_order, row_idx, comments,
                            hyperlinks, settings.color_map,
                        )
                    else:
                        row_result = self._parse_row_standard_schema(
                            row, STANDARD_COLUMN_SCHEMA, row_idx,
                            comments, hyperlinks, settings.color_map,
                        )

                    if isinstance(row_result, ParseError):
                        parse_errors.append(row_result)
                    else:
                        parsed_rows.append(row_result)

                # Detect duplicates
                duplicates = self._detect_duplicates(parsed_rows, object_id)
                duplicate_row_numbers = {d.parsed_row.row_number for d in duplicates}

                # Process parsed rows — create records
                for parsed in parsed_rows:
                    if parsed.row_number in duplicate_row_numbers:
                        dup = next(
                            d for d in duplicates
                            if d.parsed_row.row_number == parsed.row_number
                        )
                        dup.strategy = settings.duplicate_strategy
                        if settings.duplicate_strategy == "skip":
                            result.skipped += 1
                            self._store_raw_trace(
                                parsed, job_id, file_name, cls.title,
                                sheet_visibility, None,
                            )
                            continue
                        elif settings.duplicate_strategy == "update":
                            request_id = self._update_existing_request(
                                dup.existing_id, parsed, object_id
                            )
                            self._store_raw_trace(
                                parsed, job_id, file_name, cls.title,
                                sheet_visibility, request_id,
                            )
                            result.imported += 1
                            total_cost_imported += parsed.cost
                            continue

                    # Create new request
                    request_id = self._create_request(parsed, object_id)
                    self._store_raw_trace(
                        parsed, job_id, file_name, cls.title,
                        sheet_visibility, request_id,
                    )
                    result.imported += 1
                    total_cost_imported += parsed.cost

                # Record errors
                for err in parse_errors:
                    if err.severity == "error":
                        result.errors += 1
                        result.error_details.append({
                            "sheet": cls.title,
                            "row": err.row_number,
                            "error": err.error,
                        })
                    else:
                        result.warnings += 1
                        result.warning_details.append({
                            "sheet": cls.title,
                            "row": err.row_number,
                            "warning": err.error,
                        })
                    # Store raw trace for error rows too
                    self._store_raw_trace_from_error(
                        err, job_id, file_name, cls.title, sheet_visibility,
                    )

            wb.close()

            # Build reconciliation report
            reconciliation = ReconciliationReport(
                total_rows_in_file=total_rows_in_file,
                rows_imported=result.imported,
                rows_skipped=result.skipped,
                rows_with_errors=result.errors,
                rows_with_warnings=result.warnings,
                total_losses_amount=total_losses_amount,
                total_cost_imported=total_cost_imported,
                sheets_processed=sheets_processed,
                sheets_skipped=sheets_skipped,
                hidden_sheets_found=hidden_sheets_found,
            )

            # Update job status to completed
            with app_session() as session:
                job = session.query(TicketImportJob).filter_by(id=job_id).first()
                if job is not None:
                    job.status = "completed"
                    job.result_json = json.dumps({
                        "imported": result.imported,
                        "skipped": result.skipped,
                        "errors": result.errors,
                        "warnings": result.warnings,
                        "reconciliation": {
                            "total_rows_in_file": reconciliation.total_rows_in_file,
                            "rows_imported": reconciliation.rows_imported,
                            "rows_skipped": reconciliation.rows_skipped,
                            "rows_with_errors": reconciliation.rows_with_errors,
                            "rows_with_warnings": reconciliation.rows_with_warnings,
                            "total_losses_amount": str(reconciliation.total_losses_amount),
                            "total_cost_imported": str(reconciliation.total_cost_imported),
                            "sheets_processed": reconciliation.sheets_processed,
                            "sheets_skipped": reconciliation.sheets_skipped,
                            "hidden_sheets_found": reconciliation.hidden_sheets_found,
                        },
                    }, ensure_ascii=False)

        except Exception as exc:
            logger.exception("Import failed for job %s: %s", job_id, exc)
            with app_session() as session:
                job = session.query(TicketImportJob).filter_by(id=job_id).first()
                if job is not None:
                    job.status = "failed"
                    job.result_json = json.dumps(
                        {"error": str(exc)}, ensure_ascii=False
                    )
            raise

        return result

    # ------------------------------------------------------------------
    # Row parsing
    # ------------------------------------------------------------------

    def _parse_row(
        self,
        row: tuple,
        header_map: dict[int, str],
        row_number: int,
        comments: dict[str, str],
        hyperlinks: dict[str, list[str]],
        color_map: dict[str, str],
    ) -> ParsedRow | ParseError:
        """Parse a single row using header-based column mapping.

        Args:
            row: Tuple of cells from openpyxl.
            header_map: Mapping of column index → field name.
            row_number: 1-based row number in the sheet.
            comments: Cell comments dict (coord → text).
            hyperlinks: Cell hyperlinks dict (coord → list of URLs).
            color_map: Confirmed color → status mapping.

        Returns:
            ParsedRow on success, ParseError if mandatory field (ФИО) is missing.
        """
        raw_cells: dict[str, Any] = {}
        cell_colors: dict[str, str] = {}
        cell_formulas: dict[str, str] = {}
        cell_comments: dict[str, str] = {}
        cell_hyperlinks_row: dict[str, str] = {}
        cell_addresses: dict[str, str] = {}

        field_values: dict[str, Any] = {}

        for idx, cell in enumerate(row):
            coord = cell.coordinate
            cell_addresses[f"col_{idx}"] = coord
            raw_cells[coord] = cell.value

            # Extract color
            color_hex = self._get_cell_color(cell)
            if color_hex:
                cell_colors[coord] = color_hex

            # Extract formula
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell_formulas[coord] = cell.value

            # Extract comment
            if coord in comments:
                cell_comments[coord] = comments[coord]

            # Extract hyperlink
            if coord in hyperlinks:
                cell_hyperlinks_row[coord] = hyperlinks[coord][0] if hyperlinks[coord] else ""

            # Map to field
            field_name = header_map.get(idx)
            if field_name:
                field_values[field_name] = cell.value

        # Validate mandatory field: ФИО
        full_name_raw = field_values.get("full_name")
        if not full_name_raw or (isinstance(full_name_raw, str) and not full_name_raw.strip()):
            return ParseError(
                row_number=row_number,
                error="Обязательное поле ФИО отсутствует",
                severity="warning",
                raw_cells=raw_cells,
                cell_colors=cell_colors,
                cell_formulas=cell_formulas,
                cell_comments=cell_comments,
                cell_hyperlinks=cell_hyperlinks_row,
                cell_addresses=cell_addresses,
            )

        full_name = str(full_name_raw).strip()

        # Normalize optional fields
        needs_review = False
        review_fields: list[str] = []

        # Dates
        submitted_at, sr = self._normalize_value(field_values.get("submitted_at"), "date")
        if sr:
            needs_review = True
            review_fields.append("submitted_at")

        departure_date, sr = self._normalize_value(field_values.get("departure_date"), "date")
        if sr:
            needs_review = True
            review_fields.append("departure_date")

        arrival_date, sr = self._normalize_value(field_values.get("arrival_date"), "date")
        if sr:
            needs_review = True
            review_fields.append("arrival_date")

        # Cost
        cost, sr = self._normalize_value(field_values.get("cost"), "decimal")
        if sr:
            needs_review = True
            review_fields.append("cost")
        if cost is None:
            cost = Decimal("0.00")

        # Phone
        phone, sr = self._normalize_value(field_values.get("phone"), "phone")
        if sr:
            needs_review = True
            review_fields.append("phone")

        # Passport
        passport, sr = self._normalize_value(field_values.get("passport"), "passport")
        if sr:
            needs_review = True
            review_fields.append("passport")

        # Route and assignee (simple string fields)
        route = str(field_values.get("route", "") or "").strip() or None
        assignee = str(field_values.get("assignee", "") or "").strip() or None
        status_note = str(field_values.get("status_note", "") or "").strip() or None

        # Determine status from color (first colored cell in the row)
        status: str | None = status_note
        is_urgent = False
        for coord, color in cell_colors.items():
            mapped = self._map_color_to_status(color, color_map)
            if mapped == "urgent_flag":
                is_urgent = True
            elif mapped:
                status = mapped
        if status and status not in VALID_REQUEST_STATUSES:
            needs_review = True
            review_fields.append("status")
            status = "new"

        return ParsedRow(
            full_name=full_name,
            submitted_at=submitted_at,
            departure_date=departure_date,
            arrival_date=arrival_date,
            route=route,
            cost=cost,
            status=status,
            assignee=assignee,
            phone=phone,
            passport=passport,
            is_urgent=is_urgent,
            needs_review=needs_review,
            review_fields=review_fields,
            row_number=row_number,
            raw_cells=raw_cells,
            cell_colors=cell_colors,
            cell_formulas=cell_formulas,
            cell_comments=cell_comments,
            cell_hyperlinks=cell_hyperlinks_row,
            cell_addresses=cell_addresses,
        )

    def _parse_row_standard_schema(
        self,
        row: tuple,
        col_order: list[str],
        row_number: int,
        comments: dict[str, str],
        hyperlinks: dict[str, list[str]],
        color_map: dict[str, str],
    ) -> ParsedRow | ParseError:
        """Parse a row from a headerless sheet using positional column mapping.

        Args:
            row: Tuple of cells from openpyxl.
            col_order: List of field names in column order.
            row_number: 1-based row number.
            comments: Cell comments dict.
            hyperlinks: Cell hyperlinks dict.
            color_map: Confirmed color → status mapping.

        Returns:
            ParsedRow on success, ParseError if ФИО is missing.
        """
        # Build a header_map from col_order (index → field_name)
        header_map: dict[int, str] = {}
        for idx, field_name in enumerate(col_order):
            header_map[idx] = field_name

        return self._parse_row(
            row, header_map, row_number, comments, hyperlinks, color_map
        )

    def _parse_losses_sheet(self, worksheet: Worksheet) -> list[FinancialOpRow]:
        """Parse ПОТЕРИ sheet rows into FinancialOpRow dicts.

        Expected columns: ФИО, сумма, причина, дата, тип операции.
        Tries to detect headers; falls back to positional mapping.

        Args:
            worksheet: The ПОТЕРИ worksheet.

        Returns:
            List of FinancialOpRow parsed from the sheet.
        """
        results: list[FinancialOpRow] = []
        headers = self._extract_headers(worksheet)

        # Build column mapping for losses sheet
        losses_field_map: dict[str, str] = {
            "фио": "full_name",
            "ф.и.о.": "full_name",
            "сотрудник": "full_name",
            "сумма": "amount",
            "стоимость": "amount",
            "причина": "reason",
            "дата": "op_date",
            "тип": "op_type",
            "тип операции": "op_type",
        }

        col_map: dict[int, str] = {}
        has_headers = False
        for idx, h in enumerate(headers):
            h_lower = h.lower().strip()
            if h_lower in losses_field_map:
                col_map[idx] = losses_field_map[h_lower]
                has_headers = True

        # Fallback positional: ФИО, сумма, причина, дата, тип
        if not has_headers:
            default_order = ["full_name", "amount", "reason", "op_date", "op_type"]
            col_map = {i: f for i, f in enumerate(default_order)}

        start_row = 2 if has_headers else 1

        for row_idx, row in self._iter_data_rows(worksheet, min_row=start_row):
            field_values: dict[str, Any] = {}
            for idx, cell in enumerate(row):
                field_name = col_map.get(idx)
                if field_name:
                    field_values[field_name] = cell.value

            full_name_raw = field_values.get("full_name")
            if not full_name_raw or (isinstance(full_name_raw, str) and not full_name_raw.strip()):
                continue  # Skip rows without ФИО

            full_name = str(full_name_raw).strip()

            # Parse amount
            amount_raw = field_values.get("amount")
            amount = Decimal("0.00")
            if amount_raw is not None:
                try:
                    amount = Decimal(str(amount_raw).replace(",", ".").strip())
                except (InvalidOperation, ValueError):
                    amount = Decimal("0.00")

            # Parse date
            op_date: datetime | None = None
            date_raw = field_values.get("op_date")
            if isinstance(date_raw, datetime):
                op_date = date_raw
            elif date_raw is not None:
                op_date, _ = self._normalize_value(date_raw, "date")

            # Parse reason and op_type
            reason = str(field_values.get("reason", "") or "").strip() or None
            op_type_raw = str(field_values.get("op_type", "") or "").strip().lower()
            op_type = "loss"
            if "возврат" in op_type_raw or "refund" in op_type_raw:
                op_type = "refund"
            elif "обмен" in op_type_raw or "exchange" in op_type_raw:
                op_type = "exchange"

            results.append(FinancialOpRow(
                full_name=full_name,
                amount=amount,
                reason=reason,
                op_date=op_date,
                op_type=op_type,
                row_number=row_idx,
            ))

        return results

    # ------------------------------------------------------------------
    # Duplicate detection
    # ------------------------------------------------------------------

    def _detect_duplicates(
        self, parsed_rows: list[ParsedRow], object_id: int
    ) -> list[DuplicateMatch]:
        """Detect duplicates by (full_name + object_id + submitted_at).

        Args:
            parsed_rows: List of parsed rows to check.
            object_id: The object ID for this sheet.

        Returns:
            List of DuplicateMatch for rows that match existing DB records.
        """
        if not parsed_rows:
            return []

        with app_session() as session:
            rows = session.query(
                TicketRequest.id,
                TicketEmployee.full_name,
                TicketRequest.submitted_at,
            ).join(
                TicketEmployee,
                TicketRequest.employee_id == TicketEmployee.id,
            ).filter(
                TicketRequest.object_id == object_id,
            ).all()

        existing_by_key = {
            (full_name, submitted_at): request_id
            for request_id, full_name, submitted_at in rows
        }

        duplicates: list[DuplicateMatch] = []
        for parsed in parsed_rows:
            existing_id = existing_by_key.get((parsed.full_name, parsed.submitted_at))
            if existing_id is not None:
                duplicates.append(DuplicateMatch(
                    parsed_row=parsed,
                    existing_id=existing_id,
                ))

        return duplicates

    # ------------------------------------------------------------------
    # Value normalization
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_value(value: Any, expected_type: str) -> tuple[Any, bool]:
        """Normalize a cell value to the expected type.

        Handles common Excel data corruption:
        - phone-as-date: Excel serial number → phone string
        - passport-as-number: numeric → formatted string
        - date-as-text: text → datetime

        Args:
            value: Raw cell value.
            expected_type: One of 'date', 'decimal', 'phone', 'passport', 'text'.

        Returns:
            Tuple of (normalized_value, needs_review: bool).
            needs_review is True if the value required non-trivial conversion.
        """
        if value is None:
            return (None, False)

        needs_review = False

        if expected_type == "date":
            if isinstance(value, datetime):
                return (value, False)
            if isinstance(value, (int, float)):
                # Excel serial number → datetime
                try:
                    days = int(value)
                    if 1 <= days <= 200000:  # Reasonable date range
                        dt = _EXCEL_EPOCH + timedelta(days=days)
                        return (dt, True)  # needs_review: was a number
                except (ValueError, OverflowError):
                    pass
                return (None, True)
            if isinstance(value, str):
                # Try common date formats
                value_stripped = value.strip()
                for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                    try:
                        dt = datetime.strptime(value_stripped, fmt).replace(
                            tzinfo=timezone.utc
                        )
                        return (dt, True)  # needs_review: was text
                    except ValueError:
                        continue
                return (None, True)
            return (None, True)

        if expected_type == "decimal":
            if isinstance(value, Decimal):
                return (value, False)
            if isinstance(value, (int, float)):
                try:
                    return (Decimal(str(value)), False)
                except (InvalidOperation, ValueError):
                    return (Decimal("0.00"), True)
            if isinstance(value, str):
                cleaned = value.strip().replace(",", ".").replace(" ", "")
                # Remove currency symbols
                cleaned = re.sub(r"[^\d.\-]", "", cleaned)
                try:
                    return (Decimal(cleaned), True)
                except (InvalidOperation, ValueError):
                    return (Decimal("0.00"), True)
            return (Decimal("0.00"), True)

        if expected_type == "phone":
            if isinstance(value, (int, float)):
                # Phone stored as number (e.g. Excel serial date or plain number)
                num = int(value)
                # Check if it looks like an Excel date serial (phone-as-date)
                if 30000 < num < 50000:
                    # This is likely a date serial, not a phone
                    # Convert serial to a date string representation
                    phone_str = str(num)
                    return (phone_str, True)
                # Plain number — format as phone
                phone_str = str(num)
                if len(phone_str) == 10:
                    phone_str = "+7" + phone_str
                elif len(phone_str) == 11 and phone_str.startswith("8"):
                    phone_str = "+7" + phone_str[1:]
                elif len(phone_str) == 11 and phone_str.startswith("7"):
                    phone_str = "+" + phone_str
                return (phone_str, True)
            if isinstance(value, str):
                return (value.strip(), False)
            if isinstance(value, datetime):
                # Phone stored as date — extract the serial number
                delta = value - _EXCEL_EPOCH
                serial = int(delta.total_seconds() / 86400)
                return (str(serial), True)
            return (str(value), True)

        if expected_type == "passport":
            if isinstance(value, (int, float)):
                # Passport stored as number — format with leading zeros
                num_str = str(int(value))
                # Russian passport: 4 digits series + 6 digits number = 10 digits
                if len(num_str) <= 10:
                    num_str = num_str.zfill(10)
                    return (f"{num_str[:4]} {num_str[4:]}", True)
                return (num_str, True)
            if isinstance(value, str):
                return (value.strip(), False)
            return (str(value) if value else None, True)

        # Default: text
        if isinstance(value, str):
            return (value.strip(), False)
        return (str(value), False)

    # ------------------------------------------------------------------
    # Raw trace storage
    # ------------------------------------------------------------------

    def _store_raw_trace(
        self,
        parsed: ParsedRow,
        job_id: str,
        file_name: str,
        sheet_name: str,
        sheet_visibility: str,
        request_id: int | None,
    ) -> None:
        """Store raw trace for an imported row."""
        with app_session() as session:
            trace = TicketImportRawTrace(
                job_id=job_id,
                request_id=request_id,
                source_file=file_name,
                sheet_name=sheet_name,
                row_number=parsed.row_number,
                raw_cells_json=json.dumps(
                    {k: _serialize_value(v) for k, v in parsed.raw_cells.items()},
                    ensure_ascii=False,
                ),
                cell_colors_json=json.dumps(parsed.cell_colors, ensure_ascii=False),
                cell_formulas_json=json.dumps(parsed.cell_formulas, ensure_ascii=False),
                cell_comments_json=json.dumps(parsed.cell_comments, ensure_ascii=False),
                cell_hyperlinks_json=json.dumps(parsed.cell_hyperlinks, ensure_ascii=False),
                cell_addresses_json=json.dumps(parsed.cell_addresses, ensure_ascii=False),
                sheet_visibility=sheet_visibility,
            )
            session.add(trace)

    def _store_raw_trace_from_error(
        self,
        err: ParseError,
        job_id: str,
        file_name: str,
        sheet_name: str,
        sheet_visibility: str,
    ) -> None:
        """Store raw trace for a row that failed parsing."""
        with app_session() as session:
            trace = TicketImportRawTrace(
                job_id=job_id,
                request_id=None,
                source_file=file_name,
                sheet_name=sheet_name,
                row_number=err.row_number,
                raw_cells_json=json.dumps(
                    {k: _serialize_value(v) for k, v in err.raw_cells.items()},
                    ensure_ascii=False,
                ),
                cell_colors_json=json.dumps(err.cell_colors, ensure_ascii=False),
                cell_formulas_json=json.dumps(err.cell_formulas, ensure_ascii=False),
                cell_comments_json=json.dumps(err.cell_comments, ensure_ascii=False),
                cell_hyperlinks_json=json.dumps(err.cell_hyperlinks, ensure_ascii=False),
                cell_addresses_json=json.dumps(err.cell_addresses, ensure_ascii=False),
                sheet_visibility=sheet_visibility,
            )
            session.add(trace)

    # ------------------------------------------------------------------
    # DB record creation
    # ------------------------------------------------------------------

    def _create_request(self, parsed: ParsedRow, object_id: int) -> int:
        """Create a TicketRequest + TicketEmployee from a parsed row.

        Returns the created request ID.
        """
        full_name = parsed.full_name[:150]
        phone = parsed.phone[:30] if parsed.phone else None
        route = parsed.route[:255] if parsed.route else None
        status = parsed.status if parsed.status in VALID_REQUEST_STATUSES else "new"
        with app_session() as session:
            # Find or create employee
            employee = session.query(TicketEmployee).filter_by(
                full_name=full_name
            ).first()
            if employee is None:
                employee = TicketEmployee(
                    full_name=full_name,
                    phone=phone,
                )
                session.add(employee)
                session.flush()

            request = TicketRequest(
                employee_id=employee.id,
                object_id=object_id,
                status=status,
                submitted_at=parsed.submitted_at,
                departure_date=parsed.departure_date,
                arrival_date=parsed.arrival_date,
                route=route,
                total_cost=parsed.cost,
                is_urgent=parsed.is_urgent,
                needs_review=parsed.needs_review,
                source="import",
            )
            session.add(request)
            session.flush()
            return request.id

    def _update_existing_request(
        self, request_id: int, parsed: ParsedRow, object_id: int
    ) -> int:
        """Update an existing TicketRequest with new data from import."""
        with app_session() as session:
            request = session.query(TicketRequest).filter_by(id=request_id).first()
            if request is None:
                # Fallback: create new
                return self._create_request(parsed, object_id)

            if parsed.departure_date is not None:
                request.departure_date = parsed.departure_date
            if parsed.arrival_date is not None:
                request.arrival_date = parsed.arrival_date
            if parsed.route:
                request.route = parsed.route[:255]
            if parsed.cost > Decimal("0.00"):
                request.total_cost = parsed.cost
            if parsed.status in VALID_REQUEST_STATUSES:
                request.status = parsed.status
            request.is_urgent = parsed.is_urgent
            if parsed.needs_review:
                request.needs_review = True

            return request_id

    def _store_losses(
        self,
        fin_ops: list[FinancialOpRow],
        job_id: str,
        file_name: str,
        sheet_name: str,
        sheet_visibility: str,
        comments: dict[str, str],
        hyperlinks: dict[str, list[str]],
    ) -> None:
        """Store parsed ПОТЕРИ rows as TicketFinancialOp records and raw traces."""
        with app_session() as session:
            for op in fin_ops:
                # Find or create employee
                employee = session.query(TicketEmployee).filter_by(
                    full_name=op.full_name
                ).first()
                if employee is None:
                    employee = TicketEmployee(full_name=op.full_name)
                    session.add(employee)
                    session.flush()

                fin_record = TicketFinancialOp(
                    employee_id=employee.id,
                    op_type=op.op_type,
                    amount=op.amount,
                    reason=op.reason,
                    op_date=op.op_date,
                )
                session.add(fin_record)

                # Store raw trace for losses row
                trace = TicketImportRawTrace(
                    job_id=job_id,
                    request_id=None,
                    source_file=file_name,
                    sheet_name=sheet_name,
                    row_number=op.row_number,
                    raw_cells_json=json.dumps(
                        {"full_name": op.full_name, "amount": str(op.amount),
                         "reason": op.reason, "op_type": op.op_type},
                        ensure_ascii=False,
                    ),
                    cell_colors_json="{}",
                    cell_formulas_json="{}",
                    cell_comments_json="{}",
                    cell_hyperlinks_json="{}",
                    cell_addresses_json="{}",
                    sheet_visibility=sheet_visibility,
                )
                session.add(trace)

    def _build_reconciliation_report(self, result: ImportResult) -> ReconciliationReport:
        """Build a reconciliation report from import results.

        This is a convenience method — the actual report is built inline
        in execute_import() with access to all counters.
        """
        return ReconciliationReport(
            rows_imported=result.imported,
            rows_skipped=result.skipped,
            rows_with_errors=result.errors,
            rows_with_warnings=result.warnings,
        )

    # ------------------------------------------------------------------
    # Color mapping helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
        """Convert a 6-character hex color string to an (R, G, B) tuple."""
        hex_color = hex_color.strip().lstrip("#").upper()
        if len(hex_color) == 8:
            hex_color = hex_color[2:]
        if len(hex_color) != 6:
            return (0, 0, 0)
        try:
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            return (r, g, b)
        except ValueError:
            return (0, 0, 0)

    @staticmethod
    def _color_distance(hex1: str, hex2: str) -> float:
        """Calculate Euclidean distance between two hex colors in RGB space."""
        r1, g1, b1 = TicketsImportService._hex_to_rgb(hex1)
        r2, g2, b2 = TicketsImportService._hex_to_rgb(hex2)
        return math.sqrt((r1 - r2) ** 2 + (g1 - g2) ** 2 + (b1 - b2) ** 2)

    def _extract_unique_colors(self, worksheet: Worksheet) -> set[str]:
        """Extract unique cell background fill colors from a worksheet."""
        colors: set[str] = set()
        for (_row_idx, col_idx), cell in getattr(worksheet, "_cells", {}).items():
            if col_idx > MAX_IMPORT_SCAN_COLUMNS:
                continue
            fill = cell.fill
            if fill is None:
                continue
            fg_color = fill.fgColor
            if fg_color is None:
                continue
            color_type = fg_color.type
            if color_type is None:
                continue
            hex_val: str | None = None
            if color_type == "rgb" and fg_color.rgb:
                raw = str(fg_color.rgb).upper()
                if raw == "00000000":
                    continue
                if len(raw) == 8:
                    hex_val = raw[2:]
                elif len(raw) == 6:
                    hex_val = raw
            elif color_type == "indexed":
                continue
            elif color_type == "theme":
                continue
            if hex_val is None:
                continue
            if hex_val in ("FFFFFF", "000000"):
                continue
            colors.add(hex_val)
        return colors

    def _suggest_color_mapping(
        self,
        unique_colors: list[str],
        reference_colors: dict[str, str] | None = None,
    ) -> dict[str, str]:
        """Suggest a color→status mapping using Euclidean distance in RGB space."""
        if reference_colors is None:
            reference_colors = DEFAULT_REFERENCE_COLORS
        if not unique_colors or not reference_colors:
            return {}
        result: dict[str, str] = {}
        for color in unique_colors:
            best_status: str | None = None
            best_distance: float = float("inf")
            for ref_hex, status in reference_colors.items():
                dist = self._color_distance(color, ref_hex)
                if dist < best_distance:
                    best_distance = dist
                    best_status = status
            if best_status is not None:
                result[color] = best_status
        return result

    @staticmethod
    def _map_color_to_status(hex_color: str, color_map: dict[str, str]) -> str | None:
        """Look up a hex color in the confirmed color map."""
        if not hex_color or not color_map:
            return None
        normalized = hex_color.strip().lstrip("#").upper()
        if len(normalized) == 8:
            normalized = normalized[2:]
        return color_map.get(normalized)

    # ------------------------------------------------------------------
    # Header extraction and mapping
    # ------------------------------------------------------------------

    @classmethod
    def _extract_headers(cls, worksheet: Worksheet) -> list[str]:
        """Extract header values from the first row of a worksheet."""
        headers: list[str] = []
        max_col = cls._max_data_column(worksheet)
        for col_idx in range(1, max_col + 1):
            val = worksheet.cell(row=1, column=col_idx).value
            headers.append(str(val).strip() if val is not None else "")
        if not any(headers):
            return []
        return headers

    @staticmethod
    def _has_valid_headers(headers: list[str]) -> bool:
        """Check if the headers contain at least one recognized field name."""
        for h in headers:
            if h.lower().strip() in _HEADER_ALIASES:
                return True
        return False

    @staticmethod
    def _build_header_map(headers: list[str]) -> dict[int, str]:
        """Build a mapping of column index → field name from headers."""
        result: dict[int, str] = {}
        for idx, h in enumerate(headers):
            h_lower = h.lower().strip()
            if h_lower in _HEADER_ALIASES:
                result[idx] = _HEADER_ALIASES[h_lower]
        return result

    # ------------------------------------------------------------------
    # Sheet classification
    # ------------------------------------------------------------------

    def _classify_sheets(self, workbook: Workbook) -> list[SheetClassification]:
        """Classify each sheet in the workbook."""
        results: list[SheetClassification] = []
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            sheet_state = getattr(ws, "sheet_state", "visible") or "visible"
            row_count = self._count_data_rows(ws)
            if sheet_state != "visible":
                classification = "скрытый"
            elif "ПОТЕРИ" in sheet_name or "потери" in sheet_name:
                classification = "ПОТЕРИ"
            elif row_count <= 1:
                classification = "пустой"
            elif sheet_name.startswith("_") or "служебн" in sheet_name.lower():
                classification = "служебный"
            else:
                classification = "рабочий"
            results.append(SheetClassification(
                title=sheet_name,
                classification=classification,
                row_count=row_count,
            ))
        return results

    def _detect_hidden_sheets(self, workbook: Workbook) -> list[HiddenSheetInfo]:
        """Detect visibility state of all sheets."""
        results: list[HiddenSheetInfo] = []
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            state = getattr(ws, "sheet_state", "visible") or "visible"
            results.append(HiddenSheetInfo(title=sheet_name, state=state))
        return results

    # ------------------------------------------------------------------
    # Cell metadata extraction
    # ------------------------------------------------------------------

    def _extract_cell_hyperlinks(self, worksheet: Worksheet) -> dict[str, list[str]]:
        """Extract hyperlinks from all cells in a worksheet."""
        result: dict[str, list[str]] = {}
        if hasattr(worksheet, "hyperlinks"):
            for link in worksheet.hyperlinks:
                ref = link.ref or ""
                target = link.target or ""
                if ref and target:
                    coord = ref.split(":")[0]
                    if coord in result:
                        result[coord].append(target)
                    else:
                        result[coord] = [target]
        for (_row_idx, col_idx), cell in getattr(worksheet, "_cells", {}).items():
            if col_idx > MAX_IMPORT_SCAN_COLUMNS or cell.hyperlink is None:
                continue
            target = ""
            if hasattr(cell.hyperlink, "target") and cell.hyperlink.target:
                target = cell.hyperlink.target
            elif isinstance(cell.hyperlink, str):
                target = cell.hyperlink
            if target:
                coord = cell.coordinate
                if coord in result:
                    if target not in result[coord]:
                        result[coord].append(target)
                else:
                    result[coord] = [target]
        return result

    def _extract_cell_comments(self, worksheet: Worksheet) -> dict[str, str]:
        """Extract comments from all cells in a worksheet."""
        result: dict[str, str] = {}
        for (_row_idx, col_idx), cell in getattr(worksheet, "_cells", {}).items():
            if col_idx > MAX_IMPORT_SCAN_COLUMNS or cell.comment is None:
                continue
            text = cell.comment.text or ""
            if text:
                result[cell.coordinate] = text
        return result

    def _extract_threaded_comments(self, worksheet: Worksheet) -> dict[str, list[str]]:
        """Extract threaded comments from a worksheet (if available)."""
        result: dict[str, list[str]] = {}
        try:
            wb = worksheet.parent
            if wb is None:
                return result
            threaded_comments = getattr(wb, "_threaded_comments", None)
            if threaded_comments is None:
                return result
            sheet_title = worksheet.title
            for comment_thread in threaded_comments:
                ref = getattr(comment_thread, "ref", None)
                sheet_ref = getattr(comment_thread, "sheet", None)
                if sheet_ref and sheet_ref != sheet_title:
                    continue
                if ref:
                    texts = []
                    replies = getattr(comment_thread, "comments", []) or []
                    for reply in replies:
                        text = getattr(reply, "text", None) or getattr(reply, "content", "")
                        if text:
                            texts.append(str(text))
                    if texts:
                        result[str(ref)] = texts
        except (AttributeError, TypeError, StopIteration):
            pass
        return result

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _count_data_rows(worksheet: Worksheet) -> int:
        """Count rows that contain at least one non-empty cell."""
        return len({
            row_idx
            for (row_idx, _col_idx), cell in getattr(worksheet, "_cells", {}).items()
            if cell.value is not None
        })

    @staticmethod
    def _max_data_column(worksheet: Worksheet) -> int:
        """Return the highest populated column to scan, capped for styled Excel tails."""
        max_col = 0
        for (_row_idx, col_idx), cell in getattr(worksheet, "_cells", {}).items():
            if col_idx <= MAX_IMPORT_SCAN_COLUMNS and cell.value is not None:
                max_col = max(max_col, col_idx)
        if max_col:
            return max_col
        return min(int(getattr(worksheet, "max_column", 1) or 1), MAX_IMPORT_SCAN_COLUMNS)

    @classmethod
    def _iter_data_rows(cls, worksheet: Worksheet, *, min_row: int = 1):
        """Yield populated worksheet rows without expanding huge empty Excel ranges."""
        populated_rows = sorted({
            row_idx
            for (row_idx, _col_idx), cell in getattr(worksheet, "_cells", {}).items()
            if row_idx >= min_row and cell.value is not None
        })
        if not populated_rows:
            return
        max_col = cls._max_data_column(worksheet)
        for row_idx in populated_rows:
            row = tuple(
                worksheet.cell(row=row_idx, column=col_idx)
                for col_idx in range(1, max_col + 1)
            )
            if any(cell.value is not None for cell in row):
                yield row_idx, row

    @staticmethod
    def _get_cell_color(cell) -> str | None:
        """Extract the background fill color hex from a cell."""
        fill = cell.fill
        if fill is None:
            return None
        fg_color = fill.fgColor
        if fg_color is None:
            return None
        color_type = fg_color.type
        if color_type is None:
            return None
        if color_type == "rgb" and fg_color.rgb:
            raw = str(fg_color.rgb).upper()
            if raw == "00000000":
                return None
            if len(raw) == 8:
                return raw[2:]
            if len(raw) == 6:
                return raw
        return None


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------


def _serialize_value(value: Any) -> Any:
    """Serialize a cell value for JSON storage."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


_OPENPYXL_TZ_PATCHED = False


def _patch_openpyxl_timezone_datetimes() -> None:
    """Normalize aware datetimes before openpyxl writes Excel cells.

    Excel files do not store timezone metadata. Current openpyxl versions raise
    at save time for timezone-aware datetimes; the importer treats those values
    as plain Excel datetimes, so stripping tzinfo is the least surprising
    compatibility behavior.
    """
    global _OPENPYXL_TZ_PATCHED
    if _OPENPYXL_TZ_PATCHED:
        return

    try:
        import openpyxl.cell._writer as writer
    except Exception:
        return

    original_set_attributes = writer._set_attributes

    def _set_attributes_without_tz(cell, styled=None):
        value = getattr(cell, "value", None)
        if isinstance(value, datetime) and value.tzinfo is not None:
            cell.value = value.replace(tzinfo=None)
        return original_set_attributes(cell, styled)

    writer._set_attributes = _set_attributes_without_tz
    _OPENPYXL_TZ_PATCHED = True
