from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TASK_STATUS_LABELS = {
    "queued": "В очереди",
    "delivered": "Доставлено агенту",
    "acknowledged": "Выполняется",
    "completed": "Завершено",
    "failed": "Ошибка",
    "expired": "Просрочено",
}
INCIDENT_STATUS_LABELS = {
    "new": "Новый",
    "ack": "Просмотрен",
    "resolved_deleted": "Удалён",
    "resolved_clean": "Очищен",
    "resolved_moved": "Перемещён",
}
SEVERITY_LABELS = {
    "high": "Высокая",
    "medium": "Средняя",
    "low": "Низкая",
    "none": "Не указана",
}
OBSERVATION_TYPE_LABELS = {
    "found_new": "Найден впервые",
    "found_duplicate": "Найден повторно",
    "deleted": "Удалён",
    "cleaned": "Очищен",
    "moved": "Перемещён",
}
SOURCE_KIND_LABELS = {
    "pdf_slice": "Фрагмент PDF",
    "pdf": "PDF",
    "text": "Текстовый файл",
    "metadata": "Метаданные",
    "unknown": "Неизвестный источник",
}
CATEGORY_LABELS = {
    "secrets": "Конфиденциальные данные",
    "policy_match": "Совпадение с политикой",
}


def _format_ts(value: Any) -> str:
    try:
        ts = int(value or 0)
    except Exception:
        ts = 0
    if ts <= 0:
        return ""
    return datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M:%S")


def _format_date_compact(value: Any) -> str:
    try:
        ts = int(value or 0)
    except Exception:
        ts = 0
    if ts <= 0:
        return datetime.now().strftime("%Y%m%d")
    return datetime.fromtimestamp(ts).strftime("%Y%m%d")


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return str(value)
    return str(value)


def _pattern_summary(value: Any) -> str:
    patterns = value if isinstance(value, list) else []
    out: List[str] = []
    for item in patterns:
        if not isinstance(item, dict):
            continue
        name = str(item.get("pattern_name") or item.get("pattern") or item.get("id") or "").strip()
        matched_value = str(item.get("value") or item.get("snippet") or "").strip()
        if name and matched_value:
            out.append(f"{name}: {matched_value}")
        elif name:
            out.append(name)
    return "\n".join(out)


def _localized_label(value: Any, labels: Dict[str, str]) -> str:
    normalized = str(value or "").strip().lower()
    return labels.get(normalized, str(value or ""))


def _localized_observation_types(value: Any) -> str:
    raw_types = [item.strip() for item in str(value or "").split(",") if item.strip()]
    return ", ".join(_localized_label(item, OBSERVATION_TYPE_LABELS) for item in raw_types)


def _sanitize_filename_part(value: Any, fallback: str) -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
    text = text.strip("._-")
    return text[:80] or fallback


def _fit_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 10
        for cell in column_cells:
            max_len = max(max_len, min(len(str(cell.value or "")), 80))
        ws.column_dimensions[letter].width = max_len + 2


def _append_table(ws, headers: Iterable[str], rows: Iterable[Iterable[Any]], *, empty_message: str = "Нет данных") -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF8")
    ws.append(list(headers))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    row_count = 0
    for row in rows:
        ws.append(list(row))
        row_count += 1
    if row_count == 0:
        ws.append([empty_message])
    _fit_columns(ws)


def build_scan_task_incidents_excel(report: Dict[str, Any]) -> Tuple[bytes, str]:
    task = report.get("task") if isinstance(report.get("task"), dict) else {}
    result = task.get("result") if isinstance(task.get("result"), dict) else {}
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    incidents = report.get("incidents") if isinstance(report.get("incidents"), list) else []
    observations = report.get("observations") if isinstance(report.get("observations"), list) else []

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Сводка"
    summary_ws.append(["Отчёт по запуску сканирования"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([])

    summary_rows = [
        ("ID запуска", task.get("id")),
        ("Агент", task.get("agent_id")),
        ("Хост", task.get("hostname") or summary.get("hostname")),
        ("Статус", _localized_label(task.get("status"), TASK_STATUS_LABELS)),
        ("Начало запуска", _format_ts(summary.get("scan_started_at") or task.get("created_at"))),
        (
            "Конец запуска",
            _format_ts(summary.get("scan_finished_at") or task.get("completed_at") or task.get("updated_at")),
        ),
        ("Создан", _format_ts(task.get("created_at"))),
        ("Обновлен", _format_ts(task.get("updated_at"))),
        ("Завершен", _format_ts(task.get("completed_at"))),
        ("Проверено", result.get("scanned")),
        ("Пропущено", result.get("skipped")),
        ("Заданий обработки — всего", summary.get("jobs_total")),
        ("Заданий без инцидента", summary.get("jobs_done_clean")),
        ("Заданий с инцидентом", summary.get("jobs_done_with_incident")),
        ("Заданий с ошибкой", summary.get("jobs_failed")),
        ("Не удалось проверить", summary.get("jobs_incomplete")),
        ("Новые находки", summary.get("found_new")),
        ("Повторные находки", summary.get("found_duplicate")),
        ("Удалены", summary.get("deleted")),
        ("Очищены", summary.get("cleaned")),
        ("Перемещены", summary.get("moved")),
        ("Инцидентов в отчете", summary.get("incidents_total")),
    ]
    for label, value in summary_rows:
        summary_ws.append([label, value if value is not None else ""])

    summary_ws.append([])
    _append_table(
        summary_ws,
        ["Критичность", "Количество"],
        [
            (_localized_label(name, SEVERITY_LABELS), count)
            for name, count in sorted((summary.get("severity_counts") or {}).items())
        ],
        empty_message="Инцидентов нет",
    )
    summary_ws.append([])
    _append_table(
        summary_ws,
        ["Статус", "Количество"],
        [
            (_localized_label(name, INCIDENT_STATUS_LABELS), count)
            for name, count in sorted((summary.get("status_counts") or {}).items())
        ],
        empty_message="Инцидентов нет",
    )
    _fit_columns(summary_ws)

    incidents_ws = wb.create_sheet("Инциденты")
    incident_headers = [
        "Время",
        "Событие запуска",
        "Файл",
        "Путь",
        "Хост",
        "Пользователь",
        "Филиал",
        "Критичность",
        "Статус",
        "Источник",
        "Категория",
        "Причина",
        "Паттерны/фрагменты",
        "ID задания обработки",
        "ID инцидента",
    ]
    incident_rows = []
    for item in incidents:
        if not isinstance(item, dict):
            continue
        incident_rows.append([
            _format_ts(item.get("observation_created_at") or item.get("created_at")),
            _localized_observation_types(item.get("observation_types") or item.get("observation_type")),
            item.get("file_name") or "",
            item.get("file_path") or "",
            item.get("hostname") or "",
            item.get("user_full_name") or item.get("user_login") or "",
            item.get("branch") or "",
            _localized_label(item.get("severity"), SEVERITY_LABELS),
            _localized_label(item.get("status"), INCIDENT_STATUS_LABELS),
            _localized_label(item.get("source_kind"), SOURCE_KIND_LABELS),
            _localized_label(item.get("category"), CATEGORY_LABELS),
            item.get("short_reason") or "",
            _pattern_summary(item.get("matched_patterns")),
            item.get("job_id") or "",
            item.get("id") or "",
        ])
    _append_table(incidents_ws, incident_headers, incident_rows, empty_message="Инцидентов в этом запуске нет")

    observations_ws = wb.create_sheet("Наблюдения")
    observation_headers = [
        "Время",
        "Тип события",
        "Путь",
        "Контрольная сумма",
        "Источник",
        "Критичность",
        "Хост",
        "ID события",
        "ID задания обработки",
        "ID инцидента",
    ]
    observation_rows = []
    for item in observations:
        if not isinstance(item, dict):
            continue
        observation_rows.append([
            _format_ts(item.get("created_at")),
            _localized_label(item.get("observation_type"), OBSERVATION_TYPE_LABELS),
            item.get("file_path") or "",
            item.get("file_hash") or "",
            _localized_label(item.get("source_kind"), SOURCE_KIND_LABELS),
            _localized_label(item.get("severity"), SEVERITY_LABELS),
            item.get("hostname") or task.get("hostname") or summary.get("hostname") or "",
            item.get("event_id") or "",
            item.get("linked_job_id") or "",
            item.get("linked_incident_id") or "",
        ])
    _append_table(
        observations_ws,
        observation_headers,
        observation_rows,
        empty_message="Наблюдений в этом запуске нет",
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    hostname = _sanitize_filename_part(task.get("hostname") or summary.get("hostname"), "host")
    task_part = _sanitize_filename_part(str(task.get("id") or "task")[:8], "task")
    scan_date = _format_date_compact(summary.get("scan_started_at") or task.get("created_at"))
    return buffer.getvalue(), f"scan_{hostname}_{scan_date}_{task_part}.xlsx"
