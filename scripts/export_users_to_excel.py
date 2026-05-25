#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспорт пользователей веб-приложения в Excel с разбивкой по правам.

Запуск из корня проекта:
    python scripts/export_users_to_excel.py
    python scripts/export_users_to_excel.py --out my_report.xlsx

Источник данных определяется автоматически:
  - если задан APP_DATABASE_URL (из .env или переменных окружения) → PostgreSQL
  - иначе → data/web_users.json (legacy SQLite/JSON хранилище)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Пути
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"
DATA_DIR = PROJECT_ROOT / "data"
WEB_USERS_JSON = DATA_DIR / "web_users.json"

# Добавляем пути для импорта backend-кода (нужен если читаем из PostgreSQL)
for _p in [str(PROJECT_ROOT), str(WEB_ROOT)]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

# ---------------------------------------------------------------------------
# Загрузка .env (если есть)
# ---------------------------------------------------------------------------
def _expand_vars(val: str, env_map: dict) -> str:
    """Resolve ${VAR} and $VAR references from env_map then os.environ."""
    import re
    def _repl(m):
        name = m.group(1) or m.group(2)
        return env_map.get(name) or os.environ.get(name) or m.group(0)
    return re.sub(r'\$\{([^}]+)\}|\$([A-Za-z_][A-Za-z0-9_]*)', _repl, val)


def _load_dotenv() -> None:
    candidates = [
        PROJECT_ROOT / ".env",
        WEB_ROOT / "backend" / ".env",
        WEB_ROOT / ".env",
    ]
    for env_path in candidates:
        if env_path.exists():
            env_map: dict = {}
            with open(env_path, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, val = line.partition("=")
                    key = key.strip()
                    val = val.strip().strip('"').strip("'")
                    if key:
                        env_map[key] = val
            # Resolve cross-references and set into os.environ
            for key, val in env_map.items():
                resolved = _expand_vars(val, env_map)
                if key not in os.environ:
                    os.environ[key] = resolved
            print(f"[INFO] Загружен .env: {env_path}")
            break


# ---------------------------------------------------------------------------
# Определение прав по роли (дублируем логику из authorization_service.py)
# ---------------------------------------------------------------------------
_VIEWER_PERMISSIONS = {
    "dashboard.read",
    "tasks.read",
    "settings.read",
    "chat.read",
    "chat.write",
    "chat.ai.use",
    "mail.access",
}

_OPERATOR_EXTRA = {
    "announcements.write",
    "tasks.write",
    "database.read",
    "database.write",
    "networks.read",
    "networks.write",
    "computers.read",
    "scan.read",
    "scan.ack",
    "scan.tasks",
    "statistics.read",
    "kb.read",
    "kb.write",
    "vcs.read",
}

_ADMIN_EXTRA = {
    "tasks.review",
    "tasks.manage_all",
    "computers.read_all",
    "kb.publish",
    "kb.manage_all",
    "departments.manage",
    "settings.users.manage",
    "settings.sessions.manage",
    "settings.ai.manage",
    "mail.access",
    "ad_users.read",
    "ad_users.manage",
    "vcs.manage",
    "tickets.read",
    "tickets.write",
    "tickets.personal_data.read",
}

_ROLE_PERMISSIONS: dict[str, set[str]] = {
    "viewer": _VIEWER_PERMISSIONS,
    "operator": _VIEWER_PERMISSIONS | _OPERATOR_EXTRA,
    "admin": _VIEWER_PERMISSIONS | _OPERATOR_EXTRA | _ADMIN_EXTRA,
}

# Все известные права в логическом порядке для заголовков таблицы
_ALL_PERMISSIONS_ORDERED = [
    "dashboard.read",
    "tasks.read",
    "tasks.write",
    "tasks.review",
    "tasks.manage_all",
    "database.read",
    "database.write",
    "networks.read",
    "networks.write",
    "computers.read",
    "computers.read_all",
    "scan.read",
    "scan.ack",
    "scan.tasks",
    "statistics.read",
    "kb.read",
    "kb.write",
    "kb.publish",
    "kb.manage_all",
    "departments.manage",
    "settings.read",
    "settings.users.manage",
    "settings.sessions.manage",
    "settings.ai.manage",
    "mail.access",
    "chat.read",
    "chat.write",
    "chat.ai.use",
    "ad_users.read",
    "ad_users.manage",
    "vcs.read",
    "vcs.manage",
    "tickets.read",
    "tickets.write",
    "tickets.personal_data.read",
    "announcements.write",
]


def _get_effective_permissions(user: dict) -> set[str]:
    if user.get("use_custom_permissions"):
        raw = user.get("custom_permissions") or []
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                raw = []
        return set(raw) if isinstance(raw, list) else set()
    role = str(user.get("role") or "viewer").lower()
    return _ROLE_PERMISSIONS.get(role, _VIEWER_PERMISSIONS)


# ---------------------------------------------------------------------------
# Чтение пользователей
# ---------------------------------------------------------------------------
def _load_users_from_json() -> list[dict]:
    if not WEB_USERS_JSON.exists():
        print(f"[WARN] Файл не найден: {WEB_USERS_JSON}")
        return []
    with open(WEB_USERS_JSON, encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def _load_users_from_postgres(db_url: str) -> list[dict]:
    try:
        from sqlalchemy import create_engine, select, text
        from sqlalchemy.orm import Session
    except ImportError:
        print("[ERROR] sqlalchemy не установлен. Установите: pip install sqlalchemy psycopg2-binary")
        sys.exit(1)

    try:
        engine = create_engine(db_url, pool_pre_ping=True)
        with Session(engine) as session:
            result = session.execute(
                text(
                    "SELECT id, username, email, full_name, department, job_title, "
                    "is_active, role, use_custom_permissions, custom_permissions_json, "
                    "auth_source, telegram_id, assigned_database, created_at, updated_at "
                    "FROM app.users ORDER BY id"
                )
            )
            rows = result.mappings().all()
    except Exception as exc:
        print(f"[ERROR] Ошибка подключения к БД: {exc}")
        print("[INFO] Пробую читать из JSON файла...")
        return _load_users_from_json()

    users = []
    for row in rows:
        row_dict = dict(row)
        cp_raw = row_dict.get("custom_permissions_json") or "[]"
        try:
            cp = json.loads(cp_raw) if isinstance(cp_raw, str) else cp_raw
        except Exception:
            cp = []
        row_dict["custom_permissions"] = cp if isinstance(cp, list) else []
        users.append(row_dict)
    return users


def load_users() -> list[dict]:
    _load_dotenv()
    db_url = os.environ.get("APP_DATABASE_URL", "").strip()
    # Skip unresolved template references
    if db_url and db_url.startswith("${"):
        db_url = ""
    if not db_url:
        db_url = os.environ.get("CHAT_DATABASE_URL", "").strip()
        if db_url and not db_url.startswith("${"):
            print(f"[INFO] APP_DATABASE_URL пуст, использую CHAT_DATABASE_URL")
    if db_url and not db_url.startswith("${"):
        print(f"[INFO] Читаю пользователей из PostgreSQL ({db_url[:50]}...)")
        return _load_users_from_postgres(db_url)
    print(f"[INFO] DB URL не задан, читаю из {WEB_USERS_JSON}")
    return _load_users_from_json()


# ---------------------------------------------------------------------------
# Формирование Excel
# ---------------------------------------------------------------------------
def _fmt_dt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y %H:%M")
    s = str(val)
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S"):
        try:
            dt = datetime.strptime(s[:26], fmt[:len(s[:26])])
            return dt.strftime("%d.%m.%Y %H:%M")
        except ValueError:
            continue
    return s[:16]


def _role_ru(role: str) -> str:
    return {"admin": "Администратор", "operator": "Оператор", "viewer": "Просмотр"}.get(role, role)


def _auth_ru(source: str) -> str:
    return {"local": "Локальная", "ldap": "LDAP/AD"}.get(source, source)


def build_excel(users: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl не установлен. Установите: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Цветовая палитра
    # -----------------------------------------------------------------------
    CLR_HEADER_MAIN = "1F3864"   # тёмно-синий
    CLR_HEADER_PERM = "2E4C8A"   # средне-синий
    CLR_ADMIN      = "E2EFDA"    # светло-зелёный
    CLR_OPERATOR   = "EBF3FB"    # светло-голубой
    CLR_VIEWER     = "FAFAFA"    # почти белый
    CLR_INACTIVE   = "F4CCCC"    # розоватый
    CLR_YES        = "C6EFCE"    # зелёный — есть право
    CLR_NO         = "FFFFFF"    # белый — нет права
    CLR_CUSTOM_YES = "FFF2CC"    # жёлтый — кастомное право
    THIN = Side(border_style="thin", color="CCCCCC")

    def _border():
        return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _hdr_font(white=True):
        return Font(bold=True, color="FFFFFF" if white else "1F3864", size=10)

    def _cell_fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # -----------------------------------------------------------------------
    # Лист 1 — Сводная таблица с чекбоксами прав
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Пользователи и права"

    # Группы прав
    PERM_GROUPS = {
        "Задачи": ["tasks.read", "tasks.write", "tasks.review", "tasks.manage_all"],
        "База данных": ["database.read", "database.write"],
        "Сеть": ["networks.read", "networks.write"],
        "Компьютеры": ["computers.read", "computers.read_all"],
        "Сканирование": ["scan.read", "scan.ack", "scan.tasks"],
        "Статистика": ["statistics.read"],
        "База знаний": ["kb.read", "kb.write", "kb.publish", "kb.manage_all"],
        "Отделы": ["departments.manage"],
        "Настройки": ["settings.read", "settings.users.manage", "settings.sessions.manage", "settings.ai.manage"],
        "Почта": ["mail.access"],
        "Чат": ["chat.read", "chat.write", "chat.ai.use"],
        "AD Users": ["ad_users.read", "ad_users.manage"],
        "VCS": ["vcs.read", "vcs.manage"],
        "Прочее": ["announcements.write", "dashboard.read"],
    }

    # Формируем финальный список прав с заголовками групп
    perm_columns: list[str] = []  # список кодов прав
    group_spans: list[tuple[str, int, int]] = []  # (группа, col_start, col_end) 1-indexed offset from FIRST_PERM_COL
    for grp_name, perms in PERM_GROUPS.items():
        start = len(perm_columns)
        for p in perms:
            if p in _ALL_PERMISSIONS_ORDERED:
                perm_columns.append(p)
        end = len(perm_columns)
        if end > start:
            group_spans.append((grp_name, start, end - 1))

    BASE_COLS = ["ID", "Логин", "Полное имя", "Отдел", "Должность", "Email",
                 "Роль", "Статус", "Авторизация", "2FA", "Telegram ID",
                 "База данных", "Кастомные права", "Дата создания", "Дата изменения"]
    N_BASE = len(BASE_COLS)
    FIRST_PERM_COL = N_BASE + 1  # 1-indexed

    # --- Строка 1: группы прав (merged) ---
    # Пустые ячейки для базовых колонок
    for c in range(1, N_BASE + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _cell_fill(CLR_HEADER_MAIN)
        cell.border = _border()

    for grp_name, gs, ge in group_spans:
        col_s = FIRST_PERM_COL + gs
        col_e = FIRST_PERM_COL + ge
        if col_s == col_e:
            cell = ws.cell(row=1, column=col_s, value=grp_name)
            cell.font = _hdr_font()
            cell.fill = _cell_fill(CLR_HEADER_PERM)
            cell.alignment = _center()
            cell.border = _border()
        else:
            ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
            cell = ws.cell(row=1, column=col_s, value=grp_name)
            cell.font = _hdr_font()
            cell.fill = _cell_fill(CLR_HEADER_PERM)
            cell.alignment = _center()
            cell.border = _border()

    # --- Строка 2: заголовки ---
    for c, hdr in enumerate(BASE_COLS, start=1):
        cell = ws.cell(row=2, column=c, value=hdr)
        cell.font = _hdr_font()
        cell.fill = _cell_fill(CLR_HEADER_MAIN)
        cell.alignment = _center()
        cell.border = _border()
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    for i, perm in enumerate(perm_columns):
        col = FIRST_PERM_COL + i
        short = perm.split(".")[-1]
        cell = ws.cell(row=2, column=col, value=short)
        cell.font = _hdr_font()
        cell.fill = _cell_fill(CLR_HEADER_PERM)
        cell.alignment = _center()
        cell.border = _border()

    # --- Строки данных ---
    for row_idx, user in enumerate(users, start=3):
        username = str(user.get("username") or "")
        if username.startswith("__ai_bot__"):
            continue

        role = str(user.get("role") or "viewer").lower()
        is_active = bool(user.get("is_active", True))
        use_custom = bool(user.get("use_custom_permissions", False))
        eff_perms = _get_effective_permissions(user)
        cp_list = user.get("custom_permissions") or []
        if isinstance(cp_list, str):
            try:
                cp_list = json.loads(cp_list)
            except Exception:
                cp_list = []

        # Цвет строки
        if not is_active:
            row_color = CLR_INACTIVE
        elif role == "admin":
            row_color = CLR_ADMIN
        elif role == "operator":
            row_color = CLR_OPERATOR
        else:
            row_color = CLR_VIEWER

        base_vals = [
            user.get("id", ""),
            username,
            user.get("full_name") or "",
            user.get("department") or "",
            user.get("job_title") or "",
            user.get("email") or "",
            _role_ru(role),
            "Активен" if is_active else "Отключён",
            _auth_ru(str(user.get("auth_source") or "local")),
            "Да" if user.get("is_2fa_enabled") else "Нет",
            user.get("telegram_id") or "",
            user.get("assigned_database") or "",
            "Да" if use_custom else "Нет",
            _fmt_dt(user.get("created_at")),
            _fmt_dt(user.get("updated_at")),
        ]

        for c, val in enumerate(base_vals, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill = _cell_fill(row_color)
            cell.alignment = _left() if c > 1 else _center()
            cell.border = _border()
            cell.font = Font(size=10)

        for i, perm in enumerate(perm_columns):
            col = FIRST_PERM_COL + i
            has = perm in eff_perms
            if has and use_custom:
                mark = "✓"
                fill_clr = CLR_CUSTOM_YES
            elif has:
                mark = "✓"
                fill_clr = CLR_YES
            else:
                mark = ""
                fill_clr = CLR_NO if is_active else CLR_INACTIVE
            cell = ws.cell(row=row_idx, column=col, value=mark)
            cell.fill = _cell_fill(fill_clr)
            cell.alignment = _center()
            cell.border = _border()
            cell.font = Font(size=10, color="2D7B2D" if has else "CCCCCC")

    # Ширины колонок
    col_widths = [6, 18, 22, 18, 20, 24, 16, 10, 12, 5, 12, 16, 10, 16, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(perm_columns)):
        ws.column_dimensions[get_column_letter(FIRST_PERM_COL + i)].width = 7

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 45
    ws.freeze_panes = "B3"

    # -----------------------------------------------------------------------
    # Лист 2 — Разбивка по ролям
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("По ролям")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 15

    hdrs2 = ["Роль", "Логин", "Полное имя", "Email", "Статус"]
    for c, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = _hdr_font()
        cell.fill = _cell_fill(CLR_HEADER_MAIN)
        cell.alignment = _center()
        cell.border = _border()

    role_order = ["admin", "operator", "viewer"]
    row2 = 2
    for role_key in role_order:
        role_users = [u for u in users
                      if str(u.get("role") or "viewer").lower() == role_key
                      and not str(u.get("username") or "").startswith("__ai_bot__")]
        if not role_users:
            continue
        ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
        cell = ws2.cell(row=row2, column=1, value=f"── {_role_ru(role_key)} ({len(role_users)}) ──")
        cell.font = Font(bold=True, size=11, color="1F3864")
        cell.fill = _cell_fill(CLR_HEADER_PERM if role_key == "admin" else ("D9EAD3" if role_key == "operator" else "EAF4FB"))
        cell.alignment = _center()
        row2 += 1
        for u in role_users:
            is_active = bool(u.get("is_active", True))
            fill_clr = CLR_INACTIVE if not is_active else (CLR_ADMIN if role_key == "admin" else (CLR_OPERATOR if role_key == "operator" else CLR_VIEWER))
            for c, val in enumerate([
                _role_ru(role_key),
                str(u.get("username") or ""),
                str(u.get("full_name") or ""),
                str(u.get("email") or ""),
                "Активен" if is_active else "Отключён",
            ], 1):
                cell = ws2.cell(row=row2, column=c, value=val)
                cell.fill = _cell_fill(fill_clr)
                cell.alignment = _left()
                cell.border = _border()
                cell.font = Font(size=10)
            row2 += 1

    # -----------------------------------------------------------------------
    # Лист 3 — Кастомные права (только пользователи с use_custom_permissions)
    # -----------------------------------------------------------------------
    custom_users = [u for u in users
                    if bool(u.get("use_custom_permissions"))
                    and not str(u.get("username") or "").startswith("__ai_bot__")]
    ws3 = wb.create_sheet("Кастомные права")
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 35

    hdrs3 = ["ID", "Логин", "Полное имя", "Кастомные права (список)"]
    for c, h in enumerate(hdrs3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = _hdr_font()
        cell.fill = _cell_fill(CLR_HEADER_MAIN)
        cell.alignment = _center()
        cell.border = _border()

    if not custom_users:
        ws3.cell(row=2, column=1, value="Нет пользователей с кастомными правами")
    else:
        for ri, u in enumerate(custom_users, 2):
            cp = u.get("custom_permissions") or []
            if isinstance(cp, str):
                try:
                    cp = json.loads(cp)
                except Exception:
                    cp = []
            for c, val in enumerate([
                u.get("id", ""),
                str(u.get("username") or ""),
                str(u.get("full_name") or ""),
                ", ".join(sorted(cp)) if isinstance(cp, list) else str(cp),
            ], 1):
                cell = ws3.cell(row=ri, column=c, value=val)
                cell.fill = _cell_fill(CLR_CUSTOM_YES)
                cell.alignment = _left()
                cell.border = _border()
                cell.font = Font(size=10)

    # -----------------------------------------------------------------------
    # Сохранение
    # -----------------------------------------------------------------------
    wb.save(out_path)
    print(f"\n[OK] Excel сохранён: {out_path}")
    print(f"     Листы: '{ws.title}', '{ws2.title}', '{ws3.title}'")


# ---------------------------------------------------------------------------
# Точка входа
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Экспорт пользователей веб-приложения в Excel")
    parser.add_argument(
        "--out", "-o",
        default=str(PROJECT_ROOT / "exports" / f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
        help="Путь к выходному .xlsx файлу",
    )
    args = parser.parse_args()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    users = load_users()
    # Исключаем системных ботов
    visible = [u for u in users if not str(u.get("username") or "").startswith("__ai_bot__")]
    print(f"[INFO] Найдено пользователей: {len(visible)}")
    if not visible:
        print("[WARN] Список пользователей пуст. Проверьте источник данных.")
        return

    build_excel(visible, out_path)

    roles_stat = {}
    for u in visible:
        r = str(u.get("role") or "viewer")
        roles_stat[r] = roles_stat.get(r, 0) + 1
    print("\n  Разбивка по ролям:")
    for r, cnt in sorted(roles_stat.items()):
        print(f"    {_role_ru(r):20s}: {cnt}")
    active = sum(1 for u in visible if u.get("is_active", True))
    print(f"\n  Активных: {active} / {len(visible)}")


if __name__ == "__main__":
    main()
