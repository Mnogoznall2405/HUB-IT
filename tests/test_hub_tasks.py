from __future__ import annotations

import sys
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
import importlib
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.api import deps
from backend.api.v1 import hub
from backend.models.auth import User

hub_service_module = importlib.import_module("backend.services.hub_service")
access_policy_module = importlib.import_module("backend.services.access_policy_service")


TASKS_READ = "tasks.read"
TASKS_CREATE = "tasks.create"
TASKS_WRITE = "tasks.write"
TASKS_REVIEW = "tasks.review"
DASHBOARD_READ = "dashboard.read"


def _raw_user(
    user_id: int,
    username: str,
    full_name: str,
    role: str,
    permissions: list[str],
) -> dict:
    return {
        "id": user_id,
        "username": username,
        "email": None,
        "full_name": full_name,
        "is_active": True,
        "role": role,
        "permissions": permissions,
        "use_custom_permissions": True,
        "custom_permissions": permissions,
        "auth_source": "local",
        "telegram_id": None,
        "assigned_database": None,
        "mailbox_email": None,
        "mailbox_login": None,
        "mail_profile_mode": "manual",
        "mail_signature_html": None,
        "mail_is_configured": False,
        "created_at": None,
        "updated_at": None,
        "mail_updated_at": None,
    }


def _public_user(raw: dict) -> User:
    permissions = list(raw.get("custom_permissions") or raw.get("permissions") or [])
    return User(
        id=int(raw["id"]),
        username=str(raw["username"]),
        email=None,
        full_name=str(raw["full_name"]),
        role=str(raw["role"]),
        is_active=True,
        permissions=permissions,
        use_custom_permissions=True,
        custom_permissions=permissions,
        auth_source="local",
        telegram_id=None,
        assigned_database=None,
        mailbox_email=None,
        mailbox_login=None,
        mail_profile_mode="manual",
        mail_signature_html=None,
        mail_is_configured=False,
        created_at=None,
        updated_at=None,
        mail_updated_at=None,
    )


@pytest.fixture
def task_env(temp_dir, monkeypatch):
    monkeypatch.setenv("TASK_DISCUSSION_CHAT_ENABLED", "0")
    monkeypatch.setenv("CHAT_MODULE_ENABLED", "0")
    monkeypatch.setenv("TASK_EMAIL_AUTODISPATCH_ENABLED", "0")
    task_discussion_module = importlib.import_module("backend.chat.task_discussion")
    monkeypatch.setattr(task_discussion_module, "is_task_discussion_chat_enabled", lambda: False)
    monkeypatch.setattr(task_discussion_module.config.chat, "task_discussion_enabled", False, raising=False)

    async def _run_in_threadpool_inline(func, *args, **kwargs):
        # TestClient + run_in_threadpool can deadlock on Windows (anyio thread pool).
        return func(*args, **kwargs)

    monkeypatch.setattr(hub, "run_in_threadpool", _run_in_threadpool_inline)

    raw_users = {
        1: _raw_user(1, "author", "Task Author", "operator", [DASHBOARD_READ, TASKS_READ, TASKS_WRITE]),
        2: _raw_user(2, "assignee", "Task Assignee", "viewer", [DASHBOARD_READ, TASKS_READ]),
        3: _raw_user(3, "controller", "Task Controller", "viewer", [DASHBOARD_READ, TASKS_READ, TASKS_REVIEW]),
        4: _raw_user(4, "outsider", "Task Outsider", "viewer", [DASHBOARD_READ, TASKS_READ]),
        5: _raw_user(5, "admin", "Task Admin", "admin", [DASHBOARD_READ, TASKS_READ, TASKS_WRITE, TASKS_REVIEW]),
        6: _raw_user(6, "taskonly", "Task Only User", "viewer", [TASKS_READ]),
        7: _raw_user(7, "assistant", "Task Assistant", "viewer", [TASKS_READ]),
        8: _raw_user(8, "createonly", "Task Create Only", "viewer", [DASHBOARD_READ, TASKS_READ, TASKS_CREATE]),
    }
    users_by_id = dict(raw_users)
    store = SimpleNamespace(
        db_path=str(Path(temp_dir) / "hub_tasks.db"),
        data_dir=str(Path(temp_dir) / "data"),
    )
    Path(store.data_dir).mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(hub_service_module, "get_local_store", lambda: store)
    monkeypatch.setattr(hub_service_module, "is_app_database_configured", lambda: False)
    monkeypatch.setattr(hub_service_module.user_service, "list_users", lambda: list(users_by_id.values()))
    monkeypatch.setattr(hub_service_module.user_service, "get_by_id", lambda user_id: users_by_id.get(int(user_id)))

    service = hub_service_module.HubService()
    monkeypatch.setattr(hub, "hub_service", service)
    monkeypatch.setattr(hub_service_module, "hub_service", service)
    monkeypatch.setattr(service, "_schedule_task_email_outbox_dispatch", lambda: None)

    app = FastAPI()
    app.include_router(hub.router, prefix="/hub")

    current = {"user": _public_user(raw_users[1])}

    def _override_current_user() -> User:
        return current["user"]

    app.dependency_overrides[deps.get_current_active_user] = _override_current_user
    client = TestClient(app)

    def set_user(user_id: int) -> None:
        current["user"] = _public_user(raw_users[user_id])

    yield {
        "client": client,
        "set_user": set_user,
        "raw_users": raw_users,
        "service": service,
    }

    client.close()


def _create_task(
    client: TestClient,
    *,
    assignee_user_id: int = 2,
    controller_user_id: int = 3,
    title: str = "Task Alpha",
    due_at: str | None = None,
    email_deadline_remind_hours: int | None = None,
    observer_user_ids: list[int] | None = None,
) -> dict:
    project_code = f"{title.lower().replace(' ', '-')}-{uuid4().hex[:8]}"
    project_response = client.post(
        "/hub/task-projects",
        json={"name": f"Project for {title} {project_code[-4:]}", "code": project_code},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()
    response = client.post(
        "/hub/tasks",
        json={
            "title": title,
            "description": "Task body",
            "assignee_user_ids": [assignee_user_id],
            "controller_user_id": controller_user_id,
            "project_id": project["id"],
            "protocol_date": "2026-03-01",
            "priority": "high",
            **({"due_at": due_at} if due_at else {}),
            **({"email_deadline_remind_hours": email_deadline_remind_hours} if email_deadline_remind_hours is not None else {}),
            **({"observer_user_ids": observer_user_ids} if observer_user_ids else {}),
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["created"] == 1
    return payload["items"][0]


def _submit_task(client: TestClient, task_id: str, *, comment: str = "Report done") -> dict:
    response = client.post(
        f"/hub/tasks/{task_id}/submit",
        data={"comment": comment},
        files={"file": ("report.txt", b"report-bytes", "text/plain")},
    )
    assert response.status_code == 200, response.text
    return response.json()


def _approve_task(client: TestClient, task_id: str, *, comment: str = "Approved") -> dict:
    response = client.post(
        f"/hub/tasks/{task_id}/review",
        json={"decision": "approve", "comment": comment},
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_reopen_completed_task_by_participant(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    task = _create_task(client, title="Reopen Me")
    task_id = task["id"]

    set_user(2)
    _submit_task(client, task_id, comment="Done once")

    set_user(3)
    approved = _approve_task(client, task_id)
    assert approved["status"] == "done"

    set_user(2)
    detail = client.get(f"/hub/tasks/{task_id}")
    assert detail.status_code == 200
    assert detail.json()["capabilities"]["can_reopen"] is True

    reopened = client.post(f"/hub/tasks/{task_id}/reopen", json={"due_at": "2026-07-01T19:00"})
    assert reopened.status_code == 200, reopened.text
    reopened_payload = reopened.json()
    assert reopened_payload["status"] == "in_progress"
    assert reopened_payload["due_at"] is not None
    assert reopened_payload["capabilities"]["can_reopen"] is False
    assert reopened_payload["capabilities"]["can_submit"] is True

    status_log = client.get(f"/hub/tasks/{task_id}/status-log")
    assert status_log.status_code == 200
    transitions = [(row["old_status"], row["new_status"]) for row in status_log.json()["items"]]
    assert ("done", "in_progress") in transitions


def test_reopen_completed_task_denied_for_outsider(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    task = _create_task(client, title="Protected Reopen")
    task_id = task["id"]

    set_user(2)
    _submit_task(client, task_id)

    set_user(3)
    _approve_task(client, task_id)

    set_user(4)
    denied = client.post(f"/hub/tasks/{task_id}/reopen")
    assert denied.status_code == 403


def test_reopen_non_completed_task_returns_400(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    task = _create_task(client, title="Still Open")
    task_id = task["id"]

    set_user(2)
    bad = client.post(f"/hub/tasks/{task_id}/reopen")
    assert bad.status_code == 400


def test_reopen_completed_task_queues_email_notification(task_env, monkeypatch):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]
    raw_users = task_env["raw_users"]

    raw_users[1]["mailbox_email"] = "creator@example.test"
    raw_users[3]["mailbox_email"] = "controller@example.test"
    monkeypatch.setenv("TASK_EMAIL_NOTIFICATIONS_ENABLED", "1")
    monkeypatch.setenv("TASK_EMAIL_AUTODISPATCH_ENABLED", "0")
    monkeypatch.setenv("TASK_EMAIL_APP_URL", "https://hub.example.test")

    set_user(1)
    task = _create_task(client, title="Reopen Email Task")
    task_id = task["id"]

    set_user(2)
    _submit_task(client, task_id)

    set_user(3)
    _approve_task(client, task_id)

    set_user(2)
    reopened = client.post(
        f"/hub/tasks/{task_id}/reopen",
        json={"due_at": "2026-08-15T19:00"},
    )
    assert reopened.status_code == 200, reopened.text

    rows = [row for row in _task_email_rows(service) if row.get("event_type") == "task.reopened"]
    assert len(rows) >= 1
    assert {row["recipient_email"] for row in rows} & {"creator@example.test", "controller@example.test"}
    assert all("Reopen Email Task" in row["subject"] for row in rows)
    assert all("возвращено в работу" in row["body_text"].lower() for row in rows)
    assert all(f"/tasks?task={task_id}" in row["body_text"] for row in rows)


def _task_email_rows(service) -> list[dict]:
    with service._lock, service._connect() as conn:
        rows = conn.execute(
            f"SELECT * FROM {service._TASK_EMAIL_OUTBOX_TABLE} ORDER BY created_at ASC, id ASC"
        ).fetchall()
    return [dict(row) for row in rows]


def test_create_task_detail_access_and_role_scopes(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, title="Task Access")
    task_id = created["id"]

    set_user(3)
    controller_view = client.get("/hub/tasks", params={"scope": "my", "role_scope": "controller"})
    assert controller_view.status_code == 200
    assert controller_view.json()["total"] == 1

    both_view = client.get("/hub/tasks", params={"scope": "my", "role_scope": "both"})
    assert both_view.status_code == 200
    assert both_view.json()["total"] == 1

    detail_response = client.get(f"/hub/tasks/{task_id}")
    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert detail_payload["controller_user_id"] == 3
    assert detail_payload["created_by_user_id"] == 1
    assert "attachments_count" in detail_payload
    assert detail_payload["capabilities"] == {
        "can_edit": False,
        "can_start": False,
        "can_submit": False,
        "can_review": False,
        "can_reopen": False,
        "can_upload_files": True,
        "can_update_checklist": True,
        "can_open_discussion": False,
    }

    set_user(4)
    denied = client.get(f"/hub/tasks/{task_id}")
    assert denied.status_code == 403


def test_create_only_user_can_create_task_without_taxonomy_management(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(8)
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Create Only Project", "code": f"create-only-{uuid4().hex[:8]}"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    created = client.post(
        "/hub/tasks",
        json={
            "title": "Create Only Task",
            "description": "Viewer role can create tasks",
            "assignee_user_ids": [8],
            "controller_user_id": None,
            "project_id": project["id"],
            "protocol_date": "2026-03-01",
            "priority": "high",
            "checklist_items": [
                {"id": "step-1", "text": "Collect inventory", "done": False},
                {"id": "step-2", "text": "Confirm result", "done": True},
            ],
        },
    )
    assert created.status_code == 200, created.text
    payload = created.json()
    assert payload["created"] == 1
    assert payload["items"][0]["controller_user_id"] == 0
    assert payload["items"][0]["controller_full_name"] == ""
    task = payload["items"][0]
    assert task["created_by_user_id"] == 8
    assert task["assignee_user_id"] == 8
    assert task["project_id"] == project["id"]
    assert task["priority"] == "high"
    assert task["checklist_total"] == 2
    assert task["checklist_done"] == 1
    assert task["checklist_items"][0]["text"] == "Collect inventory"


def test_create_task_without_department_allows_cross_assignee(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(8)
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Cross Assignee Project", "code": f"cross-assignee-{uuid4().hex[:8]}"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    created = client.post(
        "/hub/tasks",
        json={
            "title": "Cross Assignee Task",
            "description": "No department required",
            "assignee_user_ids": [2],
            "controller_user_id": None,
            "project_id": project["id"],
            "protocol_date": "2026-03-01",
            "priority": "normal",
            "department_id": None,
            "visibility_scope": "private",
        },
    )
    assert created.status_code == 200, created.text
    payload = created.json()
    assert payload["created"] == 1
    task = payload["items"][0]
    assert task["assignee_user_id"] == 2
    assert task["created_by_user_id"] == 8
    assert not task.get("department_id")
    assert task.get("visibility_scope") == "private"


def test_assignee_can_update_task_checklist_only(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    task = _create_task(client, title="Checklist Access")
    seed_response = client.patch(
        f"/hub/tasks/{task['id']}",
        json={"checklist_items": [{"id": "step-1", "text": "Check workstation", "done": False}]},
    )
    assert seed_response.status_code == 200, seed_response.text
    assert seed_response.json()["checklist_done"] == 0

    clear_controller = client.patch(
        f"/hub/tasks/{task['id']}",
        json={"controller_user_id": None},
    )
    assert clear_controller.status_code == 200, clear_controller.text
    assert clear_controller.json()["controller_user_id"] == 0
    assert clear_controller.json()["controller_full_name"] == ""

    set_user(2)
    updated = client.patch(
        f"/hub/tasks/{task['id']}",
        json={"checklist_items": [{"id": "step-1", "text": "Check workstation", "done": True}]},
    )
    assert updated.status_code == 200, updated.text
    payload = updated.json()
    assert payload["checklist_done"] == 1
    assert payload["checklist_items"][0]["done"] is True

    denied = client.patch(
        f"/hub/tasks/{task['id']}",
        json={"title": "Blocked rename"},
    )
    assert denied.status_code == 403


def test_existing_tasks_without_project_are_backfilled_to_general_project(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]

    set_user(1)
    created = _create_task(client, title="Legacy Project Backfill")
    task_id = created["id"]

    with service._connect() as conn:
        conn.execute(f"UPDATE {service._TASKS_TABLE} SET project_id = NULL WHERE id = ?", (task_id,))
        conn.commit()

    service._ensure_schema()

    task = service.get_task(task_id, user_id=1)
    assert task is not None
    assert task["project_id"] == service._DEFAULT_TASK_PROJECT_ID
    assert task["project_name"] == service._DEFAULT_TASK_PROJECT_NAME

    projects = service.list_task_projects(include_inactive=True)
    assert any(item["id"] == service._DEFAULT_TASK_PROJECT_ID for item in projects)


def test_opening_task_marks_task_notifications_read(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, assignee_user_id=2, title="Read On Open")
    task_id = created["id"]

    set_user(2)
    unread_before = client.get("/hub/notifications/unread-counts")
    assert unread_before.status_code == 200
    assert unread_before.json()["notifications_unread_total"] == 1

    detail_response = client.get(f"/hub/tasks/{task_id}")
    assert detail_response.status_code == 200

    unread_after = client.get("/hub/notifications/unread-counts")
    assert unread_after.status_code == 200
    assert unread_after.json()["notifications_unread_total"] == 0


def test_task_only_user_can_access_notification_endpoints(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    _create_task(client, assignee_user_id=6, title="Task Only Notification")

    set_user(6)
    unread_counts = client.get("/hub/notifications/unread-counts")
    assert unread_counts.status_code == 200
    assert unread_counts.json()["notifications_unread_total"] == 1

    notification_poll = client.get("/hub/notifications/poll", params={"limit": 20})
    assert notification_poll.status_code == 200
    items = notification_poll.json()["items"]
    assert any(item["entity_type"] == "task" for item in items)


def test_author_can_review_and_outsider_cannot_access_task_artifacts(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]
    raw_users = task_env["raw_users"]

    set_user(1)
    created = _create_task(client, title="Author Review")
    task_id = created["id"]
    deletable = _create_task(client, title="Author Delete")

    attachment_response = client.post(
        f"/hub/tasks/{task_id}/attachments",
        files={"file": ("spec.txt", b"spec-bytes", "text/plain")},
    )
    assert attachment_response.status_code == 200, attachment_response.text
    attachment_id = attachment_response.json()["id"]

    set_user(2)
    submitted = _submit_task(client, task_id, comment="Need approval")
    report_id = submitted["latest_report"]["id"]
    assert submitted["status"] == "review"

    set_user(1)
    reviewed = client.post(
        f"/hub/tasks/{task_id}/review",
        json={"decision": "approve", "comment": "Accepted by author"},
    )
    assert reviewed.status_code == 200, reviewed.text
    reviewed_payload = reviewed.json()
    assert reviewed_payload["status"] == "done"
    assert reviewed_payload["reviewer_user_id"] == 1
    assert reviewed_payload["reviewer_full_name"] == "Task Author"

    raw_users[1]["permissions"] = [DASHBOARD_READ, TASKS_READ]
    raw_users[1]["custom_permissions"] = [DASHBOARD_READ, TASKS_READ]
    set_user(1)
    assignee_users = client.get("/hub/users/assignees", params={"q": "author"})
    assert assignee_users.status_code == 200, assignee_users.text
    assignee_payload = assignee_users.json()
    assert isinstance(assignee_payload.get("items"), list)
    assert "total" in assignee_payload
    assignee_by_id = client.get("/hub/users/assignees", params={"ids": "4"})
    assert assignee_by_id.status_code == 200, assignee_by_id.text
    assert len(assignee_by_id.json().get("items") or []) >= 1
    controller_users = client.get("/hub/users/controllers")
    assert controller_users.status_code == 200, controller_users.text
    author_update = client.patch(f"/hub/tasks/{task_id}", json={"title": "Author Review Updated"})
    assert author_update.status_code == 200, author_update.text
    assert author_update.json()["title"] == "Author Review Updated"
    author_delete = client.delete(f"/hub/tasks/{deletable['id']}")
    assert author_delete.status_code == 200, author_delete.text

    comments_response = client.post(f"/hub/tasks/{task_id}/comments", json={"body": "Looks good"})
    assert comments_response.status_code == 200

    history_response = client.get(f"/hub/tasks/{task_id}/status-log")
    assert history_response.status_code == 200
    transitions = [item["new_status"] for item in history_response.json()["items"]]
    assert "review" in transitions
    assert "done" in transitions

    set_user(4)
    assert client.patch(f"/hub/tasks/{task_id}", json={"title": "Intrusion"}).status_code == 403
    assert client.post(f"/hub/tasks/{task_id}/comments", json={"body": "I should not see this"}).status_code == 403
    assert client.get(f"/hub/tasks/{task_id}/comments").status_code == 403
    assert client.get(f"/hub/tasks/{task_id}/status-log").status_code == 403
    assert client.get(f"/hub/tasks/reports/{report_id}/file").status_code == 403
    assert client.get(f"/hub/tasks/{task_id}/attachments/{attachment_id}/file").status_code == 403


def test_controller_review_and_counts_include_author_and_controller_roles(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    review_task = _create_task(client, title="Review Required")
    open_task = _create_task(client, title="Still Open")

    set_user(2)
    submitted = _submit_task(client, review_task["id"], comment="Ready for controller")
    assert submitted["status"] == "review"

    set_user(3)
    reviewed = client.post(
        f"/hub/tasks/{review_task['id']}/review",
        json={"decision": "reject", "comment": "Controller rejected"},
    )
    assert reviewed.status_code == 200, reviewed.text
    reviewed_payload = reviewed.json()
    assert reviewed_payload["status"] == "in_progress"
    assert reviewed_payload["reviewer_user_id"] == 3
    assert reviewed_payload["reviewer_full_name"] == "Task Controller"

    set_user(1)
    counts = client.get("/hub/notifications/unread-counts")
    assert counts.status_code == 200
    counts_payload = counts.json()
    assert counts_payload["tasks_created_open"] == 2
    assert counts_payload["tasks_open_total"] == 2
    assert counts_payload["tasks_open"] == 2
    assert counts_payload["tasks_review_required"] == 0

    dashboard = client.get("/hub/dashboard", params={"tasks_limit": 10, "announcements_limit": 5})
    assert dashboard.status_code == 200
    assert dashboard.json()["my_tasks"]["total"] == 2

    set_user(3)
    controller_counts = client.get("/hub/notifications/unread-counts")
    assert controller_counts.status_code == 200
    controller_payload = controller_counts.json()
    assert controller_payload["tasks_controller_open"] == 2
    assert controller_payload["tasks_open_total"] == 2

    set_user(4)
    denied_review = client.post(
        f"/hub/tasks/{open_task['id']}/review",
        json={"decision": "approve", "comment": "Outsider"},
    )
    assert denied_review.status_code == 403


def test_compute_unread_counts_sql_task_metrics(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]

    set_user(1)
    overdue_task = _create_task(client, title="Overdue SQL", due_at="2020-01-01T12:00:00+00:00")
    review_task = _create_task(client, title="Review SQL")
    done_overdue = _create_task(client, title="Done Overdue SQL", due_at="2020-01-01T12:00:00+00:00")

    set_user(2)
    assert _submit_task(client, review_task["id"], comment="Ready")["status"] == "review"
    assert _submit_task(client, done_overdue["id"], comment="Finished")["status"] == "review"

    set_user(3)
    approved = client.post(
        f"/hub/tasks/{done_overdue['id']}/review",
        json={"decision": "approve", "comment": "Closed"},
    )
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "done"

    set_user(1)
    counts = service.get_unread_counts(user_id=1)
    assert counts["tasks_created_open"] == 2
    assert counts["tasks_open_total"] == 2
    assert counts["tasks_overdue"] == 1
    assert counts["tasks_review_required"] == 1

    set_user(2)
    comment_response = client.post(
        f"/hub/tasks/{overdue_task['id']}/comments",
        json={"body": "Overdue ping"},
    )
    assert comment_response.status_code == 200, comment_response.text

    set_user(1)
    assert service.get_unread_counts(user_id=1)["tasks_with_unread_comments"] == 1
    mark_seen = client.post(f"/hub/tasks/{overdue_task['id']}/comments/mark-seen")
    assert mark_seen.status_code == 200, mark_seen.text
    assert service.get_unread_counts(user_id=1)["tasks_with_unread_comments"] == 0


def test_assignee_cannot_submit_task_twice_while_waiting_for_review(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, title="Single Submit")

    set_user(2)
    first_submit = _submit_task(client, created["id"], comment="Initial report")
    assert first_submit["status"] == "review"

    second_submit = client.post(
        f"/hub/tasks/{created['id']}/submit",
        data={"comment": "Second report"},
        files={"file": ("report-2.txt", b"second-report", "text/plain")},
    )
    assert second_submit.status_code == 400
    assert second_submit.json()["detail"] == "Task is already waiting for review"


def test_task_comment_summary_unread_seen_and_notifications(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, title="Comment Visibility")
    task_id = created["id"]

    set_user(2)
    comment_response = client.post(
        f"/hub/tasks/{task_id}/comments",
        json={"body": "Need update\nline 2"},
    )
    assert comment_response.status_code == 200, comment_response.text

    assignee_tasks = client.get("/hub/tasks", params={"scope": "my", "role_scope": "assignee"})
    assert assignee_tasks.status_code == 200
    assignee_item = assignee_tasks.json()["items"][0]
    assert assignee_item["comments_count"] == 1
    assert assignee_item["latest_comment_preview"] == "Need update line 2"
    assert assignee_item["latest_comment_user_id"] == 2
    assert assignee_item["latest_comment_full_name"] == "Task Assignee"
    assert assignee_item["has_unread_comments"] is False

    assignee_poll = client.get("/hub/notifications/poll", params={"limit": 20})
    assert assignee_poll.status_code == 200
    assert not any(
        item["event_type"] == "task.comment_added" and item["entity_id"] == task_id
        for item in assignee_poll.json()["items"]
    )

    set_user(1)
    author_detail = client.get(f"/hub/tasks/{task_id}")
    assert author_detail.status_code == 200
    author_payload = author_detail.json()
    assert author_payload["comments_count"] == 1
    assert author_payload["latest_comment_preview"] == "Need update line 2"
    assert author_payload["latest_comment_user_id"] == 2
    assert author_payload["latest_comment_full_name"] == "Task Assignee"
    assert author_payload["has_unread_comments"] is True

    author_counts_with_comment = client.get("/hub/notifications/unread-counts")
    assert author_counts_with_comment.status_code == 200
    assert author_counts_with_comment.json()["tasks_with_unread_comments"] == 1

    author_poll = client.get("/hub/notifications/poll", params={"limit": 20})
    assert author_poll.status_code == 200
    assert not any(
        item["event_type"] == "task.comment_added" and item["entity_id"] == task_id and item["unread"] == 1
        for item in author_poll.json()["items"]
    )
    author_counts_before = client.get("/hub/notifications/unread-counts")
    assert author_counts_before.status_code == 200
    assert author_counts_before.json()["notifications_unread_total"] == 0

    mark_seen = client.post(f"/hub/tasks/{task_id}/comments/mark-seen")
    assert mark_seen.status_code == 200, mark_seen.text
    assert mark_seen.json()["has_unread_comments"] is False

    author_detail_after = client.get(f"/hub/tasks/{task_id}")
    assert author_detail_after.status_code == 200
    assert author_detail_after.json()["has_unread_comments"] is False

    author_counts_after = client.get("/hub/notifications/unread-counts")
    assert author_counts_after.status_code == 200
    assert author_counts_after.json()["notifications_unread_total"] == 0
    assert author_counts_after.json()["tasks_with_unread_comments"] == 0

    set_user(3)
    controller_tasks = client.get("/hub/tasks", params={"scope": "my", "role_scope": "controller"})
    assert controller_tasks.status_code == 200
    controller_item = controller_tasks.json()["items"][0]
    assert controller_item["comments_count"] == 1
    assert controller_item["has_unread_comments"] is True

    set_user(4)
    assert client.post(f"/hub/tasks/{task_id}/comments/mark-seen").status_code == 403


def test_notifications_poll_unread_only_excludes_read_items(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, title="Unread Only")
    task_id = created["id"]

    set_user(2)
    comment_response = client.post(
        f"/hub/tasks/{task_id}/comments",
        json={"body": "Fresh comment"},
    )
    assert comment_response.status_code == 200, comment_response.text

    set_user(1)
    unread_poll = client.get("/hub/notifications/poll", params={"limit": 20, "unread_only": True})
    assert unread_poll.status_code == 200, unread_poll.text
    unread_items = unread_poll.json()["items"]
    target_item = next(
        item for item in unread_items
        if item["event_type"] == "task.comment_added" and item["entity_id"] == task_id
    )
    assert target_item["unread"] == 1

    mark_read = client.post(f"/hub/notifications/{target_item['id']}/read")
    assert mark_read.status_code == 200, mark_read.text

    unread_poll_after = client.get("/hub/notifications/poll", params={"limit": 20, "unread_only": True})
    assert unread_poll_after.status_code == 200, unread_poll_after.text
    assert not any(item["id"] == target_item["id"] for item in unread_poll_after.json()["items"])


def test_notifications_read_all_marks_everything_read(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    first_task = _create_task(client, title="Read all 1")
    second_task = _create_task(client, title="Read all 2")

    set_user(2)
    for task_id in (first_task["id"], second_task["id"]):
        comment_response = client.post(
            f"/hub/tasks/{task_id}/comments",
            json={"body": f"Update for {task_id}"},
        )
        assert comment_response.status_code == 200, comment_response.text

    set_user(1)
    unread_before = client.get("/hub/notifications/unread-counts")
    assert unread_before.status_code == 200
    assert unread_before.json()["notifications_unread_total"] >= 2

    read_all = client.post("/hub/notifications/read-all")
    assert read_all.status_code == 200, read_all.text
    assert read_all.json()["marked_count"] >= 2

    unread_after = client.get("/hub/notifications/unread-counts")
    assert unread_after.status_code == 200
    assert unread_after.json()["notifications_unread_total"] == 0

    unread_poll_after = client.get("/hub/notifications/poll", params={"limit": 20, "unread_only": True})
    assert unread_poll_after.status_code == 200, unread_poll_after.text
    assert unread_poll_after.json()["items"] == []


def test_task_projects_protocol_date_and_analytics(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Проект Север", "code": "north"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    object_response = client.post(
        "/hub/task-objects",
        json={"project_id": project["id"], "name": "Объект 17", "code": "obj-17"},
    )
    assert object_response.status_code == 200, object_response.text
    task_object = object_response.json()

    missing_project = client.post(
        "/hub/tasks",
        json={
            "title": "No Project",
            "description": "Should fail",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
            "protocol_date": "2026-03-01",
        },
    )
    assert missing_project.status_code == 400
    assert missing_project.json()["detail"] == "project_id is required"

    created = client.post(
        "/hub/tasks",
        json={
            "title": "Task With Project",
            "description": "Task body",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
                "project_id": project["id"],
                "object_id": task_object["id"],
                "protocol_date": "2026-03-01",
                "due_at": "2030-04-10T12:00:00+00:00",
                "priority": "high",
            },
        )
    assert created.status_code == 200, created.text
    task = created.json()["items"][0]
    assert task["project_id"] == project["id"]
    assert task["project_name"] == "Проект Север"
    assert task["object_id"] == task_object["id"]
    assert task["object_name"] == "Объект 17"
    assert task["protocol_date"] == "2026-03-01"

    set_user(2)
    submit = client.post(
        f"/hub/tasks/{task['id']}/submit",
        data={"comment": "Ready"},
        files={"file": ("report.txt", b"ok", "text/plain")},
    )
    assert submit.status_code == 200, submit.text

    set_user(3)
    review = client.post(
        f"/hub/tasks/{task['id']}/review",
        json={"decision": "approve", "comment": "Approved"},
    )
    assert review.status_code == 200, review.text
    reviewed = review.json()
    assert reviewed["status"] == "done"
    assert reviewed["completed_at"]
    assert reviewed["completed_on_time"] is True

    set_user(1)
    analytics = client.get(
        "/hub/tasks/analytics",
        params={
            "start_date": "2026-03-01",
            "end_date": "2026-03-31",
            "date_basis": "protocol_date",
            "project_id": project["id"],
        },
    )
    assert analytics.status_code == 200, analytics.text
    payload = analytics.json()
    assert payload["summary"]["total"] == 1
    assert payload["summary"]["done"] == 1
    assert payload["summary"]["done_on_time"] == 1
    assert payload["summary"]["completion_percent"] == 100.0
    assert payload["summary"]["completion_on_time_percent"] == 100.0
    assert payload["by_project"][0]["project_id"] == project["id"]
    assert payload["by_project"][0]["project_name"] == "Проект Север"
    assert payload["by_object"][0]["object_id"] == task_object["id"]
    assert payload["by_object"][0]["object_name"] == "Объект 17"
    assert payload["by_participant"][0]["participant_user_id"] == 2


def test_task_analytics_backfills_completed_at_and_returns_extended_metrics(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]

    now_utc = datetime.now(timezone.utc)
    protocol_day = now_utc.date().isoformat()
    on_time_due = (now_utc + timedelta(days=10)).isoformat()
    overdue_due = (now_utc - timedelta(days=2)).isoformat()
    legacy_completed_at = (now_utc - timedelta(days=1)).isoformat()

    set_user(1)
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Проект Аналитика", "code": "analytics"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    object_response = client.post(
        "/hub/task-objects",
        json={"project_id": project["id"], "name": "Объект А1", "code": "a1"},
    )
    assert object_response.status_code == 200, object_response.text
    task_object = object_response.json()

    on_time_created = client.post(
        "/hub/tasks",
        json={
            "title": "On Time Task",
            "description": "Done on time",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
            "project_id": project["id"],
            "object_id": task_object["id"],
            "protocol_date": protocol_day,
            "due_at": on_time_due,
        },
    )
    assert on_time_created.status_code == 200, on_time_created.text
    on_time_task = on_time_created.json()["items"][0]

    set_user(2)
    submit = client.post(
        f"/hub/tasks/{on_time_task['id']}/submit",
        data={"comment": "Ready"},
        files={"file": ("report.txt", b"ok", "text/plain")},
    )
    assert submit.status_code == 200, submit.text

    set_user(3)
    review = client.post(
        f"/hub/tasks/{on_time_task['id']}/review",
        json={"decision": "approve", "comment": "Approved"},
    )
    assert review.status_code == 200, review.text

    set_user(1)
    legacy_created = client.post(
        "/hub/tasks",
        json={
            "title": "Legacy Done Without Due",
            "description": "Legacy done task",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
            "project_id": project["id"],
            "object_id": task_object["id"],
            "protocol_date": protocol_day,
        },
    )
    assert legacy_created.status_code == 200, legacy_created.text
    legacy_task = legacy_created.json()["items"][0]

    overdue_created = client.post(
        "/hub/tasks",
        json={
            "title": "Overdue Task",
            "description": "Still open",
            "assignee_user_ids": [6],
            "controller_user_id": 3,
            "project_id": project["id"],
            "protocol_date": protocol_day,
            "due_at": overdue_due,
        },
    )
    assert overdue_created.status_code == 200, overdue_created.text
    overdue_task = overdue_created.json()["items"][0]

    with service._connect() as conn:
        conn.execute(
            f"""
            UPDATE {service._TASKS_TABLE}
            SET status = 'done',
                reviewed_at = ?,
                completed_at = NULL,
                completed_at_source = NULL,
                due_at = NULL,
                updated_at = ?
            WHERE id = ?
            """,
            (legacy_completed_at, legacy_completed_at, legacy_task["id"]),
        )
        conn.commit()

    service._ensure_schema()

    legacy_detail = service.get_task(legacy_task["id"], user_id=1)
    assert legacy_detail is not None
    assert legacy_detail["completed_at"] == legacy_completed_at
    assert legacy_detail["completed_at_source"] == "reviewed_at"
    assert legacy_detail["done_without_due"] is True

    analytics = client.get(
        "/hub/tasks/analytics",
        params={
            "start_date": protocol_day,
            "end_date": protocol_day,
            "date_basis": "protocol_date",
            "project_id": project["id"],
        },
    )
    assert analytics.status_code == 200, analytics.text
    payload = analytics.json()
    assert payload["summary"]["total"] == 3
    assert payload["summary"]["open"] == 1
    assert payload["summary"]["new"] == 1
    assert payload["summary"]["in_progress"] == 0
    assert payload["summary"]["review"] == 0
    assert payload["summary"]["done"] == 2
    assert payload["summary"]["done_on_time"] == 1
    assert payload["summary"]["done_without_due"] == 1
    assert payload["summary"]["overdue"] == 1
    assert payload["summary"]["with_due_total"] == 2
    assert payload["summary"]["completion_percent"] == pytest.approx(66.67, abs=0.01)
    assert payload["summary"]["completion_on_time_percent"] == pytest.approx(50.0, abs=0.01)
    assert payload["trend"]["granularity"] == "day"
    assert payload["trend"]["items"]
    assert {item["status"]: item["value"] for item in payload["status_breakdown"]} == {
        "new": 1,
        "in_progress": 0,
        "review": 0,
        "done": 2,
    }

    participant_filtered = client.get(
        "/hub/tasks/analytics",
        params={
            "start_date": protocol_day,
            "end_date": protocol_day,
            "date_basis": "protocol_date",
            "project_id": project["id"],
            "object_id": task_object["id"],
            "participant_user_id": 2,
        },
    )
    assert participant_filtered.status_code == 200, participant_filtered.text
    filtered_payload = participant_filtered.json()
    assert filtered_payload["summary"]["total"] == 2
    assert filtered_payload["summary"]["done"] == 2
    assert filtered_payload["summary"]["done_without_due"] == 1
    assert filtered_payload["by_participant"][0]["participant_user_id"] == 2
    assert filtered_payload["by_object"][0]["object_id"] == task_object["id"]

    overdue_detail = service.get_task(overdue_task["id"], user_id=1)
    assert overdue_detail is not None
    assert overdue_detail["is_overdue"] is True


def test_task_analytics_export_returns_excel(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Проект Экспорт", "code": "export"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    object_response = client.post(
        "/hub/task-objects",
        json={"project_id": project["id"], "name": "Объект Экспорт", "code": "exp-1"},
    )
    assert object_response.status_code == 200, object_response.text
    task_object = object_response.json()

    created = client.post(
        "/hub/tasks",
        json={
            "title": "Exported Task",
            "description": "Task for excel export",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
            "project_id": project["id"],
            "object_id": task_object["id"],
            "protocol_date": "2026-03-10",
            "due_at": "2026-03-20T10:00:00+00:00",
        },
    )
    assert created.status_code == 200, created.text
    task = created.json()["items"][0]

    set_user(2)
    submit = client.post(
        f"/hub/tasks/{task['id']}/submit",
        data={"comment": "Ready"},
        files={"file": ("report.txt", b"ok", "text/plain")},
    )
    assert submit.status_code == 200, submit.text

    set_user(3)
    review = client.post(
        f"/hub/tasks/{task['id']}/review",
        json={"decision": "approve", "comment": "Approved"},
    )
    assert review.status_code == 200, review.text

    set_user(1)
    response = client.get(
        "/hub/tasks/analytics/export",
        params={
            "start_date": "2026-03-01",
            "end_date": "2026-03-31",
            "date_basis": "protocol_date",
            "project_id": project["id"],
            "object_id": task_object["id"],
            "participant_user_id": 2,
        },
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=" in response.headers.get("content-disposition", "")

    workbook = load_workbook(filename=BytesIO(response.content))
    assert workbook.sheetnames == ["Сводка", "По участникам", "По проектам", "По объектам", "Тренд"]

    summary_sheet = workbook["Сводка"]
    summary_rows = list(summary_sheet.iter_rows(values_only=True))
    flattened = [value for row in summary_rows for value in row if value not in (None, "")]
    assert "Аналитика задач" in flattened
    assert "Фильтры отчёта" in flattened
    assert "Ключевые показатели" in flattened

    assert workbook["По участникам"].max_row >= 1
    assert workbook["По проектам"].max_row >= 1
    assert workbook["По объектам"].max_row >= 1
    assert workbook["Тренд"].max_row >= 1


def test_delegate_receives_task_notifications_and_read_only_access(task_env, monkeypatch):
    client = task_env["client"]
    set_user = task_env["set_user"]

    monkeypatch.setattr(hub_service_module.user_service, "get_delegate_user_ids", lambda owner_user_id, active_only=True: [7] if int(owner_user_id) == 2 else [])
    monkeypatch.setattr(hub_service_module.user_service, "get_delegate_owner_ids", lambda delegate_user_id, active_only=True: [2] if int(delegate_user_id) == 7 else [])

    set_user(1)
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Проект Делегат", "code": "delegate"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    created = client.post(
        "/hub/tasks",
        json={
            "title": "Delegate Visible Task",
            "description": "Task body",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
            "project_id": project["id"],
            "protocol_date": "2026-03-05",
        },
    )
    assert created.status_code == 200, created.text
    task = created.json()["items"][0]

    set_user(7)
    unread = client.get("/hub/notifications/unread-counts")
    assert unread.status_code == 200
    assert unread.json()["notifications_unread_total"] == 1

    polled = client.get("/hub/notifications/poll", params={"limit": 20, "unread_only": True})
    assert polled.status_code == 200, polled.text
    items = polled.json()["items"]
    assert any(item["event_type"] == "task.assigned" and item["entity_id"] == task["id"] for item in items)

    my_tasks = client.get("/hub/tasks", params={"scope": "my", "role_scope": "assignee"})
    assert my_tasks.status_code == 200, my_tasks.text
    assert any(item["id"] == task["id"] for item in my_tasks.json()["items"])

    detail = client.get(f"/hub/tasks/{task['id']}")
    assert detail.status_code == 200, detail.text
    assert detail.json()["id"] == task["id"]

    update = client.patch(f"/hub/tasks/{task['id']}", json={"title": "Delegate Edit"})
    assert update.status_code == 403


def test_task_email_outbox_queues_assignment_and_dispatches(task_env, monkeypatch):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]
    raw_users = task_env["raw_users"]

    raw_users[2]["mailbox_email"] = "assignee@example.test"
    raw_users[3]["email"] = "controller@example.test"
    monkeypatch.setenv("TASK_EMAIL_NOTIFICATIONS_ENABLED", "1")
    monkeypatch.setenv("TASK_EMAIL_AUTODISPATCH_ENABLED", "0")
    monkeypatch.setenv("TASK_EMAIL_APP_URL", "https://hub.example.test")

    set_user(1)
    task = _create_task(client, title="Email Assigned Task")

    rows = _task_email_rows(service)
    assert {row["event_type"] for row in rows} == {"task.assigned", "task.controller_assigned"}
    assert {row["recipient_email"] for row in rows} == {"assignee@example.test", "controller@example.test"}
    assert all(row["status"] == "pending" for row in rows)
    assert all(f"/tasks?task={task['id']}" in row["body_text"] for row in rows)

    sent: list[dict] = []
    monkeypatch.setattr(
        hub_service_module.task_email_service,
        "send_task_email",
        lambda **kwargs: sent.append(kwargs) or True,
    )

    result = service.dispatch_task_email_outbox()

    assert result == {"claimed": 2, "sent": 2, "failed": 0}
    assert len(sent) == 2
    assert {item["recipient_email"] for item in sent} == {"assignee@example.test", "controller@example.test"}
    assert all("Email Assigned Task" in item["subject"] for item in sent)
    assert {row["status"] for row in _task_email_rows(service)} == {"sent"}


def test_read_paths_skip_inline_task_due_ensure(task_env, monkeypatch):
    service = task_env["service"]
    calls: list[int] = []

    def _track(conn, *, user_id: int) -> None:
        calls.append(int(user_id))

    monkeypatch.setattr(service, "_maybe_ensure_task_due_notifications_for_user", _track)
    service.get_unread_counts(user_id=2)
    service.poll_notifications(user_id=2, limit=20)
    assert calls == []


def test_list_tasks_returns_explicit_list_columns(task_env):
    client = task_env["client"]
    service = task_env["service"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, title="List Columns Task")
    detail = service.get_task(created["id"], user_id=1, is_admin=False)
    listed = service.list_tasks(user_id=1, scope="my", role_scope="both", limit=20)["items"][0]

    for column in hub_service_module._TASK_LIST_SELECT_COLUMNS:
        assert column in listed
    assert listed["latest_report"] is None
    assert listed["attachments"] == []
    assert listed["attachments_count"] == 0
    assert listed["project_name"]
    assert listed["checklist_total"] == detail["checklist_total"]
    assert listed["has_unread_comments"] is detail["has_unread_comments"]


def test_department_scope_list_excludes_foreign_department_tasks(task_env, monkeypatch):
    service = task_env["service"]
    raw_users = task_env["raw_users"]
    dept_visible = "dept-visible"
    dept_foreign = "dept-foreign"

    monkeypatch.setattr(
        hub_service_module.department_service,
        "get_department",
        lambda department_id: {"id": department_id, "name": department_id}
        if str(department_id) in {dept_visible, dept_foreign}
        else None,
    )
    monkeypatch.setattr(
        hub_service_module.department_service,
        "get_user_department_ids",
        lambda user, roles=None: {dept_visible}
        if int(user.get("id") if isinstance(user, dict) else user) == 4
        else set(),
    )
    monkeypatch.setattr(
        hub_service_module,
        "can_create_task_for_department",
        lambda *args, **kwargs: True,
    )
    monkeypatch.setattr(
        access_policy_module,
        "user_is_department_manager",
        lambda user, department_id: False,
    )
    monkeypatch.setattr(
        access_policy_module,
        "user_is_department_member",
        lambda user, department_id: (
            str(department_id) == dept_visible and int(user.get("id") if isinstance(user, dict) else user) == 4
        ),
    )

    actor = raw_users[1]
    project = service.create_task_project(name="Department Scope Project", code=f"dept-scope-{uuid4().hex[:8]}")
    visible = service.create_task(
        title="Visible Department Task",
        description="Visible in department scope",
        assignee_user_id=2,
        controller_user_id=3,
        due_at=None,
        project_id=project["id"],
        department_id=dept_visible,
        visibility_scope="department",
        actor=actor,
    )
    foreign = service.create_task(
        title="Foreign Department Task",
        description="Must stay hidden",
        assignee_user_id=2,
        controller_user_id=3,
        due_at=None,
        project_id=project["id"],
        department_id=dept_foreign,
        visibility_scope="department",
        actor=actor,
    )
    assignee_private = service.create_task(
        title="Assignee Private Foreign Task",
        description="Visible as assignee",
        assignee_user_id=4,
        controller_user_id=3,
        due_at=None,
        project_id=project["id"],
        department_id=dept_foreign,
        visibility_scope="private",
        actor=actor,
    )

    payload = service.list_tasks(user_id=4, scope="department", limit=50)
    task_ids = {item["id"] for item in payload["items"]}
    assert visible["id"] in task_ids
    assert foreign["id"] not in task_ids
    assert assignee_private["id"] in task_ids


def test_department_scope_pagination_offset_and_total(task_env, monkeypatch):
    service = task_env["service"]
    raw_users = task_env["raw_users"]
    dept_visible = "dept-pagination"

    monkeypatch.setattr(
        hub_service_module.department_service,
        "get_department",
        lambda department_id: {"id": department_id, "name": department_id}
        if str(department_id) == dept_visible
        else None,
    )
    monkeypatch.setattr(
        hub_service_module.department_service,
        "get_user_department_ids",
        lambda user, roles=None: {dept_visible}
        if int(user.get("id") if isinstance(user, dict) else user) == 4
        else set(),
    )
    monkeypatch.setattr(
        hub_service_module,
        "can_create_task_for_department",
        lambda *args, **kwargs: True,
    )
    monkeypatch.setattr(
        access_policy_module,
        "user_is_department_manager",
        lambda user, department_id: False,
    )
    monkeypatch.setattr(
        access_policy_module,
        "user_is_department_member",
        lambda user, department_id: (
            str(department_id) == dept_visible and int(user.get("id") if isinstance(user, dict) else user) == 4
        ),
    )

    actor = raw_users[1]
    project = service.create_task_project(name="Department Pagination Project", code=f"dept-page-{uuid4().hex[:8]}")
    created_ids: list[str] = []
    for index in range(5):
        created = service.create_task(
            title=f"Department Pagination {index}",
            description="Pagination task",
            assignee_user_id=2,
            controller_user_id=3,
            due_at=None,
            project_id=project["id"],
            department_id=dept_visible,
            visibility_scope="department",
            actor=actor,
        )
        created_ids.append(created["id"])

    page_one = service.list_tasks(user_id=4, scope="department", limit=2, offset=0, sort_by="updated_at", sort_dir="asc")
    page_two = service.list_tasks(user_id=4, scope="department", limit=2, offset=2, sort_by="updated_at", sort_dir="asc")
    page_three = service.list_tasks(user_id=4, scope="department", limit=2, offset=4, sort_by="updated_at", sort_dir="asc")

    assert page_one["total"] == 5
    assert page_two["total"] == 5
    assert page_three["total"] == 5
    assert len(page_one["items"]) == 2
    assert len(page_two["items"]) == 2
    assert len(page_three["items"]) == 1
    assert {item["id"] for item in page_one["items"]}.isdisjoint({item["id"] for item in page_two["items"]})
    returned_ids = [item["id"] for item in page_one["items"] + page_two["items"] + page_three["items"]]
    assert len(returned_ids) == len(set(returned_ids))
    assert set(returned_ids).issubset(set(created_ids))


def test_task_due_notification_cycle_includes_assignee_and_delegate(task_env, monkeypatch):
    service = task_env["service"]
    client = task_env["client"]
    set_user = task_env["set_user"]

    monkeypatch.setattr(hub_service_module.user_service, "get_delegate_user_ids", lambda owner_user_id, active_only=True: [7] if int(owner_user_id) == 2 else [])
    monkeypatch.setattr(hub_service_module.user_service, "get_delegate_owner_ids", lambda delegate_user_id, active_only=True: [2] if int(delegate_user_id) == 7 else [])

    set_user(1)
    due_at = (datetime.now(timezone.utc) + timedelta(hours=12)).isoformat()
    _create_task(client, title="Delegate Due Cycle", due_at=due_at)

    with service._lock, service._connect() as conn:
        user_ids = service._collect_task_due_notification_user_ids(conn)
    assert user_ids == [2, 7]

    processed = service.run_task_due_notification_cycle()
    assert processed == 2


def test_task_email_deadline_soon_and_overdue_digest_are_deduped(task_env, monkeypatch):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]
    raw_users = task_env["raw_users"]

    raw_users[2]["mailbox_email"] = "assignee@example.test"
    monkeypatch.setenv("TASK_EMAIL_NOTIFICATIONS_ENABLED", "1")
    monkeypatch.setenv("TASK_EMAIL_AUTODISPATCH_ENABLED", "0")
    monkeypatch.setenv("TASK_EMAIL_DEADLINE_SOON_HOURS", "24")

    set_user(1)
    soon_due = (datetime.now(timezone.utc) + timedelta(hours=23)).isoformat()
    overdue_due = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
    soon_task = _create_task(client, title="Email Deadline Soon", due_at=soon_due)
    overdue_task = _create_task(client, title="Email Overdue Digest", due_at=overdue_due)

    set_user(2)
    service.run_task_due_notifications_for_user(2, force=True)

    rows = _task_email_rows(service)
    deadline_rows = [
        row for row in rows
        if row["event_type"] == "task.deadline_soon" and row["task_id"] == soon_task["id"]
    ]
    digest_rows = [row for row in rows if row["event_type"] == "task.overdue_digest"]

    assert len(deadline_rows) == 1
    assert deadline_rows[0]["recipient_email"] == "assignee@example.test"
    assert "Email Deadline Soon" in deadline_rows[0]["subject"]
    assert len(digest_rows) == 1
    assert digest_rows[0]["recipient_email"] == "assignee@example.test"
    assert "Email Overdue Digest" in digest_rows[0]["body_text"]
    assert overdue_task["id"] in digest_rows[0]["body_text"]


def test_create_task_stores_custom_email_deadline_remind_hours(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]

    set_user(1)
    due_at = (datetime.now(timezone.utc) + timedelta(days=2)).isoformat()
    project_response = client.post(
        "/hub/task-projects",
        json={"name": "Email Remind Project", "code": "email-remind-project"},
    )
    assert project_response.status_code == 200, project_response.text
    project = project_response.json()

    response = client.post(
        "/hub/tasks",
        json={
            "title": "Custom Email Remind",
            "description": "Task body",
            "assignee_user_ids": [2],
            "controller_user_id": 3,
            "project_id": project["id"],
            "protocol_date": "2026-03-01",
            "due_at": due_at,
            "email_deadline_remind_hours": 48,
        },
    )
    assert response.status_code == 200, response.text
    task = response.json()["items"][0]
    assert task["email_deadline_remind_hours"] == 48

    with service._lock, service._connect() as conn:
        row = conn.execute(
            f"SELECT email_deadline_remind_hours FROM {service._TASKS_TABLE} WHERE id = ?",
            (task["id"],),
        ).fetchone()
    assert row is not None
    assert int(row["email_deadline_remind_hours"]) == 48


def test_deadline_soon_email_respects_per_task_off_and_custom_hours(task_env, monkeypatch):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]
    raw_users = task_env["raw_users"]

    raw_users[2]["mailbox_email"] = "assignee@example.test"
    monkeypatch.setenv("TASK_EMAIL_NOTIFICATIONS_ENABLED", "1")
    monkeypatch.setenv("TASK_EMAIL_AUTODISPATCH_ENABLED", "0")
    monkeypatch.setenv("TASK_EMAIL_DEADLINE_SOON_HOURS", "24")

    set_user(1)
    due_soon = (datetime.now(timezone.utc) + timedelta(hours=20)).isoformat()
    due_later = (datetime.now(timezone.utc) + timedelta(hours=60)).isoformat()

    off_task = _create_task(
        client,
        title="Email Remind Off",
        due_at=due_soon,
        email_deadline_remind_hours=0,
    )
    custom_task = _create_task(
        client,
        title="Email Remind Custom",
        due_at=due_later,
        email_deadline_remind_hours=48,
    )

    set_user(2)
    service.run_task_due_notifications_for_user(2, force=True)

    rows = _task_email_rows(service)
    deadline_rows = [row for row in rows if row["event_type"] == "task.deadline_soon"]

    assert not any(row["task_id"] == off_task["id"] for row in deadline_rows)
    assert not any(row["task_id"] == custom_task["id"] for row in deadline_rows)

    notifications = service.poll_notifications(user_id=2, limit=50)
    assert any(
        item.get("event_type") == "task.deadline_soon" and item.get("entity_id") == off_task["id"]
        for item in notifications.get("items", [])
    )


def test_task_observer_can_view_list_and_chat_capabilities(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]

    set_user(1)
    created = _create_task(client, title="Observer Task", observer_user_ids=[4])
    task_id = created["id"]
    assert created["observer_user_ids"] == [4]
    assert len(created.get("observers") or []) == 1
    assert created["observers"][0]["user_id"] == 4

    set_user(4)
    my_tasks = client.get("/hub/tasks", params={"scope": "my", "role_scope": "both"})
    assert my_tasks.status_code == 200, my_tasks.text
    assert my_tasks.json()["total"] == 1

    detail = client.get(f"/hub/tasks/{task_id}")
    assert detail.status_code == 200, detail.text
    payload = detail.json()
    assert payload["is_observer"] is True
    assert payload["capabilities"] == {
        "can_edit": False,
        "can_start": False,
        "can_submit": False,
        "can_review": False,
        "can_reopen": False,
        "can_upload_files": False,
        "can_update_checklist": False,
        "can_open_discussion": False,
    }

    start_denied = client.post(f"/hub/tasks/{task_id}/start")
    assert start_denied.status_code == 403


def test_observer_membership_clause_uses_postgresql_json_array_elements(monkeypatch):
    service = hub_service_module.HubService()
    monkeypatch.setattr(service, "_uses_postgresql", lambda: True)
    clause, params = service._observer_membership_clause(5)
    assert "json_array_elements_text" in clause
    assert "json_each" not in clause
    assert params == [5]


def test_users_by_id_reuses_cached_active_user_directory(task_env, monkeypatch):
    service = task_env["service"]
    calls = {"count": 0}

    def _list_users():
        calls["count"] += 1
        return [
            {"id": 1, "username": "author", "full_name": "Task Author", "is_active": True},
            {"id": 4, "username": "outsider", "full_name": "Task Outsider", "is_active": True},
        ]

    monkeypatch.setattr(hub_service_module.user_service, "list_users", _list_users)

    first = service._users_by_id()
    second = service._users_by_id()

    assert calls["count"] == 1
    assert first[1]["username"] == "author"
    assert second[4]["username"] == "outsider"


def test_task_observer_added_notification_on_create_and_patch(task_env):
    client = task_env["client"]
    set_user = task_env["set_user"]
    service = task_env["service"]

    set_user(1)
    created = _create_task(client, title="Observer Notify", observer_user_ids=[4, 7])
    task_id = created["id"]

    notifications = service.poll_notifications(user_id=4, limit=20)
    assert any(
        item.get("event_type") == "task.observer_added" and item.get("entity_id") == task_id
        for item in notifications.get("items", [])
    )

    set_user(1)
    patch_response = client.patch(
        f"/hub/tasks/{task_id}",
        json={"observer_user_ids": [4, 7, 6]},
    )
    assert patch_response.status_code == 200, patch_response.text

    taskonly_notifications = service.poll_notifications(user_id=6, limit=20)
    assert any(
        item.get("event_type") == "task.observer_added" and item.get("entity_id") == task_id
        for item in taskonly_notifications.get("items", [])
    )
