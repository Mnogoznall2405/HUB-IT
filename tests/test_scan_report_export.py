from __future__ import annotations

import time
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import scan_server.app as scan_app
from scan_server.database import ScanStore
from scan_server.report_export import XLSX_MEDIA_TYPE, build_scan_task_incidents_excel


def _make_store(temp_dir: str) -> ScanStore:
    root = Path(temp_dir)
    return ScanStore(
        db_path=root / "scan-server.db",
        archive_dir=root / "archive",
        task_ack_timeout_sec=300,
        agent_online_timeout_sec=1800,
    )


def _get_job(store: ScanStore, job_id: str) -> dict:
    with store._lock, store._connect() as conn:
        row = conn.execute("SELECT * FROM scan_jobs WHERE id=?", (job_id,)).fetchone()
    return dict(row)


def _seed_task_incident(store: ScanStore, *, task_id: str, event_id: str = "event-1") -> dict:
    queued = store.queue_job(
        {
            "agent_id": "agent-1",
            "hostname": "HOST-01",
            "branch": "Тюмень",
            "user_login": "corp\\petrov",
            "user_full_name": "Петров П.П.",
            "file_path": r"C:\Docs\secret.pdf",
            "file_name": "secret.pdf",
            "file_hash": "hash-secret",
            "file_size": 2048,
            "source_kind": "pdf_slice",
            "event_id": event_id,
            "scan_task_id": task_id,
        }
    )
    job = _get_job(store, queued["job_id"])
    incident = store.create_finding_and_incident(
        job=job,
        severity="high",
        category="secrets",
        matched_patterns=[{"pattern_name": "password", "value": "secret"}],
        short_reason="pattern match",
    )
    store.finalize_job(job_id=queued["job_id"], status="done_with_incident", summary="incident")
    return {"job": job, "incident": incident}


def _create_scan_task(store: ScanStore) -> dict:
    store.upsert_agent_heartbeat(
        {
            "agent_id": "agent-1",
            "hostname": "HOST-01",
            "branch": "Тюмень",
            "ip_address": "10.10.1.1",
            "version": "1.2.3",
            "status": "online",
            "last_seen_at": int(time.time()),
        }
    )
    task = store.create_task(agent_id="agent-1", command="scan_now", dedupe_key="scan_now:agent-1")
    store.report_task_result(
        agent_id="agent-1",
        task_id=task["id"],
        status="acknowledged",
        result={"phase": "server_processing", "scanned": 1, "skipped": 0},
        error_text="",
    )
    return task


def test_scan_task_incident_report_contains_linked_job_incidents(temp_dir):
    store = _make_store(temp_dir)
    task = _create_scan_task(store)
    seeded = _seed_task_incident(store, task_id=task["id"])

    report = store.get_scan_task_incident_report(task_id=task["id"])

    assert report is not None
    assert report["task"]["id"] == task["id"]
    assert report["task"]["hostname"] == "HOST-01"
    assert report["summary"]["jobs_done_with_incident"] == 1
    assert report["summary"]["incidents_total"] == 1
    assert report["summary"]["severity_counts"] == {"high": 1}
    assert report["incidents"][0]["id"] == seeded["incident"]["incident_id"]
    assert report["incidents"][0]["observation_types"] == "found_new"
    assert report["incidents"][0]["matched_patterns"][0]["pattern_name"] == "password"


def test_scan_task_incident_report_includes_duplicate_and_resolved_observations(temp_dir):
    store = _make_store(temp_dir)
    task = _create_scan_task(store)

    previous = store.queue_job(
        {
            "agent_id": "agent-1",
            "hostname": "HOST-01",
            "branch": "Тюмень",
            "file_path": r"C:\Docs\old.pdf",
            "file_name": "old.pdf",
            "file_hash": "hash-duplicate",
            "file_size": 2048,
            "source_kind": "pdf_slice",
            "event_id": "duplicate-event",
        }
    )
    previous_incident = store.create_finding_and_incident(
        job=_get_job(store, previous["job_id"]),
        severity="medium",
        category="secrets",
        matched_patterns=[{"pattern_name": "loan", "value": "loan"}],
        short_reason="duplicate pattern",
    )
    store.finalize_job(job_id=previous["job_id"], status="done_with_incident", summary="incident")

    duplicate = store.queue_job(
        {
            "agent_id": "agent-1",
            "hostname": "HOST-01",
            "branch": "Тюмень",
            "file_path": r"C:\Docs\old.pdf",
            "file_name": "old.pdf",
            "file_hash": "hash-duplicate",
            "file_size": 2048,
            "source_kind": "pdf_slice",
            "event_id": "duplicate-event",
            "scan_task_id": task["id"],
        }
    )
    assert duplicate["deduped"] is True

    store.report_task_result(
        agent_id="agent-1",
        task_id=task["id"],
        status="completed",
        result={
            "phase": "completed",
            "force_rescan": True,
            "scanned": 1,
            "deleted_file_events": [
                {
                    "event_id": "duplicate-event",
                    "file_path": r"C:\Docs\old.pdf",
                    "file_hash": "hash-duplicate",
                    "source_kind": "pdf_slice",
                }
            ],
        },
        error_text="",
    )

    report = store.get_scan_task_incident_report(task_id=task["id"])

    assert report is not None
    assert report["summary"]["found_duplicate"] == 1
    assert report["summary"]["deleted"] == 1
    incident = report["incidents"][0]
    assert incident["id"] == previous_incident["incident_id"]
    assert "found_duplicate" in incident["observation_types"]
    assert "deleted" in incident["observation_types"]


def test_scan_task_incident_report_handles_many_linked_observations(temp_dir):
    store = _make_store(temp_dir)
    task = _create_scan_task(store)
    seeded = _seed_task_incident(store, task_id=task["id"])

    with store._lock, store._connect() as conn:
        for index in range(1005):
            store._record_scan_observation_locked(
                conn,
                scan_task_id=task["id"],
                agent_id="agent-1",
                hostname="HOST-01",
                file_path=rf"C:\Docs\missing-{index}.pdf",
                file_hash=f"missing-hash-{index}",
                event_id=f"missing-event-{index}",
                observation_type="found_duplicate",
                linked_incident_id=f"missing-incident-{index}",
                source_kind="pdf_slice",
                severity="medium",
            )
        conn.commit()

    report = store.get_scan_task_incident_report(task_id=task["id"])

    assert report is not None
    assert report["summary"]["observations_total"] == 1006
    assert [item["id"] for item in report["incidents"]] == [seeded["incident"]["incident_id"]]


def test_scan_task_incidents_excel_has_summary_and_incident_sheet(temp_dir):
    store = _make_store(temp_dir)
    task = _create_scan_task(store)
    _seed_task_incident(store, task_id=task["id"])
    report = store.get_scan_task_incident_report(task_id=task["id"])

    file_bytes, filename = build_scan_task_incidents_excel(report)
    workbook = load_workbook(BytesIO(file_bytes))

    assert filename.startswith("scan_incidents_HOST-01_")
    assert workbook.sheetnames == ["Сводка", "Инциденты"]
    assert workbook["Сводка"]["A1"].value == "Отчет по запуску скана"
    assert workbook["Инциденты"]["A1"].value == "Время"
    assert workbook["Инциденты"]["D2"].value == r"C:\Docs\secret.pdf"


def _override_scan_read_dependency():
    for route in scan_app.app.routes:
        if getattr(route, "path", "") == "/api/v1/scan/tasks/{task_id}/incidents/export":
            for dependency in route.dependant.dependencies:
                scan_app.app.dependency_overrides[dependency.call] = lambda: {"id": 1, "role": "admin"}
            return
    raise AssertionError("scan task incident export route not found")


def test_scan_task_incidents_export_endpoint_returns_xlsx(monkeypatch, temp_dir):
    store = _make_store(temp_dir)
    task = _create_scan_task(store)
    _seed_task_incident(store, task_id=task["id"])
    monkeypatch.setattr(scan_app, "store", store)
    _override_scan_read_dependency()
    try:
        response = TestClient(scan_app.app).get(f"/api/v1/scan/tasks/{task['id']}/incidents/export")
    finally:
        scan_app.app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
    assert "scan_incidents_HOST-01" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Инциденты"]["D2"].value == r"C:\Docs\secret.pdf"


def test_scan_task_incidents_export_endpoint_returns_404_for_unknown_task(monkeypatch, temp_dir):
    store = _make_store(temp_dir)
    monkeypatch.setattr(scan_app, "store", store)
    _override_scan_read_dependency()
    try:
        response = TestClient(scan_app.app).get("/api/v1/scan/tasks/missing/incidents/export")
    finally:
        scan_app.app.dependency_overrides.clear()

    assert response.status_code == 404
