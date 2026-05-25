"""Tests for TicketsImportService — import execution pipeline (task 6.3).

Tests cover:
- execute_import(): full pipeline with working sheets, ПОТЕРИ, duplicates
- _parse_row(): mandatory field validation, optional field parsing
- _parse_row_standard_schema(): headerless sheet parsing
- _parse_losses_sheet(): ПОТЕРИ sheet → FinancialOpRow
- _detect_duplicates(): duplicate detection by composite key
- _normalize_value(): phone-as-date, passport-as-number, date-as-text
- Raw trace storage for every imported row
- Reconciliation report after import

Requirements: 1.2, 1.4, 1.5, 1.6, 1.8, 1.9, 1.10, 1.11
"""
from __future__ import annotations

import json
import sys
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.appdb.models import AppBase
from backend.appdb.tickets_models import (
    TicketEmployee,
    TicketFinancialOp,
    TicketImportJob,
    TicketImportRawTrace,
    TicketObject,
    TicketRequest,
)
from backend.services.tickets_import_service import (
    DuplicateMatch,
    FinancialOpRow,
    ImportResult,
    ImportSettings,
    ParsedRow,
    ParseError,
    ReconciliationReport,
    TicketsImportService,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _sqlite_url(temp_dir: str) -> str:
    return f"sqlite:///{(Path(temp_dir) / 'tickets_import_exec.db').as_posix()}"


@pytest.fixture
def db_session_factory(temp_dir, monkeypatch):
    """Set up a fresh SQLite database and return session factory."""
    import backend.appdb.db as appdb

    url = _sqlite_url(temp_dir)

    appdb._engines.clear()
    appdb._session_factories.clear()
    appdb._initialized_schema_urls.clear()

    monkeypatch.setenv("APP_DATABASE_URL", url)

    try:
        from backend.config import config
        monkeypatch.setattr(config, "app_database_url", url, raising=False)
    except (ImportError, AttributeError):
        pass

    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.close()

    AppBase.metadata.create_all(engine, checkfirst=True)

    SessionLocal = sessionmaker(bind=engine)

    @contextmanager
    def _test_app_session(database_url=None):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    monkeypatch.setattr(
        "backend.services.tickets_import_service.app_session", _test_app_session
    )

    return SessionLocal


@pytest.fixture
def service(temp_dir, db_session_factory):
    """Create a TicketsImportService with temp upload dir."""
    upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
    return TicketsImportService(upload_base_dir=str(upload_dir))


@pytest.fixture
def setup_object(db_session_factory):
    """Create a test TicketObject in the database."""
    with db_session_factory() as session:
        obj = TicketObject(
            id=1,
            code="KAM",
            name="Камчатка",
            short_name="Камч",
            region="Дальний Восток",
            is_active=True,
        )
        session.add(obj)
        session.commit()
    return 1  # object_id


@pytest.fixture
def setup_job(service, temp_dir, db_session_factory):
    """Create a previewed import job with a workbook file."""
    def _create_job(workbook: Workbook) -> str:
        # Save workbook to temp
        upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
        upload_dir.mkdir(parents=True, exist_ok=True)

        result = service.upload_file("test.xlsx", b"placeholder", user_id=1)
        job_id = result["id"]

        # Save actual workbook
        file_path = upload_dir / f"{job_id}.xlsx"
        workbook.save(str(file_path))

        # Update job to previewed status
        with db_session_factory() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            job.status = "previewed"
            job.file_path = str(file_path)
            session.commit()

        return job_id

    return _create_job


# ---------------------------------------------------------------------------
# Tests: _normalize_value
# ---------------------------------------------------------------------------


class TestNormalizeValue:
    """Tests for _normalize_value() method."""

    def test_date_from_datetime(self):
        """datetime value passes through unchanged."""
        dt = datetime(2024, 3, 15, tzinfo=timezone.utc)
        result, needs_review = TicketsImportService._normalize_value(dt, "date")
        assert result == dt
        assert needs_review is False

    def test_date_from_excel_serial(self):
        """Excel serial number is converted to datetime with needs_review."""
        # 45000 ≈ 2023-03-14
        result, needs_review = TicketsImportService._normalize_value(45000, "date")
        assert isinstance(result, datetime)
        assert needs_review is True

    def test_date_from_text_dot_format(self):
        """Date string in dd.mm.YYYY format is parsed."""
        result, needs_review = TicketsImportService._normalize_value("15.03.2024", "date")
        assert isinstance(result, datetime)
        assert result.day == 15
        assert result.month == 3
        assert result.year == 2024
        assert needs_review is True

    def test_date_from_text_iso_format(self):
        """Date string in YYYY-MM-DD format is parsed."""
        result, needs_review = TicketsImportService._normalize_value("2024-03-15", "date")
        assert isinstance(result, datetime)
        assert result.year == 2024
        assert needs_review is True

    def test_date_none_returns_none(self):
        """None value returns (None, False)."""
        result, needs_review = TicketsImportService._normalize_value(None, "date")
        assert result is None
        assert needs_review is False

    def test_date_invalid_text(self):
        """Invalid date text returns (None, True)."""
        result, needs_review = TicketsImportService._normalize_value("not a date", "date")
        assert result is None
        assert needs_review is True

    def test_phone_from_number_10_digits(self):
        """10-digit number is formatted as +7XXXXXXXXXX."""
        result, needs_review = TicketsImportService._normalize_value(9001234567, "phone")
        assert result == "+79001234567"
        assert needs_review is True

    def test_phone_from_number_11_digits_8(self):
        """11-digit number starting with 8 is formatted as +7."""
        result, needs_review = TicketsImportService._normalize_value(89001234567, "phone")
        assert result == "+79001234567"
        assert needs_review is True

    def test_phone_from_string(self):
        """String phone passes through unchanged."""
        result, needs_review = TicketsImportService._normalize_value("+79001234567", "phone")
        assert result == "+79001234567"
        assert needs_review is False

    def test_phone_from_excel_serial(self):
        """Large number (Excel date serial) is flagged for review."""
        result, needs_review = TicketsImportService._normalize_value(45000, "phone")
        assert needs_review is True
        assert isinstance(result, str)

    def test_passport_from_number(self):
        """Numeric passport is formatted with leading zeros and space."""
        # 1234567890 → "1234 567890"
        result, needs_review = TicketsImportService._normalize_value(1234567890, "passport")
        assert result == "1234 567890"
        assert needs_review is True

    def test_passport_from_short_number(self):
        """Short numeric passport is zero-padded."""
        result, needs_review = TicketsImportService._normalize_value(123456, "passport")
        assert "0000" in result  # padded to 10 digits
        assert needs_review is True

    def test_passport_from_string(self):
        """String passport passes through unchanged."""
        result, needs_review = TicketsImportService._normalize_value("1234 567890", "passport")
        assert result == "1234 567890"
        assert needs_review is False

    def test_decimal_from_int(self):
        """Integer is converted to Decimal."""
        result, needs_review = TicketsImportService._normalize_value(15000, "decimal")
        assert result == Decimal("15000")
        assert needs_review is False

    def test_decimal_from_float(self):
        """Float is converted to Decimal."""
        result, needs_review = TicketsImportService._normalize_value(15000.50, "decimal")
        assert result == Decimal("15000.5")
        assert needs_review is False

    def test_decimal_from_string_with_comma(self):
        """String with comma decimal separator is parsed."""
        result, needs_review = TicketsImportService._normalize_value("15 000,50", "decimal")
        assert result == Decimal("15000.50")
        assert needs_review is True

    def test_decimal_none(self):
        """None returns (None, False)."""
        result, needs_review = TicketsImportService._normalize_value(None, "decimal")
        assert result is None
        assert needs_review is False


# ---------------------------------------------------------------------------
# Tests: _parse_row
# ---------------------------------------------------------------------------


class TestParseRow:
    """Tests for _parse_row() method."""

    def _make_workbook_row(self, values: list):
        """Helper: create a workbook and return cells from row 2."""
        wb = Workbook()
        ws = wb.active
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=2, column=col_idx, value=val)
        return tuple(ws.iter_rows(min_row=2, max_row=2))[0]

    def test_parse_row_with_full_name(self, service):
        """Row with ФИО is parsed successfully."""
        row = self._make_workbook_row(["Иванов И.И.", "2024-01-15", None, "Москва-Камчатка", 25000])
        header_map = {0: "full_name", 1: "submitted_at", 2: "departure_date", 3: "route", 4: "cost"}

        result = service._parse_row(row, header_map, 2, {}, {}, {})
        assert isinstance(result, ParsedRow)
        assert result.full_name == "Иванов И.И."
        assert result.route == "Москва-Камчатка"
        assert result.cost == Decimal("25000")

    def test_parse_row_missing_full_name_returns_error(self, service):
        """Row without ФИО returns ParseError."""
        row = self._make_workbook_row([None, "2024-01-15", None, "Москва", 25000])
        header_map = {0: "full_name", 1: "submitted_at", 3: "route", 4: "cost"}

        result = service._parse_row(row, header_map, 3, {}, {}, {})
        assert isinstance(result, ParseError)
        assert result.row_number == 3
        assert "ФИО" in result.error

    def test_parse_row_empty_string_full_name_returns_error(self, service):
        """Row with empty string ФИО returns ParseError."""
        row = self._make_workbook_row(["   ", "2024-01-15"])
        header_map = {0: "full_name", 1: "submitted_at"}

        result = service._parse_row(row, header_map, 4, {}, {}, {})
        assert isinstance(result, ParseError)

    def test_parse_row_optional_fields_missing(self, service):
        """Row with only ФИО (optional fields missing) is still valid."""
        row = self._make_workbook_row(["Петров П.П.", None, None, None, None])
        header_map = {0: "full_name", 1: "submitted_at", 2: "departure_date", 3: "route", 4: "cost"}

        result = service._parse_row(row, header_map, 2, {}, {}, {})
        assert isinstance(result, ParsedRow)
        assert result.full_name == "Петров П.П."
        assert result.cost == Decimal("0.00")

    def test_parse_row_stores_raw_cells(self, service):
        """Raw cell values are stored in the parsed row."""
        row = self._make_workbook_row(["Иванов И.И.", 25000])
        header_map = {0: "full_name", 1: "cost"}

        result = service._parse_row(row, header_map, 2, {}, {}, {})
        assert isinstance(result, ParsedRow)
        assert len(result.raw_cells) > 0
        assert len(result.cell_addresses) > 0

    def test_parse_row_color_maps_to_status(self, service):
        """Cell color is mapped to status via color_map."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="Иванов И.И.")
        ws["A2"].fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        row = tuple(ws.iter_rows(min_row=2, max_row=2))[0]
        header_map = {0: "full_name"}
        color_map = {"C6EFCE": "purchased"}

        result = service._parse_row(row, header_map, 2, {}, {}, color_map)
        assert isinstance(result, ParsedRow)
        assert result.status == "purchased"

    def test_parse_row_needs_review_for_date_as_text(self, service):
        """Date stored as text triggers needs_review flag."""
        row = self._make_workbook_row(["Иванов И.И.", "15.03.2024"])
        header_map = {0: "full_name", 1: "submitted_at"}

        result = service._parse_row(row, header_map, 2, {}, {}, {})
        assert isinstance(result, ParsedRow)
        assert result.needs_review is True
        assert "submitted_at" in result.review_fields


# ---------------------------------------------------------------------------
# Tests: _parse_row_standard_schema
# ---------------------------------------------------------------------------


class TestParseRowStandardSchema:
    """Tests for _parse_row_standard_schema() — headerless sheets."""

    def _make_row(self, values: list):
        """Helper: create cells from a row."""
        wb = Workbook()
        ws = wb.active
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=1, column=col_idx, value=val)
        return tuple(ws.iter_rows(min_row=1, max_row=1))[0]

    def test_standard_schema_parses_by_position(self, service):
        """Values are mapped by column position per STANDARD_COLUMN_SCHEMA."""
        from backend.services.tickets_import_service import STANDARD_COLUMN_SCHEMA

        row = self._make_row(["Сидоров С.С.", "15.01.2024", "20.01.2024", "Москва-Тикси", 30000, "В работе", "Логист1"])

        result = service._parse_row_standard_schema(
            row, STANDARD_COLUMN_SCHEMA, 1, {}, {}, {}
        )
        assert isinstance(result, ParsedRow)
        assert result.full_name == "Сидоров С.С."
        assert result.route == "Москва-Тикси"

    def test_standard_schema_missing_name_returns_error(self, service):
        """Missing ФИО in first column returns ParseError."""
        from backend.services.tickets_import_service import STANDARD_COLUMN_SCHEMA

        row = self._make_row([None, "15.01.2024", "20.01.2024"])

        result = service._parse_row_standard_schema(
            row, STANDARD_COLUMN_SCHEMA, 1, {}, {}, {}
        )
        assert isinstance(result, ParseError)


# ---------------------------------------------------------------------------
# Tests: _parse_losses_sheet
# ---------------------------------------------------------------------------


class TestParseLossesSheet:
    """Tests for _parse_losses_sheet() method."""

    def test_parse_losses_with_headers(self, service):
        """ПОТЕРИ sheet with headers is parsed correctly."""
        wb = Workbook()
        ws = wb.active
        ws.title = "ПОТЕРИ"
        ws["A1"] = "ФИО"
        ws["B1"] = "Сумма"
        ws["C1"] = "Причина"
        ws["D1"] = "Дата"
        ws["E1"] = "Тип операции"
        ws["A2"] = "Иванов И.И."
        ws["B2"] = 15000.50
        ws["C2"] = "Неявка"
        ws["D2"] = datetime(2024, 3, 15, tzinfo=timezone.utc)
        ws["E2"] = "потеря"

        results = service._parse_losses_sheet(ws)
        assert len(results) == 1
        assert results[0].full_name == "Иванов И.И."
        assert results[0].amount == Decimal("15000.5")
        assert results[0].reason == "Неявка"
        assert results[0].op_type == "loss"

    def test_parse_losses_refund_type(self, service):
        """Op type 'возврат' maps to 'refund'."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ФИО"
        ws["B1"] = "Сумма"
        ws["C1"] = "Причина"
        ws["D1"] = "Дата"
        ws["E1"] = "Тип"
        ws["A2"] = "Петров П.П."
        ws["B2"] = 5000
        ws["C2"] = "Отмена рейса"
        ws["D2"] = None
        ws["E2"] = "возврат"

        results = service._parse_losses_sheet(ws)
        assert len(results) == 1
        assert results[0].op_type == "refund"

    def test_parse_losses_skips_empty_name(self, service):
        """Rows without ФИО are skipped."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ФИО"
        ws["B1"] = "Сумма"
        ws["A2"] = None
        ws["B2"] = 5000
        ws["A3"] = "Сидоров С.С."
        ws["B3"] = 3000

        results = service._parse_losses_sheet(ws)
        assert len(results) == 1
        assert results[0].full_name == "Сидоров С.С."

    def test_parse_losses_without_headers(self, service):
        """ПОТЕРИ sheet without recognized headers uses positional mapping."""
        wb = Workbook()
        ws = wb.active
        # No recognized headers
        ws["A1"] = "Иванов И.И."
        ws["B1"] = 10000
        ws["C1"] = "Опоздание"
        ws["D1"] = datetime(2024, 1, 10, tzinfo=timezone.utc)
        ws["E1"] = "потеря"

        results = service._parse_losses_sheet(ws)
        assert len(results) == 1
        assert results[0].full_name == "Иванов И.И."
        assert results[0].amount == Decimal("10000")


# ---------------------------------------------------------------------------
# Tests: _detect_duplicates
# ---------------------------------------------------------------------------


class TestDetectDuplicates:
    """Tests for _detect_duplicates() method."""

    def test_no_duplicates_empty_db(self, service, db_session_factory, setup_object):
        """No duplicates when DB is empty."""
        parsed_rows = [
            ParsedRow(full_name="Иванов И.И.", submitted_at=datetime(2024, 1, 15, tzinfo=timezone.utc), row_number=2),
        ]
        result = service._detect_duplicates(parsed_rows, setup_object)
        assert result == []

    def test_detects_duplicate_by_composite_key(self, service, db_session_factory, setup_object):
        """Duplicate is detected by (full_name + object_id + submitted_at)."""
        submitted = datetime(2024, 1, 15, tzinfo=timezone.utc)

        # Create existing record
        with db_session_factory() as session:
            emp = TicketEmployee(id=10, full_name="Иванов И.И.")
            session.add(emp)
            session.flush()
            req = TicketRequest(
                id=100,
                employee_id=emp.id,
                object_id=setup_object,
                status="new",
                submitted_at=submitted,
                source="import",
            )
            session.add(req)
            session.commit()

        parsed_rows = [
            ParsedRow(full_name="Иванов И.И.", submitted_at=submitted, row_number=2),
        ]
        result = service._detect_duplicates(parsed_rows, setup_object)
        assert len(result) == 1
        assert result[0].existing_id == 100

    def test_no_duplicate_different_name(self, service, db_session_factory, setup_object):
        """Different name is not a duplicate."""
        submitted = datetime(2024, 1, 15, tzinfo=timezone.utc)

        with db_session_factory() as session:
            emp = TicketEmployee(id=10, full_name="Иванов И.И.")
            session.add(emp)
            session.flush()
            req = TicketRequest(
                id=100, employee_id=emp.id, object_id=setup_object,
                status="new", submitted_at=submitted, source="import",
            )
            session.add(req)
            session.commit()

        parsed_rows = [
            ParsedRow(full_name="Петров П.П.", submitted_at=submitted, row_number=2),
        ]
        result = service._detect_duplicates(parsed_rows, setup_object)
        assert result == []


# ---------------------------------------------------------------------------
# Tests: execute_import (integration)
# ---------------------------------------------------------------------------


class TestExecuteImport:
    """Integration tests for execute_import() method."""

    def test_import_simple_working_sheet(self, service, db_session_factory, setup_object, setup_job):
        """Import a simple working sheet creates requests and raw traces."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Камчатка"
        ws["A1"] = "ФИО"
        ws["B1"] = "Дата подачи"
        ws["C1"] = "Маршрут"
        ws["D1"] = "Стоимость"
        ws["A2"] = "Иванов И.И."
        ws["B2"] = datetime(2024, 1, 15, tzinfo=timezone.utc)
        ws["C2"] = "Москва-Камчатка"
        ws["D2"] = 25000
        ws["A3"] = "Петров П.П."
        ws["B3"] = datetime(2024, 1, 16, tzinfo=timezone.utc)
        ws["C3"] = "Москва-Камчатка"
        ws["D3"] = 30000

        job_id = setup_job(wb)
        settings = ImportSettings(
            color_map={},
            duplicate_strategy="skip",
            sheet_object_map={"Камчатка": setup_object},
        )

        result = service.execute_import(job_id, settings, user_id=1)
        assert result.imported == 2
        assert result.errors == 0
        assert result.skipped == 0

        # Verify DB records
        with db_session_factory() as session:
            requests = session.query(TicketRequest).all()
            assert len(requests) == 2

            traces = session.query(TicketImportRawTrace).all()
            assert len(traces) == 2

            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            assert job.status == "completed"

    def test_import_skips_rows_without_name(self, service, db_session_factory, setup_object, setup_job):
        """Rows without ФИО are counted as errors."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Камчатка"
        ws["A1"] = "ФИО"
        ws["B1"] = "Стоимость"
        ws["A2"] = "Иванов И.И."
        ws["B2"] = 25000
        ws["A3"] = None  # Missing name
        ws["B3"] = 30000

        job_id = setup_job(wb)
        settings = ImportSettings(sheet_object_map={"Камчатка": setup_object})

        result = service.execute_import(job_id, settings, user_id=1)
        assert result.imported == 1
        assert result.errors == 1

    def test_import_handles_losses_sheet(self, service, db_session_factory, setup_object, setup_job):
        """ПОТЕРИ sheet creates TicketFinancialOp records."""
        wb = Workbook()
        ws = wb.active
        ws.title = "ПОТЕРИ"
        ws["A1"] = "ФИО"
        ws["B1"] = "Сумма"
        ws["C1"] = "Причина"
        ws["D1"] = "Дата"
        ws["E1"] = "Тип"
        ws["A2"] = "Иванов И.И."
        ws["B2"] = 5000
        ws["C2"] = "Неявка"
        ws["D2"] = datetime(2024, 3, 1, tzinfo=timezone.utc)
        ws["E2"] = "потеря"

        job_id = setup_job(wb)
        settings = ImportSettings()

        result = service.execute_import(job_id, settings, user_id=1)

        with db_session_factory() as session:
            fin_ops = session.query(TicketFinancialOp).all()
            assert len(fin_ops) == 1
            assert fin_ops[0].amount == Decimal("5000")
            assert fin_ops[0].op_type == "loss"

    def test_import_duplicate_skip_strategy(self, service, db_session_factory, setup_object, setup_job):
        """Duplicate rows are skipped with 'skip' strategy."""
        submitted = datetime(2024, 1, 15, tzinfo=timezone.utc)

        # Pre-create existing record
        with db_session_factory() as session:
            emp = TicketEmployee(id=10, full_name="Иванов И.И.")
            session.add(emp)
            session.flush()
            req = TicketRequest(
                id=100, employee_id=emp.id, object_id=setup_object,
                status="new", submitted_at=submitted, source="import",
            )
            session.add(req)
            session.commit()

        wb = Workbook()
        ws = wb.active
        ws.title = "Камчатка"
        ws["A1"] = "ФИО"
        ws["B1"] = "Дата подачи"
        ws["A2"] = "Иванов И.И."
        ws["B2"] = submitted

        job_id = setup_job(wb)
        settings = ImportSettings(
            duplicate_strategy="skip",
            sheet_object_map={"Камчатка": setup_object},
        )

        result = service.execute_import(job_id, settings, user_id=1)
        assert result.skipped == 1
        assert result.imported == 0

    def test_import_stores_raw_trace_for_errors(self, service, db_session_factory, setup_object, setup_job):
        """Raw traces are stored even for rows that fail parsing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Камчатка"
        ws["A1"] = "ФИО"
        ws["B1"] = "Стоимость"
        ws["A2"] = None  # Will fail
        ws["B2"] = 5000

        job_id = setup_job(wb)
        settings = ImportSettings(sheet_object_map={"Камчатка": setup_object})

        result = service.execute_import(job_id, settings, user_id=1)
        assert result.errors == 1

        with db_session_factory() as session:
            traces = session.query(TicketImportRawTrace).all()
            assert len(traces) == 1  # Error row still gets a trace

    def test_import_reconciliation_in_result_json(self, service, db_session_factory, setup_object, setup_job):
        """Reconciliation report is stored in job result_json."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Камчатка"
        ws["A1"] = "ФИО"
        ws["B1"] = "Стоимость"
        ws["A2"] = "Иванов И.И."
        ws["B2"] = 25000
        ws["A3"] = "Петров П.П."
        ws["B3"] = 30000

        job_id = setup_job(wb)
        settings = ImportSettings(sheet_object_map={"Камчатка": setup_object})

        service.execute_import(job_id, settings, user_id=1)

        with db_session_factory() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            result_data = json.loads(job.result_json)
            assert "reconciliation" in result_data
            recon = result_data["reconciliation"]
            assert recon["rows_imported"] == 2
            assert Decimal(recon["total_cost_imported"]) == Decimal("55000")

    def test_import_skips_empty_and_service_sheets(self, service, db_session_factory, setup_object, setup_job):
        """Empty and service sheets are skipped."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Камчатка"
        ws["A1"] = "ФИО"
        ws["A2"] = "Иванов И.И."

        # Add empty sheet
        ws_empty = wb.create_sheet("ПустойЛист")

        # Add service sheet
        ws_svc = wb.create_sheet("_Настройки")
        ws_svc["A1"] = "param"
        ws_svc["A2"] = "value"

        job_id = setup_job(wb)
        settings = ImportSettings(sheet_object_map={"Камчатка": setup_object})

        result = service.execute_import(job_id, settings, user_id=1)
        assert result.imported == 1  # Only from Камчатка

    def test_import_job_status_failed_on_error(self, service, db_session_factory, setup_job):
        """Job status is set to 'failed' if an exception occurs."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws["A1"] = "data"

        job_id = setup_job(wb)

        # Corrupt the file path to cause an error
        with db_session_factory() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            job.file_path = "/nonexistent/path.xlsx"
            session.commit()

        settings = ImportSettings()
        with pytest.raises(Exception):
            service.execute_import(job_id, settings, user_id=1)

        with db_session_factory() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            assert job.status == "failed"


# ---------------------------------------------------------------------------
# Tests: _has_valid_headers and _build_header_map
# ---------------------------------------------------------------------------


class TestHeaderHelpers:
    """Tests for header detection and mapping helpers."""

    def test_has_valid_headers_true(self, service):
        """Headers with recognized field names return True."""
        headers = ["ФИО", "Дата подачи", "Маршрут", "Стоимость"]
        assert service._has_valid_headers(headers) is True

    def test_has_valid_headers_false(self, service):
        """Headers without recognized names return False."""
        headers = ["Column1", "Column2", "Column3"]
        assert service._has_valid_headers(headers) is False

    def test_build_header_map(self, service):
        """Header map correctly maps indices to field names."""
        headers = ["ФИО", "Дата подачи", "Маршрут", "Стоимость", "Ответственный"]
        result = service._build_header_map(headers)
        assert result[0] == "full_name"
        assert result[1] == "submitted_at"
        assert result[2] == "route"
        assert result[3] == "cost"
        assert result[4] == "assignee"
