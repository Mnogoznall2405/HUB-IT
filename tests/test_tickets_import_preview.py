"""Tests for TicketsImportService — import preview and color mapping (task 6.2).

Tests cover:
- get_preview(): sheet classification, object matching, color extraction, status update
- _extract_unique_colors(): color extraction from worksheet cells
- _suggest_color_mapping(): Euclidean distance-based color→status suggestion
- _map_color_to_status(): confirmed color map lookup
- _extract_headers(): header extraction from first row
- _hex_to_rgb(): hex color to RGB conversion
- _color_distance(): Euclidean distance between colors
- Unmatched sheets detection

Requirements: 1.1, 1.3, 1.7
"""
from __future__ import annotations

import sys
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.appdb.models import AppBase
from backend.appdb.tickets_models import TicketImportJob, TicketObject
from backend.services.tickets_import_service import (
    DEFAULT_REFERENCE_COLORS,
    ImportPreview,
    SheetPreviewInfo,
    TicketsImportService,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _sqlite_url(temp_dir: str) -> str:
    return f"sqlite:///{(Path(temp_dir) / 'tickets_preview.db').as_posix()}"


@pytest.fixture
def db_setup(temp_dir, monkeypatch):
    """Set up a fresh SQLite database and return engine + session factory."""
    import backend.appdb.db as appdb

    url = _sqlite_url(temp_dir)

    appdb._engines.clear()
    appdb._session_factories.clear()
    appdb._initialized_schema_urls.clear()

    monkeypatch.setenv("APP_DATABASE_URL", url)

    try:
        from backend.config import config
        monkeypatch.setattr(config, "app_database_url", url, raising=False)
    except (ImportError, AttributeError):
        pass

    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.close()

    AppBase.metadata.create_all(engine, checkfirst=True)

    SessionLocal = sessionmaker(bind=engine)
    return engine, SessionLocal


@pytest.fixture
def service(temp_dir, db_setup, monkeypatch):
    """Create a TicketsImportService with a fresh SQLite database."""
    engine, SessionLocal = db_setup

    @contextmanager
    def _test_app_session(database_url=None):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    monkeypatch.setattr(
        "backend.services.tickets_import_service.app_session", _test_app_session
    )

    upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
    svc = TicketsImportService(upload_base_dir=str(upload_dir))
    return svc


@pytest.fixture
def service_with_objects(service, db_setup, monkeypatch):
    """Service with pre-seeded TicketObject records for matching."""
    _, SessionLocal = db_setup

    with SessionLocal() as session:
        obj1 = TicketObject(
            id=1, code="KAM", name="Камчатка",
            short_name="Камч", region="ДВ", is_active=True,
        )
        obj2 = TicketObject(
            id=2, code="MAG", name="Магадан",
            short_name="Маг", region="ДВ", is_active=True,
        )
        obj3 = TicketObject(
            id=3, code="TIK", name="Тикси",
            short_name=None, region="Якутия", is_active=True,
        )
        session.add_all([obj1, obj2, obj3])
        session.commit()

    return service


# ---------------------------------------------------------------------------
# Tests: _hex_to_rgb
# ---------------------------------------------------------------------------


class TestHexToRgb:
    """Tests for _hex_to_rgb() static method."""

    def test_red(self):
        assert TicketsImportService._hex_to_rgb("FF0000") == (255, 0, 0)

    def test_green(self):
        assert TicketsImportService._hex_to_rgb("00FF00") == (0, 255, 0)

    def test_blue(self):
        assert TicketsImportService._hex_to_rgb("0000FF") == (0, 0, 255)

    def test_white(self):
        assert TicketsImportService._hex_to_rgb("FFFFFF") == (255, 255, 255)

    def test_black(self):
        assert TicketsImportService._hex_to_rgb("000000") == (0, 0, 0)

    def test_with_hash_prefix(self):
        assert TicketsImportService._hex_to_rgb("#C6EFCE") == (198, 239, 206)

    def test_argb_8_char(self):
        """8-char ARGB string strips alpha."""
        assert TicketsImportService._hex_to_rgb("FFFF0000") == (255, 0, 0)

    def test_lowercase(self):
        assert TicketsImportService._hex_to_rgb("ff0000") == (255, 0, 0)

    def test_invalid_returns_black(self):
        assert TicketsImportService._hex_to_rgb("XYZ") == (0, 0, 0)

    def test_empty_returns_black(self):
        assert TicketsImportService._hex_to_rgb("") == (0, 0, 0)


# ---------------------------------------------------------------------------
# Tests: _color_distance
# ---------------------------------------------------------------------------


class TestColorDistance:
    """Tests for _color_distance() static method."""

    def test_same_color_zero_distance(self):
        assert TicketsImportService._color_distance("FF0000", "FF0000") == 0.0

    def test_black_to_white(self):
        """Max distance is sqrt(255^2 * 3) ≈ 441.67."""
        import math
        dist = TicketsImportService._color_distance("000000", "FFFFFF")
        expected = math.sqrt(255**2 + 255**2 + 255**2)
        assert abs(dist - expected) < 0.01

    def test_red_to_green(self):
        import math
        dist = TicketsImportService._color_distance("FF0000", "00FF00")
        expected = math.sqrt(255**2 + 255**2)
        assert abs(dist - expected) < 0.01

    def test_similar_colors_small_distance(self):
        """Two similar greens should have small distance."""
        dist = TicketsImportService._color_distance("C6EFCE", "C0E8C0")
        assert dist < 30  # Very close colors

    def test_symmetric(self):
        d1 = TicketsImportService._color_distance("FF0000", "00FF00")
        d2 = TicketsImportService._color_distance("00FF00", "FF0000")
        assert d1 == d2


# ---------------------------------------------------------------------------
# Tests: _extract_unique_colors
# ---------------------------------------------------------------------------


class TestExtractUniqueColors:
    """Tests for _extract_unique_colors() method."""

    def test_no_colors(self, service):
        """Worksheet without colored cells returns empty set."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "plain text"
        ws["A2"] = "more text"

        result = service._extract_unique_colors(ws)
        assert result == set()

    def test_single_color(self, service):
        """Single colored cell returns one color."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "colored"
        ws["A1"].fill = PatternFill(
            start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"
        )

        result = service._extract_unique_colors(ws)
        assert "C6EFCE" in result

    def test_multiple_colors(self, service):
        """Multiple different colors are all extracted."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "green"
        ws["A1"].fill = PatternFill(
            start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"
        )
        ws["A2"] = "yellow"
        ws["A2"].fill = PatternFill(
            start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
        )
        ws["A3"] = "red"
        ws["A3"].fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )

        result = service._extract_unique_colors(ws)
        assert "C6EFCE" in result
        assert "FFFF00" in result
        assert "FF0000" in result

    def test_duplicate_colors_deduplicated(self, service):
        """Same color on multiple cells appears only once."""
        wb = Workbook()
        ws = wb.active
        green_fill = PatternFill(
            start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"
        )
        ws["A1"] = "cell1"
        ws["A1"].fill = green_fill
        ws["A2"] = "cell2"
        ws["A2"].fill = green_fill

        result = service._extract_unique_colors(ws)
        assert len(result) == 1
        assert "C6EFCE" in result

    def test_white_excluded(self, service):
        """White fill is excluded from results."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "white"
        ws["A1"].fill = PatternFill(
            start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
        )

        result = service._extract_unique_colors(ws)
        assert "FFFFFF" not in result

    def test_black_excluded(self, service):
        """Black fill (000000) is excluded from results."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "black"
        ws["A1"].fill = PatternFill(
            start_color="FF000000", end_color="FF000000", fill_type="solid"
        )

        result = service._extract_unique_colors(ws)
        assert "000000" not in result


# ---------------------------------------------------------------------------
# Tests: _suggest_color_mapping
# ---------------------------------------------------------------------------


class TestSuggestColorMapping:
    """Tests for _suggest_color_mapping() method."""

    def test_empty_colors_returns_empty(self, service):
        """No unique colors → empty mapping."""
        result = service._suggest_color_mapping([])
        assert result == {}

    def test_empty_reference_returns_empty(self, service):
        """No reference colors → empty mapping."""
        result = service._suggest_color_mapping(["FF0000"], {})
        assert result == {}

    def test_exact_match(self, service):
        """Color that exactly matches a reference gets correct status."""
        result = service._suggest_color_mapping(
            ["FF0000"],
            {"FF0000": "cancelled"},
        )
        assert result == {"FF0000": "cancelled"}

    def test_closest_match_green(self, service):
        """A green-ish color maps to 'purchased' (closest to C6EFCE)."""
        result = service._suggest_color_mapping(
            ["C0E8C0"],  # Similar to light green
            DEFAULT_REFERENCE_COLORS,
        )
        assert result["C0E8C0"] == "purchased"

    def test_closest_match_red(self, service):
        """A red-ish color maps to 'cancelled' (closest to FF0000)."""
        result = service._suggest_color_mapping(
            ["FF3333"],  # Close to red
            DEFAULT_REFERENCE_COLORS,
        )
        assert result["FF3333"] == "cancelled"

    def test_closest_match_yellow(self, service):
        """A yellow-ish color maps to 'in_progress' (closest to FFFF00)."""
        result = service._suggest_color_mapping(
            ["FFFF33"],  # Close to yellow
            DEFAULT_REFERENCE_COLORS,
        )
        assert result["FFFF33"] == "in_progress"

    def test_multiple_colors_mapped(self, service):
        """Multiple colors each get their closest match."""
        result = service._suggest_color_mapping(
            ["FF0000", "00FF00", "0000FF"],
            {"FF0000": "cancelled", "00FF00": "purchased", "0000FF": "data_check"},
        )
        assert result["FF0000"] == "cancelled"
        assert result["00FF00"] == "purchased"
        assert result["0000FF"] == "data_check"

    def test_uses_default_reference_when_none(self, service):
        """When reference_colors is None, uses DEFAULT_REFERENCE_COLORS."""
        result = service._suggest_color_mapping(["FF0000"], None)
        assert "FF0000" in result
        assert result["FF0000"] == "cancelled"


# ---------------------------------------------------------------------------
# Tests: _map_color_to_status
# ---------------------------------------------------------------------------


class TestMapColorToStatus:
    """Tests for _map_color_to_status() static method."""

    def test_found_in_map(self):
        color_map = {"FF0000": "cancelled", "C6EFCE": "purchased"}
        assert TicketsImportService._map_color_to_status("FF0000", color_map) == "cancelled"

    def test_not_found_returns_none(self):
        color_map = {"FF0000": "cancelled"}
        assert TicketsImportService._map_color_to_status("00FF00", color_map) is None

    def test_empty_color_returns_none(self):
        color_map = {"FF0000": "cancelled"}
        assert TicketsImportService._map_color_to_status("", color_map) is None

    def test_empty_map_returns_none(self):
        assert TicketsImportService._map_color_to_status("FF0000", {}) is None

    def test_normalizes_hash_prefix(self):
        color_map = {"FF0000": "cancelled"}
        assert TicketsImportService._map_color_to_status("#FF0000", color_map) == "cancelled"

    def test_normalizes_argb(self):
        """8-char ARGB input is normalized to 6-char."""
        color_map = {"FF0000": "cancelled"}
        assert TicketsImportService._map_color_to_status("FFFF0000", color_map) == "cancelled"

    def test_case_insensitive(self):
        color_map = {"FF0000": "cancelled"}
        assert TicketsImportService._map_color_to_status("ff0000", color_map) == "cancelled"


# ---------------------------------------------------------------------------
# Tests: _extract_headers
# ---------------------------------------------------------------------------


class TestExtractHeaders:
    """Tests for _extract_headers() static method."""

    def test_normal_headers(self, service):
        """Extracts header values from first row."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ФИО"
        ws["B1"] = "Дата"
        ws["C1"] = "Статус"

        headers = service._extract_headers(ws)
        assert headers == ["ФИО", "Дата", "Статус"]

    def test_empty_sheet(self, service):
        """Empty sheet returns empty headers list."""
        wb = Workbook()
        ws = wb.active

        headers = service._extract_headers(ws)
        assert headers == []

    def test_none_cells_become_empty_string(self, service):
        """None values in header row become empty strings."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ФИО"
        ws["B1"] = None
        ws["C1"] = "Дата"

        headers = service._extract_headers(ws)
        assert headers[0] == "ФИО"
        assert headers[1] == ""
        assert headers[2] == "Дата"

    def test_numeric_header_converted_to_string(self, service):
        """Numeric header values are converted to strings."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 123
        ws["B1"] = "Text"

        headers = service._extract_headers(ws)
        assert headers[0] == "123"
        assert headers[1] == "Text"


# ---------------------------------------------------------------------------
# Tests: get_preview
# ---------------------------------------------------------------------------


class TestGetPreview:
    """Tests for get_preview() method."""

    def _create_job_with_workbook(self, service, db_setup, temp_dir, sheets_config):
        """Helper: create a workbook, upload it, and return job_id.

        sheets_config: list of dicts with keys:
            title, rows (list of lists), colors (optional dict cell→ARGB hex)
        """
        _, SessionLocal = db_setup

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for cfg in sheets_config:
            ws = wb.create_sheet(title=cfg["title"])
            for row_idx, row_data in enumerate(cfg.get("rows", []), start=1):
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            # Apply colors
            for cell_ref, argb in cfg.get("colors", {}).items():
                ws[cell_ref].fill = PatternFill(
                    start_color=argb, end_color=argb, fill_type="solid"
                )
            # Set visibility
            if cfg.get("hidden"):
                ws.sheet_state = "hidden"

        # Save workbook
        upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / "test_preview.xlsx"
        wb.save(str(file_path))

        # Create DB record directly
        import uuid as uuid_mod
        job_id = uuid_mod.uuid4().hex
        with SessionLocal() as session:
            job = TicketImportJob(
                id=job_id,
                file_name="test_preview.xlsx",
                file_path=str(file_path),
                status="uploaded",
                user_id=1,
            )
            session.add(job)
            session.commit()

        return job_id

    def test_preview_returns_import_preview(self, service, db_setup, temp_dir):
        """get_preview returns an ImportPreview dataclass."""
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{"title": "Sheet1", "rows": [["ФИО", "Дата"], ["Иванов", "2024-01-01"]]}],
        )
        result = service.get_preview(job_id)
        assert isinstance(result, ImportPreview)
        assert result.job_id == job_id

    def test_preview_classifies_sheets(self, service, db_setup, temp_dir):
        """Preview correctly classifies sheets."""
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [
                {"title": "Рабочий", "rows": [["A", "B"], ["1", "2"]]},
                {"title": "Пустой", "rows": []},
                {"title": "ПОТЕРИ", "rows": [["A"], ["1"]]},
            ],
        )
        result = service.get_preview(job_id)
        sheets_by_title = {s.title: s for s in result.sheets}

        assert sheets_by_title["Рабочий"].classification == "рабочий"
        assert sheets_by_title["Пустой"].classification == "пустой"
        assert sheets_by_title["ПОТЕРИ"].classification == "ПОТЕРИ"

    def test_preview_extracts_headers(self, service, db_setup, temp_dir):
        """Preview extracts headers from first row."""
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{"title": "Data", "rows": [["ФИО", "Дата", "Статус"], ["Иванов", "2024", "new"]]}],
        )
        result = service.get_preview(job_id)
        assert result.sheets[0].headers == ["ФИО", "Дата", "Статус"]

    def test_preview_extracts_colors(self, service, db_setup, temp_dir):
        """Preview extracts unique colors from working sheets."""
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{
                "title": "Colored",
                "rows": [["ФИО", "Дата"], ["Иванов", "2024"]],
                "colors": {"A2": "FFC6EFCE", "B2": "FFFF0000"},
            }],
        )
        result = service.get_preview(job_id)
        assert "C6EFCE" in result.unique_colors
        assert "FF0000" in result.unique_colors

    def test_preview_suggests_color_mapping(self, service, db_setup, temp_dir):
        """Preview suggests color→status mapping."""
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{
                "title": "Colored",
                "rows": [["ФИО"], ["Иванов"]],
                "colors": {"A2": "FFFF0000"},
            }],
        )
        result = service.get_preview(job_id)
        assert "FF0000" in result.color_map
        assert result.color_map["FF0000"] == "cancelled"

    def test_preview_matches_objects_by_name(
        self, service_with_objects, db_setup, temp_dir
    ):
        """Preview matches sheet title to TicketObject by name."""
        job_id = self._create_job_with_workbook(
            service_with_objects, db_setup, temp_dir,
            [{"title": "Камчатка", "rows": [["ФИО"], ["Иванов"]]}],
        )
        result = service_with_objects.get_preview(job_id)
        sheet = result.sheets[0]
        assert sheet.matched_object_id == 1
        assert sheet.matched_object_name == "Камчатка"

    def test_preview_matches_objects_by_code(
        self, service_with_objects, db_setup, temp_dir
    ):
        """Preview matches sheet title to TicketObject by code."""
        job_id = self._create_job_with_workbook(
            service_with_objects, db_setup, temp_dir,
            [{"title": "KAM", "rows": [["ФИО"], ["Иванов"]]}],
        )
        result = service_with_objects.get_preview(job_id)
        sheet = result.sheets[0]
        assert sheet.matched_object_id == 1

    def test_preview_unmatched_sheets(
        self, service_with_objects, db_setup, temp_dir
    ):
        """Unmatched working sheets are listed in unmatched_sheets."""
        job_id = self._create_job_with_workbook(
            service_with_objects, db_setup, temp_dir,
            [
                {"title": "Камчатка", "rows": [["ФИО"], ["Иванов"]]},
                {"title": "НеизвестныйОбъект", "rows": [["ФИО"], ["Петров"]]},
            ],
        )
        result = service_with_objects.get_preview(job_id)
        assert "НеизвестныйОбъект" in result.unmatched_sheets
        assert "Камчатка" not in result.unmatched_sheets

    def test_preview_updates_job_status(self, service, db_setup, temp_dir):
        """get_preview updates job status to 'previewed'."""
        _, SessionLocal = db_setup
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{"title": "Sheet1", "rows": [["A"], ["1"]]}],
        )
        service.get_preview(job_id)

        with SessionLocal() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            assert job.status == "previewed"

    def test_preview_job_not_found_raises(self, service):
        """get_preview raises ValueError for non-existent job."""
        with pytest.raises(ValueError, match="not found"):
            service.get_preview("nonexistent_id")

    def test_preview_wrong_status_raises(self, service, db_setup, temp_dir):
        """get_preview raises ValueError if job is not in 'uploaded' status."""
        _, SessionLocal = db_setup
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{"title": "Sheet1", "rows": [["A"], ["1"]]}],
        )
        # Manually set status to 'running'
        with SessionLocal() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            job.status = "running"
            session.commit()

        with pytest.raises(ValueError, match="status"):
            service.get_preview(job_id)

    def test_preview_non_working_sheets_not_in_unmatched(
        self, service_with_objects, db_setup, temp_dir
    ):
        """Non-working sheets (empty, hidden, ПОТЕРИ) are not in unmatched_sheets."""
        job_id = self._create_job_with_workbook(
            service_with_objects, db_setup, temp_dir,
            [
                {"title": "ПОТЕРИ", "rows": [["A"], ["1"]]},
                {"title": "Пустой", "rows": []},
                {"title": "Скрытый", "rows": [["A"], ["1"]], "hidden": True},
            ],
        )
        result = service_with_objects.get_preview(job_id)
        assert result.unmatched_sheets == []

    def test_preview_case_insensitive_matching(
        self, service_with_objects, db_setup, temp_dir
    ):
        """Object matching is case-insensitive."""
        job_id = self._create_job_with_workbook(
            service_with_objects, db_setup, temp_dir,
            [{"title": "камчатка", "rows": [["ФИО"], ["Иванов"]]}],
        )
        result = service_with_objects.get_preview(job_id)
        sheet = result.sheets[0]
        assert sheet.matched_object_id == 1

    def test_preview_stores_preview_json(self, service, db_setup, temp_dir):
        """get_preview stores preview data as JSON in the job record."""
        import json
        _, SessionLocal = db_setup
        job_id = self._create_job_with_workbook(
            service, db_setup, temp_dir,
            [{"title": "Sheet1", "rows": [["ФИО", "Дата"], ["Иванов", "2024"]]}],
        )
        service.get_preview(job_id)

        with SessionLocal() as session:
            job = session.query(TicketImportJob).filter_by(id=job_id).first()
            preview_data = json.loads(job.preview_json)
            assert "sheets" in preview_data
            assert "color_map" in preview_data
            assert "unmatched_sheets" in preview_data
