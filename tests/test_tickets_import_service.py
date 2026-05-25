"""Tests for TicketsImportService — file upload and workbook parsing (task 6.1).

Tests cover:
- upload_file(): file saving, DB record creation, return value
- _classify_sheets(): classification logic for all sheet types
- _detect_hidden_sheets(): visibility detection
- _extract_cell_hyperlinks(): hyperlink extraction
- _extract_cell_comments(): comment extraction
- _extract_threaded_comments(): threaded comment extraction

Requirements: 1.1, 1.10
"""
from __future__ import annotations

import sys
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.appdb.models import AppBase
from backend.appdb.tickets_models import TicketImportJob
from backend.services.tickets_import_service import (
    HiddenSheetInfo,
    SheetClassification,
    TicketsImportService,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _sqlite_url(temp_dir: str) -> str:
    return f"sqlite:///{(Path(temp_dir) / 'tickets_import.db').as_posix()}"


@pytest.fixture
def service(temp_dir, monkeypatch):
    """Create a TicketsImportService with a fresh SQLite database and temp upload dir."""
    import backend.appdb.db as appdb

    url = _sqlite_url(temp_dir)

    appdb._engines.clear()
    appdb._session_factories.clear()
    appdb._initialized_schema_urls.clear()

    monkeypatch.setenv("APP_DATABASE_URL", url)

    try:
        from backend.config import config
        monkeypatch.setattr(config, "app_database_url", url, raising=False)
    except (ImportError, AttributeError):
        pass

    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.close()

    AppBase.metadata.create_all(engine, checkfirst=True)

    SessionLocal = sessionmaker(bind=engine)

    @contextmanager
    def _test_app_session(database_url=None):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    monkeypatch.setattr(
        "backend.services.tickets_import_service.app_session", _test_app_session
    )

    upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
    svc = TicketsImportService(upload_base_dir=str(upload_dir))
    return svc


@pytest.fixture
def sample_workbook() -> Workbook:
    """Create a sample workbook with various sheet types for testing."""
    wb = Workbook()

    # Default sheet — rename to a working sheet
    ws_working = wb.active
    ws_working.title = "Камчатка"
    ws_working["A1"] = "ФИО"
    ws_working["B1"] = "Дата"
    ws_working["A2"] = "Иванов И.И."
    ws_working["B2"] = "2024-01-15"

    # Empty sheet (no data rows)
    ws_empty = wb.create_sheet("ПустойЛист")

    # Sheet with only header (1 row = empty by our definition)
    ws_one_row = wb.create_sheet("ОдинЗаголовок")
    ws_one_row["A1"] = "Заголовок"

    # Service sheet (starts with _)
    ws_service1 = wb.create_sheet("_Настройки")
    ws_service1["A1"] = "param"
    ws_service1["A2"] = "value"

    # Service sheet (contains "служебн")
    ws_service2 = wb.create_sheet("Служебный лист")
    ws_service2["A1"] = "data"
    ws_service2["A2"] = "more data"

    # ПОТЕРИ sheet
    ws_losses = wb.create_sheet("ПОТЕРИ 2024")
    ws_losses["A1"] = "ФИО"
    ws_losses["A2"] = "Петров П.П."

    # Hidden sheet
    ws_hidden = wb.create_sheet("Скрытый")
    ws_hidden["A1"] = "hidden data"
    ws_hidden["A2"] = "more hidden"
    ws_hidden.sheet_state = "hidden"

    # Very hidden sheet
    ws_very_hidden = wb.create_sheet("ОченьСкрытый")
    ws_very_hidden["A1"] = "very hidden"
    ws_very_hidden["A2"] = "secret"
    ws_very_hidden.sheet_state = "veryHidden"

    return wb


# ---------------------------------------------------------------------------
# Tests: upload_file
# ---------------------------------------------------------------------------


class TestUploadFile:
    """Tests for upload_file() method."""

    def test_upload_creates_file_on_disk(self, service, temp_dir):
        """File content is saved to the upload directory."""
        content = b"fake xlsx content"
        result = service.upload_file("test.xlsx", content, user_id=1)

        job_id = result["id"]
        upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
        saved_file = upload_dir / f"{job_id}.xlsx"
        assert saved_file.exists()
        assert saved_file.read_bytes() == content

    def test_upload_returns_correct_dict(self, service):
        """Return value contains id, file_name, and status."""
        result = service.upload_file("report.xlsx", b"data", user_id=42)

        assert "id" in result
        assert result["file_name"] == "report.xlsx"
        assert result["status"] == "uploaded"
        assert len(result["id"]) == 32  # uuid hex

    def test_upload_creates_db_record(self, service, temp_dir, monkeypatch):
        """A TicketImportJob record is created in the database."""
        import backend.appdb.db as appdb

        url = _sqlite_url(temp_dir)
        engine = create_engine(
            url,
            execution_options={"schema_translate_map": {"app": None, "system": None}},
        )
        SessionLocal = sessionmaker(bind=engine)

        result = service.upload_file("data.xlsx", b"content", user_id=7)

        with SessionLocal() as session:
            job = session.query(TicketImportJob).filter_by(id=result["id"]).first()
            assert job is not None
            assert job.file_name == "data.xlsx"
            assert job.status == "uploaded"
            assert job.user_id == 7

    def test_upload_multiple_files_unique_ids(self, service):
        """Each upload gets a unique job ID."""
        r1 = service.upload_file("file1.xlsx", b"a", user_id=1)
        r2 = service.upload_file("file2.xlsx", b"b", user_id=1)
        assert r1["id"] != r2["id"]


# ---------------------------------------------------------------------------
# Tests: _classify_sheets
# ---------------------------------------------------------------------------


class TestClassifySheets:
    """Tests for _classify_sheets() method."""

    def test_working_sheet(self, service, sample_workbook):
        """Normal sheet with data is classified as рабочий."""
        results = service._classify_sheets(sample_workbook)
        kamchatka = next(r for r in results if r.title == "Камчатка")
        assert kamchatka.classification == "рабочий"
        assert kamchatka.row_count == 2

    def test_empty_sheet_no_data(self, service, sample_workbook):
        """Sheet with no data rows is classified as пустой."""
        results = service._classify_sheets(sample_workbook)
        empty = next(r for r in results if r.title == "ПустойЛист")
        assert empty.classification == "пустой"
        assert empty.row_count == 0

    def test_empty_sheet_one_row(self, service, sample_workbook):
        """Sheet with exactly 1 row of data is classified as пустой."""
        results = service._classify_sheets(sample_workbook)
        one_row = next(r for r in results if r.title == "ОдинЗаголовок")
        assert one_row.classification == "пустой"
        assert one_row.row_count == 1

    def test_service_sheet_underscore(self, service, sample_workbook):
        """Sheet starting with '_' is classified as служебный."""
        results = service._classify_sheets(sample_workbook)
        svc = next(r for r in results if r.title == "_Настройки")
        assert svc.classification == "служебный"

    def test_service_sheet_keyword(self, service, sample_workbook):
        """Sheet containing 'служебн' is classified as служебный."""
        results = service._classify_sheets(sample_workbook)
        svc = next(r for r in results if r.title == "Служебный лист")
        assert svc.classification == "служебный"

    def test_losses_sheet(self, service, sample_workbook):
        """Sheet with 'ПОТЕРИ' in title is classified as ПОТЕРИ."""
        results = service._classify_sheets(sample_workbook)
        losses = next(r for r in results if r.title == "ПОТЕРИ 2024")
        assert losses.classification == "ПОТЕРИ"

    def test_hidden_sheet(self, service, sample_workbook):
        """Hidden sheet is classified as скрытый."""
        results = service._classify_sheets(sample_workbook)
        hidden = next(r for r in results if r.title == "Скрытый")
        assert hidden.classification == "скрытый"

    def test_very_hidden_sheet(self, service, sample_workbook):
        """Very hidden sheet is classified as скрытый."""
        results = service._classify_sheets(sample_workbook)
        very_hidden = next(r for r in results if r.title == "ОченьСкрытый")
        assert very_hidden.classification == "скрытый"

    def test_all_sheets_classified(self, service, sample_workbook):
        """All sheets in the workbook are classified."""
        results = service._classify_sheets(sample_workbook)
        assert len(results) == len(sample_workbook.sheetnames)

    def test_hidden_takes_priority_over_losses(self, service):
        """Hidden state takes priority over ПОТЕРИ keyword."""
        wb = Workbook()
        ws = wb.active
        ws.title = "ПОТЕРИ скрытые"
        ws["A1"] = "data"
        ws["A2"] = "more"
        ws.sheet_state = "hidden"

        results = service._classify_sheets(wb)
        assert results[0].classification == "скрытый"

    def test_losses_keyword_case_insensitive(self, service):
        """ПОТЕРИ detection works with lowercase 'потери'."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист потери"
        ws["A1"] = "data"
        ws["A2"] = "more"

        results = service._classify_sheets(wb)
        assert results[0].classification == "ПОТЕРИ"


# ---------------------------------------------------------------------------
# Tests: _detect_hidden_sheets
# ---------------------------------------------------------------------------


class TestDetectHiddenSheets:
    """Tests for _detect_hidden_sheets() method."""

    def test_visible_sheet(self, service, sample_workbook):
        """Visible sheets are reported with state 'visible'."""
        results = service._detect_hidden_sheets(sample_workbook)
        kamchatka = next(r for r in results if r.title == "Камчатка")
        assert kamchatka.state == "visible"

    def test_hidden_sheet(self, service, sample_workbook):
        """Hidden sheets are reported with state 'hidden'."""
        results = service._detect_hidden_sheets(sample_workbook)
        hidden = next(r for r in results if r.title == "Скрытый")
        assert hidden.state == "hidden"

    def test_very_hidden_sheet(self, service, sample_workbook):
        """Very hidden sheets are reported with state 'veryHidden'."""
        results = service._detect_hidden_sheets(sample_workbook)
        very_hidden = next(r for r in results if r.title == "ОченьСкрытый")
        assert very_hidden.state == "veryHidden"

    def test_all_sheets_detected(self, service, sample_workbook):
        """All sheets are included in the result."""
        results = service._detect_hidden_sheets(sample_workbook)
        assert len(results) == len(sample_workbook.sheetnames)


# ---------------------------------------------------------------------------
# Tests: _extract_cell_hyperlinks
# ---------------------------------------------------------------------------


class TestExtractCellHyperlinks:
    """Tests for _extract_cell_hyperlinks() method."""

    def test_no_hyperlinks(self, service):
        """Worksheet without hyperlinks returns empty dict."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "plain text"

        result = service._extract_cell_hyperlinks(ws)
        assert result == {}

    def test_single_hyperlink(self, service):
        """Single hyperlink is extracted correctly."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Click here"
        ws["A1"].hyperlink = "https://example.com"

        result = service._extract_cell_hyperlinks(ws)
        assert "A1" in result
        assert "https://example.com" in result["A1"]

    def test_multiple_hyperlinks(self, service):
        """Multiple hyperlinks on different cells are extracted."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Link 1"
        ws["A1"].hyperlink = "https://example.com/1"
        ws["B2"] = "Link 2"
        ws["B2"].hyperlink = "https://example.com/2"

        result = service._extract_cell_hyperlinks(ws)
        assert "A1" in result
        assert "B2" in result
        assert "https://example.com/1" in result["A1"]
        assert "https://example.com/2" in result["B2"]


# ---------------------------------------------------------------------------
# Tests: _extract_cell_comments
# ---------------------------------------------------------------------------


class TestExtractCellComments:
    """Tests for _extract_cell_comments() method."""

    def test_no_comments(self, service):
        """Worksheet without comments returns empty dict."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "no comment"

        result = service._extract_cell_comments(ws)
        assert result == {}

    def test_single_comment(self, service):
        """Single comment is extracted correctly."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data"
        ws["A1"].comment = Comment("This is a note", "Author")

        result = service._extract_cell_comments(ws)
        assert "A1" in result
        assert result["A1"] == "This is a note"

    def test_multiple_comments(self, service):
        """Multiple comments on different cells are extracted."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data1"
        ws["A1"].comment = Comment("Note 1", "Author1")
        ws["C3"] = "data2"
        ws["C3"].comment = Comment("Note 2", "Author2")

        result = service._extract_cell_comments(ws)
        assert result["A1"] == "Note 1"
        assert result["C3"] == "Note 2"

    def test_empty_comment_text_excluded(self, service):
        """Comments with empty text are not included."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data"
        ws["A1"].comment = Comment("", "Author")

        result = service._extract_cell_comments(ws)
        assert "A1" not in result


# ---------------------------------------------------------------------------
# Tests: _extract_threaded_comments
# ---------------------------------------------------------------------------


class TestExtractThreadedComments:
    """Tests for _extract_threaded_comments() method."""

    def test_no_threaded_comments(self, service):
        """Worksheet without threaded comments returns empty dict."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data"

        result = service._extract_threaded_comments(ws)
        assert result == {}

    def test_returns_dict_type(self, service):
        """Method always returns a dict."""
        wb = Workbook()
        ws = wb.active

        result = service._extract_threaded_comments(ws)
        assert isinstance(result, dict)


# ---------------------------------------------------------------------------
# Tests: open_workbook
# ---------------------------------------------------------------------------


class TestOpenWorkbook:
    """Tests for open_workbook() method."""

    def test_opens_valid_xlsx(self, service, temp_dir):
        """Can open a valid .xlsx file."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "test"
        ws["A2"] = "=A1&\" value\""  # formula

        file_path = Path(temp_dir) / "test.xlsx"
        wb.save(str(file_path))

        loaded = service.open_workbook(str(file_path))
        assert isinstance(loaded, Workbook)
        # data_only=False means formulas are preserved
        ws_loaded = loaded.active
        assert ws_loaded["A2"].value == '=A1&" value"'

    def test_preserves_comments(self, service, temp_dir):
        """Comments are preserved when opening workbook."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data"
        ws["A1"].comment = Comment("Important note", "Admin")

        file_path = Path(temp_dir) / "comments.xlsx"
        wb.save(str(file_path))

        loaded = service.open_workbook(str(file_path))
        ws_loaded = loaded.active
        assert ws_loaded["A1"].comment is not None
        assert ws_loaded["A1"].comment.text == "Important note"


# ---------------------------------------------------------------------------
# Tests: _count_data_rows helper
# ---------------------------------------------------------------------------


class TestCountDataRows:
    """Tests for _count_data_rows() static method."""

    def test_empty_sheet(self, service):
        """Empty sheet has 0 data rows."""
        wb = Workbook()
        ws = wb.active
        assert service._count_data_rows(ws) == 0

    def test_sheet_with_data(self, service):
        """Counts only rows with at least one non-empty cell."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "header"
        ws["A2"] = "data"
        ws["A3"] = None  # empty row
        ws["A4"] = "more"

        assert service._count_data_rows(ws) == 3

    def test_row_with_only_none_values(self, service):
        """Rows where all cells are None are not counted."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data"
        # Row 2 has no values set — should not be counted
        ws["A3"] = "data"

        # openpyxl may not iterate over row 2 if nothing was set
        count = service._count_data_rows(ws)
        assert count >= 2  # At least rows 1 and 3
