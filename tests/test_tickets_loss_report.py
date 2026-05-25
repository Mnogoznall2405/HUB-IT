"""Tests for TicketsService loss report and export methods (task 8.1).

Tests cover:
- get_losses_report() with period filtering, object filter, op_type filter
- Default period (current month) when no dates provided
- Max 366 days period validation
- Totals calculation: total_losses, total_refunds, balance (Decimal arithmetic)
- Pagination (max 50 per page)
- export_losses_xlsx() — generates valid .xlsx with correct columns and data
- export_requests_xlsx() — generates valid .xlsx with correct columns and data
- Max 50,000 records limit for exports

Requirements: 11.1, 11.2, 11.3, 11.4, 11.5, 11.6
"""
from __future__ import annotations

import sys
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.appdb.models import AppBase
from backend.appdb.tickets_models import (
    TicketEmployee,
    TicketFinancialOp,
    TicketObject,
    TicketRequest,
)
from backend.services.tickets_service import (
    Pagination,
    RequestFilters,
    TicketsService,
    TicketsValidationError,
)


def _sqlite_url(temp_dir: str) -> str:
    return f"sqlite:///{(Path(temp_dir) / 'tickets_loss_report.db').as_posix()}"


@pytest.fixture
def service(temp_dir, monkeypatch):
    """Create a TicketsService with a fresh SQLite database."""
    from sqlalchemy import create_engine, event
    from sqlalchemy.orm import sessionmaker

    url = _sqlite_url(temp_dir)

    import backend.appdb.db as appdb
    appdb._engines.clear()
    appdb._session_factories.clear()
    appdb._initialized_schema_urls.clear()

    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    AppBase.metadata.create_all(engine, checkfirst=True)

    SessionLocal = sessionmaker(bind=engine)

    @contextmanager
    def _test_app_session(database_url=None):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    monkeypatch.setattr("backend.services.tickets_service.app_session", _test_app_session)

    return TicketsService()


@pytest.fixture
def seed_data(service, temp_dir):
    """Seed reference data and financial operations for loss report tests."""
    from sqlalchemy import create_engine, event
    from sqlalchemy.orm import sessionmaker

    url = _sqlite_url(temp_dir)
    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )
    SessionLocal = sessionmaker(bind=engine)

    @contextmanager
    def _session():
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    now = datetime.now(timezone.utc)

    with _session() as session:
        # Create two objects
        obj1 = TicketObject(
            code="KAM", name="Камчатка", region="Дальний Восток",
            is_active=True, created_at=now, updated_at=now,
        )
        obj2 = TicketObject(
            code="MAG", name="Магадан", region="Дальний Восток",
            is_active=True, created_at=now, updated_at=now,
        )
        session.add_all([obj1, obj2])
        session.flush()
        obj1_id, obj2_id = obj1.id, obj2.id

        # Create employees
        emp1 = TicketEmployee(
            full_name="Иванов Иван Иванович", date_of_birth_enc="",
            status="active", created_at=now, updated_at=now,
        )
        emp2 = TicketEmployee(
            full_name="Петров Пётр Петрович", date_of_birth_enc="",
            status="active", created_at=now, updated_at=now,
        )
        session.add_all([emp1, emp2])
        session.flush()
        emp1_id, emp2_id = emp1.id, emp2.id

        # Create a request for export test
        req = TicketRequest(
            employee_id=emp1_id, object_id=obj1_id, status="purchased",
            total_cost=Decimal("15000.00"), is_urgent=False, needs_review=False,
            source="manual", version=1, created_at=now, updated_at=now,
            departure_date=now + timedelta(days=5),
            arrival_date=now + timedelta(days=10),
        )
        session.add(req)
        session.flush()
        req_id = req.id

        # Create financial operations within current month
        base_date = now.replace(day=5, hour=12, minute=0, second=0, microsecond=0)

        ops = [
            TicketFinancialOp(
                employee_id=emp1_id, object_id=obj1_id, op_type="loss",
                amount=Decimal("5000.50"), reason="Неявка", refund_status=None,
                op_date=base_date, is_deleted=False, created_at=now, updated_at=now,
            ),
            TicketFinancialOp(
                employee_id=emp1_id, object_id=obj1_id, op_type="refund",
                amount=Decimal("3000.25"), reason="Отмена рейса", refund_status="completed",
                op_date=base_date + timedelta(days=1), is_deleted=False, created_at=now, updated_at=now,
            ),
            TicketFinancialOp(
                employee_id=emp2_id, object_id=obj2_id, op_type="loss",
                amount=Decimal("7500.00"), reason="Опоздание", refund_status=None,
                op_date=base_date + timedelta(days=2), is_deleted=False, created_at=now, updated_at=now,
            ),
            TicketFinancialOp(
                employee_id=emp2_id, object_id=obj2_id, op_type="exchange",
                amount=Decimal("1200.00"), reason="Перенос даты", refund_status=None,
                op_date=base_date + timedelta(days=3), is_deleted=False, created_at=now, updated_at=now,
            ),
            # Deleted op — should not appear
            TicketFinancialOp(
                employee_id=emp1_id, object_id=obj1_id, op_type="loss",
                amount=Decimal("9999.99"), reason="Удалённая", refund_status=None,
                op_date=base_date + timedelta(days=1), is_deleted=True, created_at=now, updated_at=now,
            ),
        ]
        session.add_all(ops)
        session.flush()

    return {
        "object1_id": obj1_id,
        "object2_id": obj2_id,
        "employee1_id": emp1_id,
        "employee2_id": emp2_id,
        "request_id": req_id,
        "base_date": base_date,
        "now": now,
    }


# ---------------------------------------------------------------------------
# get_losses_report tests
# ---------------------------------------------------------------------------


class TestGetLossesReport:
    """Tests for get_losses_report()."""

    def test_default_period_returns_current_month(self, service, seed_data):
        """With no filters, returns data for current month."""
        result = service.get_losses_report()
        assert "items" in result
        assert "totals" in result
        assert "pagination" in result
        # Should have 4 non-deleted ops (all within current month)
        assert result["pagination"]["total"] == 4

    def test_totals_calculation_decimal_precision(self, service, seed_data):
        """Totals use Decimal arithmetic: losses, refunds, balance."""
        result = service.get_losses_report()
        totals = result["totals"]
        # total_losses = 5000.50 + 7500.00 = 12500.50
        assert Decimal(totals["total_losses"]) == Decimal("12500.50")
        # total_refunds = 3000.25
        assert Decimal(totals["total_refunds"]) == Decimal("3000.25")
        # balance = losses - refunds = 12500.50 - 3000.25 = 9500.25
        assert Decimal(totals["balance"]) == Decimal("9500.25")

    def test_filter_by_object(self, service, seed_data):
        """Filter by object_id returns only ops for that object."""
        result = service.get_losses_report(filters={"object_id": seed_data["object1_id"]})
        # Object 1 has: loss 5000.50, refund 3000.25
        assert result["pagination"]["total"] == 2
        assert Decimal(result["totals"]["total_losses"]) == Decimal("5000.50")
        assert Decimal(result["totals"]["total_refunds"]) == Decimal("3000.25")

    def test_filter_by_op_type(self, service, seed_data):
        """Filter by op_type returns only matching operations."""
        result = service.get_losses_report(filters={"op_type": "loss"})
        assert result["pagination"]["total"] == 2
        # All items should be losses
        for item in result["items"]:
            assert item["op_type"] == "loss"

    def test_filter_by_period(self, service, seed_data):
        """Filter by explicit date range."""
        base = seed_data["base_date"]
        # Only include first 2 days
        result = service.get_losses_report(filters={
            "date_from": base.isoformat(),
            "date_to": (base + timedelta(days=2)).isoformat(),
        })
        # Should include ops on day 5 and day 6 (base_date and base_date+1)
        assert result["pagination"]["total"] == 2

    def test_max_period_validation(self, service, seed_data):
        """Period exceeding 366 days raises validation error."""
        now = datetime.now(timezone.utc)
        with pytest.raises(TicketsValidationError):
            service.get_losses_report(filters={
                "date_from": (now - timedelta(days=400)).strftime("%Y-%m-%d"),
                "date_to": now.strftime("%Y-%m-%d"),
            })

    def test_pagination_max_50(self, service, seed_data):
        """Page size is capped at 50 for loss report."""
        result = service.get_losses_report(
            pagination=Pagination(page=1, page_size=100)
        )
        assert result["pagination"]["page_size"] == 50

    def test_deleted_ops_excluded(self, service, seed_data):
        """Soft-deleted operations are not included in the report."""
        result = service.get_losses_report()
        # The deleted op (9999.99) should not be in totals
        assert Decimal(result["totals"]["total_losses"]) == Decimal("12500.50")
        # And not in items
        for item in result["items"]:
            assert Decimal(item["amount"]) != Decimal("9999.99")

    def test_items_contain_required_fields(self, service, seed_data):
        """Each item has all required fields for the report."""
        result = service.get_losses_report()
        required_fields = {"id", "employee_name", "object_name", "op_date", "amount", "reason", "op_type", "refund_status"}
        for item in result["items"]:
            assert required_fields.issubset(item.keys())

    def test_empty_result(self, service, seed_data):
        """Returns empty items and zero totals when no data matches."""
        # Use a date range far in the past
        result = service.get_losses_report(filters={
            "date_from": "2020-01-01",
            "date_to": "2020-02-01",
        })
        assert result["items"] == []
        assert result["pagination"]["total"] == 0
        assert Decimal(result["totals"]["total_losses"]) == Decimal("0.00")
        assert Decimal(result["totals"]["total_refunds"]) == Decimal("0.00")
        assert Decimal(result["totals"]["balance"]) == Decimal("0.00")


# ---------------------------------------------------------------------------
# export_losses_xlsx tests
# ---------------------------------------------------------------------------


class TestExportLossesXlsx:
    """Tests for export_losses_xlsx()."""

    def test_returns_valid_xlsx_bytes(self, service, seed_data):
        """Returns bytes that can be loaded as a valid .xlsx workbook."""
        from openpyxl import load_workbook

        data = service.export_losses_xlsx()
        assert isinstance(data, bytes)
        assert len(data) > 0

        wb = load_workbook(BytesIO(data))
        ws = wb.active
        assert ws.title == "Потери и возвраты"

    def test_xlsx_has_correct_headers(self, service, seed_data):
        """The .xlsx file has the correct header row."""
        from openpyxl import load_workbook

        data = service.export_losses_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        assert headers == ["ФИО", "Объект", "Дата", "Сумма", "Причина", "Тип операции", "Статус возврата"]

    def test_xlsx_contains_data_rows(self, service, seed_data):
        """The .xlsx file contains the correct number of data rows."""
        from openpyxl import load_workbook

        data = service.export_losses_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        # 1 header + 4 data rows
        assert ws.max_row == 5

    def test_xlsx_data_correctness(self, service, seed_data):
        """Data in the .xlsx matches the financial operations."""
        from openpyxl import load_workbook

        data = service.export_losses_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        # Check that employee names appear in the data
        names = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            names.add(row[0])
        assert "Иванов Иван Иванович" in names
        assert "Петров Пётр Петрович" in names

    def test_xlsx_filter_by_object(self, service, seed_data):
        """Export respects object_id filter."""
        from openpyxl import load_workbook

        data = service.export_losses_xlsx(filters={"object_id": seed_data["object1_id"]})
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        # 1 header + 2 data rows (object1 has 2 non-deleted ops)
        assert ws.max_row == 3

    def test_xlsx_max_period_validation(self, service, seed_data):
        """Export raises validation error for period > 366 days."""
        now = datetime.now(timezone.utc)
        with pytest.raises(TicketsValidationError):
            service.export_losses_xlsx(filters={
                "date_from": (now - timedelta(days=400)).strftime("%Y-%m-%d"),
                "date_to": now.strftime("%Y-%m-%d"),
            })

    def test_xlsx_op_type_labels(self, service, seed_data):
        """Op types are displayed as Russian labels."""
        from openpyxl import load_workbook

        data = service.export_losses_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        op_types = set()
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
            op_types.add(row[0])
        assert op_types.issubset({"Возврат", "Обмен", "Потеря"})


# ---------------------------------------------------------------------------
# export_requests_xlsx tests
# ---------------------------------------------------------------------------


class TestExportRequestsXlsx:
    """Tests for export_requests_xlsx()."""

    def test_returns_valid_xlsx_bytes(self, service, seed_data):
        """Returns bytes that can be loaded as a valid .xlsx workbook."""
        from openpyxl import load_workbook

        data = service.export_requests_xlsx()
        assert isinstance(data, bytes)
        assert len(data) > 0

        wb = load_workbook(BytesIO(data))
        ws = wb.active
        assert ws.title == "Заявки"

    def test_xlsx_has_correct_headers(self, service, seed_data):
        """The .xlsx file has the correct header row."""
        from openpyxl import load_workbook

        data = service.export_requests_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        assert headers == ["№", "ФИО", "Объект", "Дата вылета", "Дата прибытия", "Статус", "Ответственный", "Стоимость"]

    def test_xlsx_contains_request_data(self, service, seed_data):
        """The .xlsx file contains the seeded request."""
        from openpyxl import load_workbook

        data = service.export_requests_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        # 1 header + 1 request
        assert ws.max_row == 2

        # Check data in first data row
        row = [cell.value for cell in ws[2]]
        assert row[0] == seed_data["request_id"]  # №
        assert row[1] == "Иванов Иван Иванович"  # ФИО
        assert row[2] == "Камчатка"  # Объект
        assert row[5] == "Билет куплен"  # Статус (purchased → Билет куплен)
        assert row[7] == 15000.0  # Стоимость

    def test_xlsx_respects_filters(self, service, seed_data):
        """Export respects RequestFilters."""
        from openpyxl import load_workbook

        # Filter by a non-existent status — should return empty
        filters = RequestFilters(statuses=["archive"])
        data = service.export_requests_xlsx(filters=filters)
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        # Only header row
        assert ws.max_row == 1

    def test_xlsx_status_labels_in_russian(self, service, seed_data):
        """Status values are displayed as Russian labels."""
        from openpyxl import load_workbook

        data = service.export_requests_xlsx()
        wb = load_workbook(BytesIO(data))
        ws = wb.active

        # The request has status "purchased" which should be "Билет куплен"
        status_cell = ws.cell(row=2, column=6).value
        assert status_cell == "Билет куплен"
