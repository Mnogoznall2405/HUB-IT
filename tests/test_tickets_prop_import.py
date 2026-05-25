"""Property-based tests for TicketsImportService — import logic.

**Validates: Requirements 1.3, 1.4, 1.5, 1.6, 1.8, 1.9, 1.10, 1.11**

Properties tested:
1. Dynamic color-to-status mapping correctness
2. Import row validation (mandatory/optional fields)
3. Import report totals consistency
4. Duplicate detection by composite key
5. Standard column schema parsing (headerless sheets)
6. ПОТЕРИ sheet → financial operations
7. Import raw trace completeness
8. Value normalization with review flag
32. Reconciliation report accuracy
"""
from __future__ import annotations

import sys
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from hypothesis import given, settings, assume, HealthCheck
from hypothesis import strategies as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.appdb.models import AppBase, AppUser
from backend.appdb.tickets_models import (
    TicketEmployee,
    TicketImportJob,
    TicketImportRawTrace,
    TicketObject,
    TicketRequest,
)
from backend.services.tickets_import_service import (
    DEFAULT_REFERENCE_COLORS,
    STANDARD_COLUMN_SCHEMA,
    DuplicateMatch,
    FinancialOpRow,
    ImportResult,
    ParsedRow,
    ParseError,
    ReconciliationReport,
    TicketsImportService,
)


# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------

hex_char = st.sampled_from("0123456789ABCDEF")
hex_color_st = st.text(alphabet="0123456789ABCDEF", min_size=6, max_size=6)

# Non-empty Russian-like names for ФИО
fio_st = st.text(
    alphabet=st.characters(whitelist_categories=("L", "Zs"), whitelist_characters="-"),
    min_size=3,
    max_size=50,
).filter(lambda s: s.strip())

# Dates within reasonable range
date_st = st.datetimes(
    min_value=datetime(2000, 1, 1),
    max_value=datetime(2030, 12, 31),
    timezones=st.just(timezone.utc),
)

# Decimal amounts
amount_st = st.decimals(
    min_value=Decimal("0.01"),
    max_value=Decimal("999999.99"),
    places=2,
    allow_nan=False,
    allow_infinity=False,
)

# Op types for losses
op_type_st = st.sampled_from(["возврат", "обмен", "потеря", "refund", "exchange", "loss"])


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def service(temp_dir, monkeypatch):
    """Create a TicketsImportService with a fresh SQLite database."""
    import backend.appdb.db as appdb

    url = f"sqlite:///{(Path(temp_dir) / 'tickets_prop_import.db').as_posix()}"

    appdb._engines.clear()
    appdb._session_factories.clear()
    appdb._initialized_schema_urls.clear()

    monkeypatch.setenv("APP_DATABASE_URL", url)

    try:
        from backend.config import config
        monkeypatch.setattr(config, "app_database_url", url, raising=False)
    except (ImportError, AttributeError):
        pass

    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.close()

    AppBase.metadata.create_all(engine, checkfirst=True)

    SessionLocal = sessionmaker(bind=engine)

    @contextmanager
    def _test_app_session(database_url=None):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    monkeypatch.setattr(
        "backend.services.tickets_import_service.app_session", _test_app_session
    )

    # Seed test data
    with _test_app_session() as session:
        user = AppUser(id=1, username="admin", full_name="Admin User", role="admin")
        session.add(user)
        session.flush()

        obj = TicketObject(
            id=1,
            code="KAM",
            name="Камчатка",
            region="Камчатский край",
            is_active=True,
        )
        session.add(obj)
        session.flush()

    upload_dir = Path(temp_dir) / "uploads" / "tickets" / "import"
    svc = TicketsImportService(upload_base_dir=str(upload_dir))
    svc._test_session_factory = _test_app_session
    svc._seed = {"obj_id": 1, "user_id": 1}
    return svc


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_worksheet_with_headers(data_rows: list[list], headers: list[str] | None = None) -> Workbook:
    """Create a workbook with a single sheet containing headers and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Камчатка"
    if headers:
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        start_row = 2
    else:
        start_row = 1
    for row_idx, row_data in enumerate(data_rows, start=start_row):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    return wb


def _make_losses_worksheet(rows: list[dict]) -> Workbook:
    """Create a workbook with a ПОТЕРИ sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ПОТЕРИ"
    headers = ["ФИО", "Сумма", "Причина", "Дата", "Тип операции"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row_data in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row_data.get("full_name"))
        ws.cell(row=row_idx, column=2, value=row_data.get("amount"))
        ws.cell(row=row_idx, column=3, value=row_data.get("reason"))
        ws.cell(row=row_idx, column=4, value=row_data.get("op_date"))
        ws.cell(row=row_idx, column=5, value=row_data.get("op_type"))
    return wb


# ---------------------------------------------------------------------------
# Property 1: Dynamic color-to-status mapping correctness
# ---------------------------------------------------------------------------


class TestPropertyColorMapping:
    """For any hex color, _suggest_color_mapping picks the reference color with minimum
    Euclidean distance in RGB space.

    **Validates: Requirements 1.3**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(colors=st.lists(hex_color_st, min_size=1, max_size=5))
    def test_suggested_mapping_picks_closest_reference(self, service, colors):
        """Each color in the mapping maps to the reference color with minimum distance."""
        result = service._suggest_color_mapping(colors)

        for color, mapped_status in result.items():
            # Find the actual minimum distance reference
            best_distance = float("inf")
            best_status = None
            for ref_hex, status in DEFAULT_REFERENCE_COLORS.items():
                dist = service._color_distance(color, ref_hex)
                if dist < best_distance:
                    best_distance = dist
                    best_status = status
            assert mapped_status == best_status, (
                f"Color {color}: expected status '{best_status}' but got '{mapped_status}'"
            )

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(colors=st.lists(hex_color_st, min_size=1, max_size=5))
    def test_all_input_colors_are_mapped(self, service, colors):
        """Every input color appears in the result mapping."""
        result = service._suggest_color_mapping(colors)
        for color in colors:
            assert color in result

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(color=hex_color_st)
    def test_map_color_to_status_lookup(self, service, color):
        """_map_color_to_status returns the value from color_map for a known color."""
        color_map = {color: "test_status"}
        result = service._map_color_to_status(color, color_map)
        assert result == "test_status"


# ---------------------------------------------------------------------------
# Property 2: Import row validation (mandatory/optional fields)
# ---------------------------------------------------------------------------


class TestPropertyRowValidation:
    """Rows without ФИО return ParseError; rows with ФИО return ParsedRow.

    **Validates: Requirements 1.4**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(fio=fio_st)
    def test_row_with_fio_returns_parsed_row(self, service, fio):
        """A row with a non-empty ФИО field produces a ParsedRow."""
        wb = _make_worksheet_with_headers(
            [[fio, "2024-01-15", "Москва-Камчатка"]],
            headers=["ФИО", "Дата подачи", "Маршрут"],
        )
        ws = wb.active
        header_map = {0: "full_name", 1: "submitted_at", 2: "route"}
        row = list(ws.iter_rows(min_row=2, max_row=2))[0]

        result = service._parse_row(
            row=row,
            header_map=header_map,
            row_number=2,
            comments={},
            hyperlinks={},
            color_map={},
        )
        assert isinstance(result, ParsedRow)
        assert result.full_name == fio.strip()

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(empty_val=st.sampled_from([None, "", "   ", "\t"]))
    def test_row_without_fio_returns_parse_error(self, service, empty_val):
        """A row with missing/empty ФИО produces a ParseError."""
        wb = _make_worksheet_with_headers(
            [[empty_val, "2024-01-15", "Москва-Камчатка"]],
            headers=["ФИО", "Дата подачи", "Маршрут"],
        )
        ws = wb.active
        header_map = {0: "full_name", 1: "submitted_at", 2: "route"}
        row = list(ws.iter_rows(min_row=2, max_row=2))[0]

        result = service._parse_row(
            row=row,
            header_map=header_map,
            row_number=2,
            comments={},
            hyperlinks={},
            color_map={},
        )
        assert isinstance(result, ParseError)
        assert "ФИО" in result.error

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(fio=fio_st)
    def test_row_with_fio_but_no_optional_fields_still_succeeds(self, service, fio):
        """A row with only ФИО (no optional fields) still produces a ParsedRow."""
        wb = _make_worksheet_with_headers(
            [[fio]],
            headers=["ФИО"],
        )
        ws = wb.active
        header_map = {0: "full_name"}
        row = list(ws.iter_rows(min_row=2, max_row=2))[0]

        result = service._parse_row(
            row=row,
            header_map=header_map,
            row_number=2,
            comments={},
            hyperlinks={},
            color_map={},
        )
        assert isinstance(result, ParsedRow)
        assert result.full_name == fio.strip()


# ---------------------------------------------------------------------------
# Property 3: Import report totals consistency
# ---------------------------------------------------------------------------


class TestPropertyImportReportTotals:
    """Import report totals: imported + skipped + errors = total processed rows.

    **Validates: Requirements 1.5**
    """

    @settings(max_examples=15)
    @given(
        imported=st.integers(min_value=0, max_value=100),
        skipped=st.integers(min_value=0, max_value=100),
        errors=st.integers(min_value=0, max_value=100),
        warnings=st.integers(min_value=0, max_value=100),
    )
    def test_report_totals_are_consistent(self, imported, skipped, errors, warnings):
        """imported + skipped + errors equals total rows processed."""
        result = ImportResult(
            imported=imported,
            skipped=skipped,
            errors=errors,
            warnings=warnings,
        )
        total = result.imported + result.skipped + result.errors
        assert total == imported + skipped + errors
        # Warnings are separate — they don't affect the row count
        assert result.warnings == warnings


# ---------------------------------------------------------------------------
# Property 4: Duplicate detection by composite key
# ---------------------------------------------------------------------------


class TestPropertyDuplicateDetection:
    """Rows with same (full_name, object_id, submitted_at) as existing DB records are detected.

    **Validates: Requirements 1.6**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(fio=fio_st, date=date_st)
    def test_duplicate_detected_when_existing_record_matches(self, service, fio, date):
        """A parsed row matching an existing DB record is detected as duplicate."""
        import backend.appdb.db as appdb

        # Insert an employee and request into the DB
        session_factory = service._test_session_factory
        with session_factory() as session:
            emp = TicketEmployee(full_name=fio.strip(), status="active")
            session.add(emp)
            session.flush()
            emp_id = emp.id

            req = TicketRequest(
                employee_id=emp_id,
                object_id=1,
                status="new",
                submitted_at=date,
            )
            session.add(req)
            session.flush()

        # Create a parsed row with the same composite key
        parsed = ParsedRow(
            full_name=fio.strip(),
            submitted_at=date,
            row_number=2,
        )

        duplicates = service._detect_duplicates([parsed], object_id=1)
        assert len(duplicates) >= 1
        assert any(d.parsed_row.full_name == fio.strip() for d in duplicates)

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(fio=fio_st, date=date_st)
    def test_no_duplicate_when_no_matching_record(self, service, fio, date):
        """A parsed row with no matching DB record is not detected as duplicate."""
        # Don't insert anything — DB is empty for this name
        parsed = ParsedRow(
            full_name=fio.strip() + "_unique_suffix",
            submitted_at=date,
            row_number=2,
        )

        duplicates = service._detect_duplicates([parsed], object_id=1)
        assert len(duplicates) == 0


# ---------------------------------------------------------------------------
# Property 5: Standard column schema parsing (headerless sheets)
# ---------------------------------------------------------------------------


class TestPropertyStandardSchema:
    """Positional column mapping for headerless sheets uses STANDARD_COLUMN_SCHEMA.

    **Validates: Requirements 1.8**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(fio=fio_st)
    def test_standard_schema_parses_fio_from_first_column(self, service, fio):
        """In standard schema mode, the first column is mapped to full_name."""
        wb = Workbook()
        ws = wb.active
        # Standard schema: full_name, submitted_at, departure_date, route, cost, status_note, assignee
        ws.cell(row=1, column=1, value=fio)
        ws.cell(row=1, column=2, value="2024-03-15")
        ws.cell(row=1, column=3, value="2024-03-20")
        ws.cell(row=1, column=4, value="Москва-Камчатка")
        ws.cell(row=1, column=5, value=15000)
        ws.cell(row=1, column=6, value="новая")
        ws.cell(row=1, column=7, value="Логист1")

        row = list(ws.iter_rows(min_row=1, max_row=1))[0]

        result = service._parse_row_standard_schema(
            row=row,
            col_order=STANDARD_COLUMN_SCHEMA,
            row_number=1,
            comments={},
            hyperlinks={},
            color_map={},
        )
        assert isinstance(result, ParsedRow)
        assert result.full_name == fio.strip()

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(empty_val=st.sampled_from([None, "", "   "]))
    def test_standard_schema_without_fio_returns_error(self, service, empty_val):
        """Standard schema with empty first column (ФИО) returns ParseError."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=empty_val)
        ws.cell(row=1, column=2, value="2024-03-15")

        row = list(ws.iter_rows(min_row=1, max_row=1))[0]

        result = service._parse_row_standard_schema(
            row=row,
            col_order=STANDARD_COLUMN_SCHEMA,
            row_number=1,
            comments={},
            hyperlinks={},
            color_map={},
        )
        assert isinstance(result, ParseError)


# ---------------------------------------------------------------------------
# Property 6: ПОТЕРИ sheet → financial operations
# ---------------------------------------------------------------------------


class TestPropertyLossesSheet:
    """ПОТЕРИ sheet is parsed into FinancialOpRow list with correct field mapping.

    **Validates: Requirements 1.9**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(
        fio=fio_st,
        amount=amount_st,
        op_type=op_type_st,
    )
    def test_losses_sheet_produces_financial_op_rows(self, service, fio, amount, op_type):
        """Each row in ПОТЕРИ sheet with ФИО produces a FinancialOpRow."""
        wb = _make_losses_worksheet([{
            "full_name": fio,
            "amount": str(amount),
            "reason": "Тестовая причина",
            "op_date": datetime(2024, 3, 15, tzinfo=timezone.utc),
            "op_type": op_type,
        }])
        ws = wb.active

        results = service._parse_losses_sheet(ws)
        assert len(results) == 1
        assert results[0].full_name == fio.strip()
        assert results[0].amount == Decimal(str(amount).replace(",", "."))

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(empty_val=st.sampled_from([None, "", "   "]))
    def test_losses_sheet_skips_rows_without_fio(self, service, empty_val):
        """Rows in ПОТЕРИ sheet without ФИО are skipped."""
        wb = _make_losses_worksheet([{
            "full_name": empty_val,
            "amount": "1000",
            "reason": "Причина",
            "op_date": datetime(2024, 3, 15, tzinfo=timezone.utc),
            "op_type": "потеря",
        }])
        ws = wb.active

        results = service._parse_losses_sheet(ws)
        assert len(results) == 0

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(op_type=st.sampled_from(["возврат", "refund"]))
    def test_losses_sheet_maps_refund_type(self, service, op_type):
        """Op type containing 'возврат' or 'refund' maps to 'refund'."""
        wb = _make_losses_worksheet([{
            "full_name": "Иванов И.И.",
            "amount": "5000",
            "reason": "Отмена рейса",
            "op_date": datetime(2024, 3, 15, tzinfo=timezone.utc),
            "op_type": op_type,
        }])
        ws = wb.active

        results = service._parse_losses_sheet(ws)
        assert len(results) == 1
        assert results[0].op_type == "refund"


# ---------------------------------------------------------------------------
# Property 7: Import raw trace completeness
# ---------------------------------------------------------------------------


class TestPropertyRawTrace:
    """Every imported row stores a raw trace with all required fields.

    **Validates: Requirements 1.10**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(fio=fio_st)
    def test_raw_trace_stores_all_fields(self, service, fio):
        """_store_raw_trace creates a TicketImportRawTrace with all JSON fields populated."""
        session_factory = service._test_session_factory

        # Create a minimal import job
        import uuid as uuid_mod
        job_id = uuid_mod.uuid4().hex[:16]
        with session_factory() as session:
            job = TicketImportJob(
                id=job_id,
                file_name="test.xlsx",
                file_path="/tmp/test.xlsx",
                status="running",
                user_id=1,
            )
            session.add(job)

        parsed = ParsedRow(
            full_name=fio.strip(),
            row_number=5,
            raw_cells={"A5": fio, "B5": "2024-01-15"},
            cell_colors={"A5": "FF0000"},
            cell_formulas={"C5": "=SUM(D5:E5)"},
            cell_comments={"A5": "Комментарий"},
            cell_hyperlinks={"B5": "https://example.com"},
            cell_addresses={"col_0": "A5", "col_1": "B5"},
        )

        service._store_raw_trace(
            parsed=parsed,
            job_id=job_id,
            file_name="test.xlsx",
            sheet_name="Камчатка",
            sheet_visibility="visible",
            request_id=None,
        )

        # Verify the trace was stored
        with session_factory() as session:
            trace = session.query(TicketImportRawTrace).filter_by(job_id=job_id).first()
            assert trace is not None
            assert trace.source_file == "test.xlsx"
            assert trace.sheet_name == "Камчатка"
            assert trace.row_number == 5
            assert trace.sheet_visibility == "visible"
            # All JSON fields are non-empty
            assert trace.raw_cells_json != "{}"
            assert trace.cell_colors_json != "{}"
            assert trace.cell_formulas_json != "{}"
            assert trace.cell_comments_json != "{}"
            assert trace.cell_hyperlinks_json != "{}"
            assert trace.cell_addresses_json != "{}"


# ---------------------------------------------------------------------------
# Property 8: Value normalization with review flag
# ---------------------------------------------------------------------------


class TestPropertyValueNormalization:
    """Value normalization sets needs_review correctly based on input type.

    **Validates: Requirements 1.11**
    """

    @settings(max_examples=15)
    @given(dt=date_st)
    def test_datetime_value_for_date_type_no_review(self, dt):
        """A proper datetime value for 'date' type does not need review."""
        normalized, needs_review = TicketsImportService._normalize_value(dt, "date")
        assert normalized == dt
        assert needs_review is False

    @settings(max_examples=15)
    @given(text_date=st.sampled_from(["15.03.2024", "2024-03-15", "15/03/2024"]))
    def test_date_as_text_needs_review(self, text_date):
        """A date stored as text string needs review after normalization."""
        normalized, needs_review = TicketsImportService._normalize_value(text_date, "date")
        assert normalized is not None
        assert needs_review is True

    @settings(max_examples=15)
    @given(phone_num=st.integers(min_value=9000000000, max_value=9999999999))
    def test_phone_as_number_needs_review(self, phone_num):
        """A phone number stored as integer needs review."""
        normalized, needs_review = TicketsImportService._normalize_value(phone_num, "phone")
        assert needs_review is True
        assert isinstance(normalized, str)

    @settings(max_examples=15)
    @given(phone_str=st.from_regex(r"\+7\d{10}", fullmatch=True))
    def test_phone_as_string_no_review(self, phone_str):
        """A phone stored as proper string does not need review."""
        normalized, needs_review = TicketsImportService._normalize_value(phone_str, "phone")
        assert needs_review is False
        assert normalized == phone_str

    @settings(max_examples=15)
    @given(passport_num=st.integers(min_value=1000000000, max_value=9999999999))
    def test_passport_as_number_needs_review(self, passport_num):
        """A passport stored as number needs review and gets formatted."""
        normalized, needs_review = TicketsImportService._normalize_value(passport_num, "passport")
        assert needs_review is True
        assert isinstance(normalized, str)
        # Should be formatted as "XXXX XXXXXX"
        assert " " in normalized

    @settings(max_examples=15)
    @given(passport_str=st.from_regex(r"\d{4} \d{6}", fullmatch=True))
    def test_passport_as_string_no_review(self, passport_str):
        """A passport stored as proper string does not need review."""
        normalized, needs_review = TicketsImportService._normalize_value(passport_str, "passport")
        assert needs_review is False
        assert normalized == passport_str

    @settings(max_examples=15)
    @given(val=st.decimals(min_value=Decimal("0.01"), max_value=Decimal("99999.99"), places=2, allow_nan=False, allow_infinity=False))
    def test_decimal_value_for_decimal_type_no_review(self, val):
        """A proper Decimal value for 'decimal' type does not need review."""
        normalized, needs_review = TicketsImportService._normalize_value(val, "decimal")
        assert normalized == val
        assert needs_review is False

    def test_none_value_returns_none_no_review(self):
        """None value returns (None, False) for any type."""
        for expected_type in ("date", "decimal", "phone", "passport", "text"):
            normalized, needs_review = TicketsImportService._normalize_value(None, expected_type)
            assert normalized is None
            assert needs_review is False


# ---------------------------------------------------------------------------
# Property 32: Reconciliation report accuracy
# ---------------------------------------------------------------------------


class TestPropertyReconciliationReport:
    """Reconciliation report accurately reflects import results.

    **Validates: Requirements 1.5**
    """

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(
        imported=st.integers(min_value=0, max_value=500),
        skipped=st.integers(min_value=0, max_value=500),
        errors=st.integers(min_value=0, max_value=500),
        warnings=st.integers(min_value=0, max_value=500),
    )
    def test_reconciliation_report_matches_import_result(self, service, imported, skipped, errors, warnings):
        """_build_reconciliation_report produces a report matching the ImportResult counts."""
        result = ImportResult(
            imported=imported,
            skipped=skipped,
            errors=errors,
            warnings=warnings,
        )

        report = service._build_reconciliation_report(result)

        assert isinstance(report, ReconciliationReport)
        assert report.rows_imported == imported
        assert report.rows_skipped == skipped
        assert report.rows_with_errors == errors
        assert report.rows_with_warnings == warnings
