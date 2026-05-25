"""Property-based tests for loss report and notifications (task 8.6).

Properties tested:
- Property 28: Loss report filtering and totals (Decimal precision)
- Property 29: Loss report export round-trip
- Property 31: Notification SLA rules correctness

**Validates: Requirements 11.2, 11.3, 11.4**
"""
from __future__ import annotations

import sys
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from hypothesis import given, settings, assume, HealthCheck
from hypothesis import strategies as st

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = PROJECT_ROOT / "WEB-itinvent"

if str(WEB_ROOT) not in sys.path:
    sys.path.insert(0, str(WEB_ROOT))

from backend.appdb.models import AppBase, AppUser
from backend.appdb.tickets_models import (
    TicketEmployee,
    TicketFinancialOp,
    TicketNotificationRule,
    TicketObject,
    TicketRequest,
)
from backend.services.tickets_service import (
    FinOpFilters,
    Pagination,
    TicketsService,
)
from backend.services.tickets_notification_service import (
    RULE_DEPARTURE_SOON,
    RULE_MISSING_DATA_STALE,
    RULE_STUCK_REQUEST,
    RULE_NEW_LOSS,
    TERMINAL_STATUSES,
    TicketsNotificationService,
)


# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------

st_amount = st.decimals(
    min_value=Decimal("0.01"), max_value=Decimal("999999.99"),
    places=2, allow_nan=False, allow_infinity=False,
)

st_op_type = st.sampled_from(["refund", "exchange", "loss"])

# Non-terminal statuses for notification tests
NON_TERMINAL_STATUSES = [
    "new", "data_check", "missing_data", "ready_to_buy",
    "in_progress", "exchange_needed", "refund",
]

st_non_terminal_status = st.sampled_from(NON_TERMINAL_STATUSES)
st_terminal_status = st.sampled_from(list(TERMINAL_STATUSES))


@st.composite
def st_financial_op_for_report(draw):
    """Strategy that generates a financial operation with a date in 2024."""
    return {
        "op_type": draw(st_op_type),
        "amount": str(draw(st_amount)),
        "op_date": draw(
            st.dates(
                min_value=datetime(2024, 1, 1).date(),
                max_value=datetime(2024, 12, 31).date(),
            )
        ).isoformat(),
    }


@st.composite
def st_financial_ops_batch(draw):
    """Strategy that generates a batch of 3-12 financial operations."""
    n = draw(st.integers(min_value=3, max_value=12))
    return [draw(st_financial_op_for_report()) for _ in range(n)]


@st.composite
def st_days_threshold(draw):
    """Strategy for days threshold (1-30)."""
    return draw(st.integers(min_value=1, max_value=30))


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _sqlite_url(temp_dir: str) -> str:
    return f"sqlite:///{(Path(temp_dir) / 'tickets_prop_reports_notif.db').as_posix()}"


@pytest.fixture
def db_env(temp_dir, monkeypatch):
    """Create a fresh SQLite database with all ticket tables and seed data."""
    from sqlalchemy import create_engine, event
    from sqlalchemy.orm import sessionmaker

    import backend.appdb.db as appdb

    url = _sqlite_url(temp_dir)

    appdb._engines.clear()
    appdb._session_factories.clear()
    appdb._initialized_schema_urls.clear()

    engine = create_engine(
        url,
        execution_options={"schema_translate_map": {"app": None, "system": None}},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, connection_record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.close()

    AppBase.metadata.create_all(engine, checkfirst=True)

    SessionLocal = sessionmaker(bind=engine)

    @contextmanager
    def _test_app_session(database_url=None):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    monkeypatch.setattr("backend.services.tickets_service.app_session", _test_app_session)
    monkeypatch.setattr(
        "backend.services.tickets_notification_service.app_session", _test_app_session
    )

    # Seed base data
    now = datetime.now(timezone.utc)
    with _test_app_session() as session:
        user = AppUser(id=1, username="admin", full_name="Admin User", role="admin")
        session.add(user)
        session.flush()

        emp = TicketEmployee(
            full_name="Тестов Тест Тестович",
            phone="+79001234567",
            status="active",
        )
        session.add(emp)
        session.flush()
        emp_id = emp.id

        obj = TicketObject(
            code="KAM",
            name="Камчатка",
            region="Дальний Восток",
            is_active=True,
        )
        session.add(obj)
        session.flush()
        obj_id = obj.id

        # Seed notification rules (all enabled)
        rules = [
            TicketNotificationRule(
                rule_type=RULE_DEPARTURE_SOON,
                is_enabled=True,
                threshold_days=3,
                notify_roles="admin,operator",
            ),
            TicketNotificationRule(
                rule_type=RULE_MISSING_DATA_STALE,
                is_enabled=True,
                threshold_days=3,
                notify_roles="admin,operator",
            ),
            TicketNotificationRule(
                rule_type=RULE_STUCK_REQUEST,
                is_enabled=True,
                threshold_days=5,
                notify_roles="admin,operator",
            ),
            TicketNotificationRule(
                rule_type=RULE_NEW_LOSS,
                is_enabled=True,
                threshold_days=None,
                notify_roles="admin,operator",
            ),
        ]
        session.add_all(rules)
        session.flush()

    return {
        "session_factory": _test_app_session,
        "emp_id": emp_id,
        "obj_id": obj_id,
        "user_id": 1,
    }


@pytest.fixture
def service(db_env):
    """Create a TicketsService instance."""
    return TicketsService()


@pytest.fixture
def notif_service(db_env):
    """Create a TicketsNotificationService instance."""
    return TicketsNotificationService()


# ---------------------------------------------------------------------------
# Property 28: Loss report filtering and totals (Decimal precision)
#
# For any set of financial operations, total_losses = sum of amounts where
# op_type="loss", total_refunds = sum where op_type="refund",
# balance = losses - refunds. All using Decimal precision.
#
# **Validates: Requirements 11.2, 11.3**
# ---------------------------------------------------------------------------


class TestProperty28LossReportTotals:
    """Property 28: Loss report filtering and totals (Decimal precision)."""

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(ops=st_financial_ops_batch())
    def test_totals_match_decimal_sums(self, service, ops):
        """For any set of financial operations, the loss report totals
        (total_losses, total_refunds, balance) match exact Decimal sums.

        **Validates: Requirements 11.2, 11.3**
        """
        # Create all operations
        created_ids = []
        for op_data in ops:
            result = service.create_financial_op(op_data, user_id=1)
            created_ids.append(result["id"])

        # Compute expected totals using Decimal arithmetic
        expected_losses = sum(
            Decimal(op["amount"]) for op in ops if op["op_type"] == "loss"
        )
        expected_refunds = sum(
            Decimal(op["amount"]) for op in ops if op["op_type"] == "refund"
        )
        expected_balance = expected_losses - expected_refunds

        # Get loss report covering all of 2024
        report = service.get_losses_report(
            filters={"date_from": "2024-01-01", "date_to": "2025-01-01"},
            pagination=Pagination(page=1, page_size=50),
        )

        # Verify totals use Decimal precision (no float rounding)
        actual_losses = Decimal(report["totals"]["total_losses"])
        actual_refunds = Decimal(report["totals"]["total_refunds"])
        actual_balance = Decimal(report["totals"]["balance"])

        assert actual_losses == expected_losses, (
            f"total_losses: expected {expected_losses}, got {actual_losses}"
        )
        assert actual_refunds == expected_refunds, (
            f"total_refunds: expected {expected_refunds}, got {actual_refunds}"
        )
        assert actual_balance == expected_balance, (
            f"balance: expected {expected_balance}, got {actual_balance}"
        )

        # Clean up
        for op_id in created_ids:
            service.delete_financial_op(op_id, user_id=1)

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(ops=st_financial_ops_batch())
    def test_filter_by_op_type_returns_correct_subset(self, service, ops):
        """Filtering the loss report by op_type returns only matching items
        and correct totals for that subset.

        **Validates: Requirements 11.2, 11.3**
        """
        created_ids = []
        for op_data in ops:
            result = service.create_financial_op(op_data, user_id=1)
            created_ids.append(result["id"])

        # Filter by "loss" type only
        report = service.get_losses_report(
            filters={"date_from": "2024-01-01", "date_to": "2025-01-01", "op_type": "loss"},
            pagination=Pagination(page=1, page_size=50),
        )

        # All returned items must be of type "loss"
        for item in report["items"]:
            assert item["op_type"] == "loss"

        # Count must match expected
        expected_count = sum(1 for op in ops if op["op_type"] == "loss")
        assert report["pagination"]["total"] == expected_count

        # Clean up
        for op_id in created_ids:
            service.delete_financial_op(op_id, user_id=1)


# ---------------------------------------------------------------------------
# Property 29: Loss report export round-trip
#
# export_losses_xlsx() produces valid .xlsx bytes that can be loaded by
# openpyxl, with correct number of data rows matching the filter.
#
# **Validates: Requirements 11.4**
# ---------------------------------------------------------------------------


class TestProperty29ExportRoundTrip:
    """Property 29: Loss report export round-trip."""

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(ops=st_financial_ops_batch())
    def test_export_produces_valid_xlsx_with_correct_row_count(self, service, ops):
        """export_losses_xlsx() produces valid .xlsx bytes loadable by openpyxl,
        with data rows matching the number of filtered records.

        **Validates: Requirements 11.4**
        """
        from openpyxl import load_workbook

        # Create operations
        created_ids = []
        for op_data in ops:
            result = service.create_financial_op(op_data, user_id=1)
            created_ids.append(result["id"])

        # Export covering all of 2024
        xlsx_bytes = service.export_losses_xlsx(
            filters={"date_from": "2024-01-01", "date_to": "2025-01-01"}
        )

        # Verify it's valid xlsx bytes
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

        # Load with openpyxl
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active

        # First row is header
        rows = list(ws.iter_rows(min_row=1))
        header_row = rows[0]
        data_rows = rows[1:]

        # Header should have 7 columns
        assert len(header_row) == 7

        # Data rows should match the number of created operations
        assert len(data_rows) == len(ops), (
            f"Expected {len(ops)} data rows, got {len(data_rows)}"
        )

        # Clean up
        for op_id in created_ids:
            service.delete_financial_op(op_id, user_id=1)

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(ops=st_financial_ops_batch())
    def test_export_filtered_by_op_type_matches_count(self, service, ops):
        """When exporting with op_type filter, the xlsx contains only matching rows.

        **Validates: Requirements 11.4**
        """
        from openpyxl import load_workbook

        # Create operations
        created_ids = []
        for op_data in ops:
            result = service.create_financial_op(op_data, user_id=1)
            created_ids.append(result["id"])

        # Count expected losses
        expected_loss_count = sum(1 for op in ops if op["op_type"] == "loss")

        # Export only losses
        xlsx_bytes = service.export_losses_xlsx(
            filters={"date_from": "2024-01-01", "date_to": "2025-01-01", "op_type": "loss"}
        )

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2))  # skip header

        assert len(rows) == expected_loss_count, (
            f"Expected {expected_loss_count} loss rows, got {len(rows)}"
        )

        # Clean up
        for op_id in created_ids:
            service.delete_financial_op(op_id, user_id=1)


# ---------------------------------------------------------------------------
# Property 31: Notification SLA rules correctness
#
# check_departure_soon() only returns requests where departure_date is within
# threshold AND status is not terminal.
# check_stuck_requests() only returns requests where updated_at is older than
# threshold AND status is not terminal.
# After dismissing a notification for a user, it doesn't appear in
# get_all_pending() for that user but still appears for other users.
#
# **Validates: Requirements 11.2, 11.3**
# ---------------------------------------------------------------------------


class TestProperty31NotificationSLA:
    """Property 31: Notification SLA rules correctness."""

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(
        days_threshold=st_days_threshold(),
        status=st_non_terminal_status,
        days_until_departure=st.integers(min_value=1, max_value=40),
    )
    def test_departure_soon_only_within_threshold_non_terminal(
        self, notif_service, db_env, days_threshold, status, days_until_departure
    ):
        """check_departure_soon() returns a request only if departure_date is
        within threshold AND status is not terminal.

        Uses days_until_departure >= 1 to avoid boundary timing issues where
        departure_date == now can race with the service's own now() call.

        **Validates: Requirements 11.2, 11.3**
        """
        now = datetime.now(timezone.utc)
        session_factory = db_env["session_factory"]

        with session_factory() as session:
            req = TicketRequest(
                employee_id=db_env["emp_id"],
                object_id=db_env["obj_id"],
                status=status,
                departure_date=now + timedelta(days=days_until_departure),
                created_at=now,
                updated_at=now,
            )
            session.add(req)
            session.flush()
            req_id = req.id

        notifications = notif_service.check_departure_soon(days_threshold=days_threshold)
        found_ids = {n.request_id for n in notifications}

        # Should be found only if departure is within threshold (and >= now)
        should_be_found = days_until_departure <= days_threshold
        if should_be_found:
            assert req_id in found_ids, (
                f"Request with departure in {days_until_departure} days should be flagged "
                f"(threshold={days_threshold})"
            )
        else:
            assert req_id not in found_ids, (
                f"Request with departure in {days_until_departure} days should NOT be flagged "
                f"(threshold={days_threshold})"
            )

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(
        days_threshold=st_days_threshold(),
        status=st_terminal_status,
    )
    def test_departure_soon_excludes_terminal_statuses(
        self, notif_service, db_env, days_threshold, status
    ):
        """check_departure_soon() never returns requests in terminal statuses,
        regardless of departure date.

        **Validates: Requirements 11.2, 11.3**
        """
        now = datetime.now(timezone.utc)
        session_factory = db_env["session_factory"]

        with session_factory() as session:
            req = TicketRequest(
                employee_id=db_env["emp_id"],
                object_id=db_env["obj_id"],
                status=status,
                departure_date=now + timedelta(days=1),  # within any threshold
                created_at=now,
                updated_at=now,
            )
            session.add(req)
            session.flush()
            req_id = req.id

        notifications = notif_service.check_departure_soon(days_threshold=days_threshold)
        found_ids = {n.request_id for n in notifications}

        assert req_id not in found_ids, (
            f"Request in terminal status '{status}' should never be flagged"
        )

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(
        days_threshold=st_days_threshold(),
        status=st_non_terminal_status,
        days_since_update=st.integers(min_value=0, max_value=40),
    )
    def test_stuck_requests_only_older_than_threshold_non_terminal(
        self, notif_service, db_env, days_threshold, status, days_since_update
    ):
        """check_stuck_requests() returns a request only if updated_at is older
        than threshold AND status is not terminal.

        **Validates: Requirements 11.2, 11.3**
        """
        now = datetime.now(timezone.utc)
        session_factory = db_env["session_factory"]

        with session_factory() as session:
            req = TicketRequest(
                employee_id=db_env["emp_id"],
                object_id=db_env["obj_id"],
                status=status,
                departure_date=now + timedelta(days=30),
                created_at=now - timedelta(days=days_since_update + 5),
                updated_at=now - timedelta(days=days_since_update),
            )
            session.add(req)
            session.flush()
            req_id = req.id

        notifications = notif_service.check_stuck_requests(days_threshold=days_threshold)
        found_ids = {n.request_id for n in notifications}

        # Should be found only if days_since_update >= threshold
        should_be_found = days_since_update >= days_threshold
        if should_be_found:
            assert req_id in found_ids, (
                f"Request updated {days_since_update} days ago should be flagged "
                f"(threshold={days_threshold})"
            )
        else:
            assert req_id not in found_ids, (
                f"Request updated {days_since_update} days ago should NOT be flagged "
                f"(threshold={days_threshold})"
            )

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(
        days_threshold=st_days_threshold(),
        status=st_terminal_status,
    )
    def test_stuck_requests_excludes_terminal_statuses(
        self, notif_service, db_env, days_threshold, status
    ):
        """check_stuck_requests() never returns requests in terminal statuses.

        **Validates: Requirements 11.2, 11.3**
        """
        now = datetime.now(timezone.utc)
        session_factory = db_env["session_factory"]

        with session_factory() as session:
            req = TicketRequest(
                employee_id=db_env["emp_id"],
                object_id=db_env["obj_id"],
                status=status,
                departure_date=now + timedelta(days=30),
                created_at=now - timedelta(days=30),
                updated_at=now - timedelta(days=30),  # very old
            )
            session.add(req)
            session.flush()
            req_id = req.id

        notifications = notif_service.check_stuck_requests(days_threshold=days_threshold)
        found_ids = {n.request_id for n in notifications}

        assert req_id not in found_ids, (
            f"Request in terminal status '{status}' should never be flagged as stuck"
        )

    @settings(max_examples=15, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    @given(
        user1_id=st.integers(min_value=100, max_value=200),
        user2_id=st.integers(min_value=201, max_value=300),
    )
    def test_dismiss_notification_per_user_isolation(
        self, notif_service, db_env, user1_id, user2_id
    ):
        """After dismissing a notification for user1, it doesn't appear in
        get_all_pending() for user1 but still appears for user2.

        **Validates: Requirements 11.2, 11.3**
        """
        now = datetime.now(timezone.utc)
        session_factory = db_env["session_factory"]

        # Create a request that triggers departure_soon
        with session_factory() as session:
            req = TicketRequest(
                employee_id=db_env["emp_id"],
                object_id=db_env["obj_id"],
                status="in_progress",
                departure_date=now + timedelta(days=1),
                created_at=now,
                updated_at=now,
            )
            session.add(req)
            session.flush()
            req_id = req.id

        notification_id = f"departure_soon_{req_id}"

        user1 = {"id": user1_id, "role": "admin"}
        user2 = {"id": user2_id, "role": "admin"}

        # Both users should see the notification initially
        pending_user1 = notif_service.get_all_pending(user1)
        pending_user2 = notif_service.get_all_pending(user2)
        assert any(n.id == notification_id for n in pending_user1)
        assert any(n.id == notification_id for n in pending_user2)

        # User1 dismisses the notification
        notif_service.dismiss_notification(notification_id, user1)

        # User1 should no longer see it
        pending_user1_after = notif_service.get_all_pending(user1)
        assert not any(n.id == notification_id for n in pending_user1_after), (
            f"Dismissed notification should not appear for user1"
        )

        # User2 should still see it
        pending_user2_after = notif_service.get_all_pending(user2)
        assert any(n.id == notification_id for n in pending_user2_after), (
            f"Notification should still appear for user2 after user1 dismissed it"
        )
